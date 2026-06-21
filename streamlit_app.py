import csv
import importlib
import json
import random
import re
import threading
import time
from copy import deepcopy
from dataclasses import MISSING, asdict, dataclass, fields
from datetime import datetime
from functools import lru_cache
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from html import escape
from io import BytesIO, StringIO
from pathlib import Path

import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.amazon_scraper import DEFAULT_TEST_URL, fetch_best_sellers
from src.category_mapping import (
    AIR_FRYERS_PATH,
    AIR_FRYERS_URL,
    category_url_matches_path,
    clean_category_entries,
    is_legacy_home_path,
)
from src import category_selection as category_selection_module
from src.chrome_cdp import (
    SELLERSPRITE_INCOMPLETE_MESSAGE,
    chrome_debugger_available,
    discover_bestseller_category_links,
    is_rank_category_url,
    refresh_sellersprite_cache_pages,
)
from src.collection_quality import (
    completed_collection_page_count,
    evaluate_sellersprite_collection_quality,
    is_collection_quality_warning,
    is_sellersprite_load_failure,
)
from src.collection_retry import CollectionRetryQueue
from src.collection_status import build_collection_total_status_text
from src.sellersprite_parser import (
    load_cached_sellersprite_products,
    sellersprite_product_hydrated,
)
from src.result_pagination import (
    PAGE_SIZE_OPTIONS,
    clamp_page,
    normalize_page_size,
    page_count,
    page_range_label,
    page_start_index,
    page_slice,
)
from src.product_annotations import (
    load_favorites,
    load_notes,
    remove_favorite,
    save_favorites,
    save_notes,
    set_product_note,
    upsert_favorite,
)

category_selection_module = importlib.reload(category_selection_module)
category_path_selected = category_selection_module.category_path_selected
toggle_compact_category_selection = category_selection_module.toggle_compact_category_selection


OUTPUT_DIR = Path(__file__).resolve().parent / "outputs"
SELLERSPRITE_DOM_CACHE = OUTPUT_DIR / "sellersprite_dom.txt"
SELLERSPRITE_IMAGE_CACHE = OUTPUT_DIR / "sellersprite_images.json"
SELLERSPRITE_META_CACHE = OUTPUT_DIR / "sellersprite_cache_meta.json"
LEARNED_CATEGORY_LINKS = OUTPUT_DIR / "category_links_learned.json"
RAW_PRODUCTS_CACHE = OUTPUT_DIR / "last_raw_products.json"
RAW_PRODUCTS_HISTORY_DIR = OUTPUT_DIR / "raw_products"
RAW_PRODUCTS_HISTORY_INDEX = RAW_PRODUCTS_HISTORY_DIR / "index.json"
RAW_PRODUCTS_HISTORY_LIMIT = 5
LAST_COLLECTION_REPORT = OUTPUT_DIR / "last_collection_report.json"
STOP_COLLECTION_FLAG = OUTPUT_DIR / "stop_collection.flag"
COLLECTION_RUNNING_FLAG = OUTPUT_DIR / "collection_running.flag"
FAVORITES_PATH = OUTPUT_DIR / "favorites.json"
PRODUCT_NOTES_PATH = OUTPUT_DIR / "product_notes.json"
STOP_COLLECTION_PORT = 8765
SELLERSPRITE_PLUGIN_FIELDS_TEXT = "价格、评分、评分数、排名、销量、销售额、FBA费用、毛利率、变体数、卖家数、包装信息"
SELLERSPRITE_EXPECTED_PRODUCTS_PER_PAGE = 50
SELLERSPRITE_MIN_PRODUCTS_PER_PAGE = 45
SELLERSPRITE_MIN_PRODUCTS_TWO_PAGES = 95
SELLERSPRITE_CATEGORY_RETRY_LIMIT = 1
EMPTY_NEW_RELEASES_MESSAGE = "Amazon 明确显示该类目暂无热门新品。"

SELLERSPRITE_EXPORT_COLUMNS = [
    ("image_preview_formula", "主图"),
    ("rank", "#"),
    ("title", "产品信息"),
    ("asin", "ASIN"),
    ("brand", "品牌"),
    ("seller_name", "卖家"),
    ("prime_fba", "配送"),
    ("seller_count", "卖家数"),
    ("bsr_rank", "大类BSR"),
    ("bsr_category", "大类目"),
    ("sub_rank", "子类排名"),
    ("sub_category", "子类目"),
    ("monthly_bought", "近30天销量(父体)"),
    ("child_monthly_sales_label", "近30天销量(子体)"),
    ("sales_amount", "销售额"),
    ("variant_count", "变体数"),
    ("price", "价格"),
    ("review_count", "评分数"),
    ("rating", "评分"),
    ("fba_fee", "FBA费用"),
    ("margin_rate", "毛利率"),
    ("launched_at", "上架时间"),
    ("package_weight_lb", "包装重量(lb)"),
    ("package_dimensions", "包装尺寸"),
    ("review_status", "审核状态"),
    ("note", "备注"),
    ("potential_score", "AI分"),
    ("potential_level", "评级"),
    ("reason", "推荐理由"),
    ("risk_tags", "风险词"),
    ("category_path", "浏览同类目"),
    ("amazon_url", "Amazon链接"),
    ("image_url", "主图链接"),
    ("scraped_at", "抓取时间"),
    ("status", "状态"),
    ("error", "错误"),
]

SELLERSPRITE_TABLE_COLUMNS = [
    (field, header)
    for field, header in SELLERSPRITE_EXPORT_COLUMNS
    if field != "image_preview_formula"
]


st.set_page_config(
    page_title="Amazon Selection Agent",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded",
)


CATEGORIES = {
    "Appliances": {
        "zh": "家电",
        "count": 80335,
        "children": {
            "Appliance Warranties": {"zh": "家电保修", "count": 2, "children": {}},
            "Dishwashers": {
                "zh": "洗碗机",
                "count": 1314,
                "children": {
                    "Built-In Dishwashers": {"zh": "嵌入式洗碗机", "count": 437},
                    "Countertop Dishwashers": {"zh": "台式洗碗机", "count": 542},
                    "Portable Dishwashers": {"zh": "便携式洗碗机", "count": 103},
                },
            },
            "Small Appliances": {
                "zh": "小家电",
                "count": 18640,
                "children": {
                    "Humidifiers": {"zh": "加湿器", "count": 3182},
                    "Portable Fans": {"zh": "便携风扇", "count": 2710},
                    "Electric Kettles": {"zh": "电热水壶", "count": 1788},
                    "Vacuums": {"zh": "吸尘器", "count": 4960},
                },
            },
        },
    },
    "Home & Kitchen": {
        "zh": "家居厨房",
        "count": 245980,
        "children": {},
    },
    "Pet Supplies": {
        "zh": "宠物用品",
        "count": 115420,
        "children": {},
    },
}

CATEGORY_BESTSELLER_URLS = {
    "Appliances": "https://www.amazon.com/gp/bestsellers/appliances/ref=zg_bs_nav_appliances_0",
    "Home & Kitchen": "https://www.amazon.com/gp/bestsellers/home-garden/ref=zg_bs_nav_home-garden_0",
    AIR_FRYERS_PATH: AIR_FRYERS_URL,
    "Pet Supplies": "https://www.amazon.com/gp/bestsellers/pet-supplies/ref=zg_bs_nav_pet-supplies_0",
    "Pet Supplies > Dogs": "https://www.amazon.com/gp/bestsellers/pet-supplies/2975312011",
    "Pet Supplies > Dogs > Dog Carriers": "https://www.amazon.com/gp/bestsellers/pet-supplies/2975333011",
    "Pet Supplies > Dogs > Dog Grooming": "https://www.amazon.com/gp/bestsellers/pet-supplies/2975362011",
    "Pet Supplies > Dogs > Dog Feeding": "https://www.amazon.com/gp/bestsellers/pet-supplies/2975351011",
    "Pet Supplies > Dogs > Dog Toys": "https://www.amazon.com/gp/bestsellers/pet-supplies/2975413011",
}

CATEGORY_TITLE_ALIASES = {
    "Pet Supplies > Dogs > Dog Carriers": ["Carriers & Travel Products", "Dog Carriers & Travel Products"],
    "Pet Supplies > Dogs > Dog Grooming": ["Grooming", "Dog Grooming Supplies"],
    "Pet Supplies > Dogs > Dog Feeding": ["Feeding & Watering Supplies", "Dog Feeding & Watering Supplies"],
    "Pet Supplies > Dogs > Dog Toys": ["Toys", "Dog Toys"],
    "Pet Supplies > Cats > Cat Litter": ["Litter & Housebreaking", "Cat Litter"],
    "Pet Supplies > Cats > Cat Trees": ["Trees", "Cat Trees"],
    "Pet Supplies > Cats > Cat Grooming": ["Grooming", "Cat Grooming Supplies"],
}


TEXT = {
    "English": {
        "title": "Amazon Selection Agent",
        "caption": "Amazon product collection, filtering, selection, and Excel export.",
        "task": "Task",
        "ui_language": "UI language",
        "list_type": "List type",
        "categories": "Categories",
        "all_categories": "All categories",
        "filters": "Filters",
        "price": "Price",
        "reviews": "Reviews",
        "monthly_sales": "Monthly sales",
        "child_sales": "Child sales",
        "bsr": "BSR",
        "launched_at": "Launch date",
        "pages": "Pages per category",
        "custom_url": "Amazon URL (optional)",
        "run": "Start Collection",
        "clear": "Clear Results",
        "cards": "Product Cards",
        "table": "Table",
        "log": "Log",
        "open_amazon": "Open Amazon",
        "review_status": "Review status",
        "note": "Note",
        "include": "Include in export",
    },
    "中文": {
        "title": "亚马逊选品 AI 智能体",
        "caption": "亚马逊产品采集、筛选、勾选与 Excel 导出。",
        "task": "任务",
        "ui_language": "界面语言",
        "list_type": "榜单类型",
        "categories": "类目",
        "all_categories": "全部类目",
        "filters": "筛选条件",
        "price": "价格",
        "reviews": "评分数",
        "monthly_sales": "月销量",
        "child_sales": "子体销量",
        "bsr": "BSR",
        "launched_at": "上架时间",
        "pages": "每个类目页数",
        "custom_url": "指定 Amazon 链接（可选）",
        "run": "开始采集",
        "clear": "清空结果",
        "cards": "产品卡片",
        "table": "表格",
        "log": "日志",
        "open_amazon": "打开 Amazon",
        "review_status": "审核状态",
        "note": "备注",
        "include": "加入导出",
    },
}

BLOCKED_KEYWORDS = {
    "supplement",
    "vitamin",
    "adult",
    "book",
    "shoes",
    "clothing",
    "jewelry",
    "video game",
    "amazon basics",
    "kindle",
}


@dataclass
class Product:
    selected: bool
    review_status: str
    note: str
    list_type: str
    category_path: str
    rank: int
    title: str
    asin: str
    amazon_url: str
    image_url: str
    price: float
    rating: float
    review_count: int
    monthly_bought: int
    brand: str
    prime_fba: str
    delivery: str
    variant_count: int
    launched_at: str
    package_dimensions: str
    package_weight_lb: float
    fba_fee: float
    potential_score: int
    potential_level: str
    reason: str
    risk_tags: str
    scraped_at: str
    status: str
    error: str
    seller_name: str = ""
    seller_count: int = 0
    sales_amount: float = 0.0
    child_monthly_sales: int = 0
    child_monthly_sales_label: str = ""
    bsr_rank: int = 0
    bsr_category: str = ""
    sub_rank: int = 0
    sub_category: str = ""
    margin_rate: str = ""
    plugin_data_loaded: bool = False


class CollectionStopped(RuntimeError):
    pass


class SellerSpriteLoadTimeout(RuntimeError):
    def __init__(self, message: str, products: list[Product], page_count: int):
        super().__init__(message)
        self.products = products
        self.page_count = page_count


def clear_stop_collection_flag() -> None:
    try:
        STOP_COLLECTION_FLAG.unlink(missing_ok=True)
    except OSError:
        pass


def mark_collection_running() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    COLLECTION_RUNNING_FLAG.write_text("running", encoding="utf-8")


def clear_collection_running_flag() -> None:
    try:
        COLLECTION_RUNNING_FLAG.unlink(missing_ok=True)
    except OSError:
        pass


def collection_running() -> bool:
    return COLLECTION_RUNNING_FLAG.exists()


def stop_collection_requested() -> bool:
    return STOP_COLLECTION_FLAG.exists()


def raise_if_stop_requested() -> None:
    if stop_collection_requested():
        raise CollectionStopped("用户请求停止采集：程序已在安全检查点停下，并保留当前已完成的原始采集池。")


class StopCollectionHandler(BaseHTTPRequestHandler):
    def _send(self, status: int = 200, body: str = "ok") -> None:
        data = body.encode("utf-8")
        origin = self.headers.get("Origin") or ""
        allowed_origins = {"http://localhost:8501", "http://127.0.0.1:8501", "null"}
        allowed_origin = origin if origin in allowed_origins else "http://localhost:8501"
        self.send_response(status)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", allowed_origin)
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS, GET")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_OPTIONS(self):
        self._send(204, "")

    def do_GET(self):
        if self.path == "/status":
            if not collection_running():
                self._send(200, "idle")
            else:
                self._send(200, "stopping" if stop_collection_requested() else "running")
            return
        self._send(404, "not found")

    def do_POST(self):
        if self.path == "/stop":
            origin = self.headers.get("Origin") or ""
            if origin and origin not in {"http://localhost:8501", "http://127.0.0.1:8501", "null"}:
                self._send(403, "forbidden")
                return
            if not collection_running():
                self._send(409, "not running")
                return
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            STOP_COLLECTION_FLAG.write_text("stop", encoding="utf-8")
            self._send(200, "stop requested")
            return
        self._send(404, "not found")

    def log_message(self, format, *args):
        return


@st.cache_resource
def ensure_stop_collection_server() -> bool:
    try:
        server = ThreadingHTTPServer(("127.0.0.1", STOP_COLLECTION_PORT), StopCollectionHandler)
    except OSError:
        return False
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    return True


def request_stop_collection() -> None:
    if not st.session_state.get("collection_in_progress", False) and not collection_running():
        st.session_state.last_collection_summary = "当前没有正在运行的采集，无需停止。"
        log("Stop requested while collection is idle; ignored.")
        return
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    STOP_COLLECTION_FLAG.write_text("stop", encoding="utf-8")
    st.session_state.last_collection_summary = "已请求停止采集：程序会在当前页面或当前小类结束后停下，并保留已完成的原始采集池。"
    log("Stop collection requested from UI.")


def close_category_dialog_state() -> None:
    st.session_state.show_category_dialog = False


def reset_result_page() -> None:
    st.session_state.result_current_page = 1


def set_result_view_mode(mode: str) -> None:
    st.session_state.result_view_mode = mode


def handle_result_page_size_change() -> None:
    reset_result_page()


def handle_result_page_jump_change() -> None:
    st.session_state.result_current_page = clamp_page(
        st.session_state.get("result_page_jump", 1),
        len(st.session_state.get("products", [])),
        st.session_state.get("result_page_size", PAGE_SIZE_OPTIONS[0]),
    )


def prepare_collection_run() -> None:
    clear_stop_collection_flag()
    mark_collection_running()
    st.session_state.collection_in_progress = True
    st.session_state.collection_start_requested = True


def render_stop_collection_button() -> None:
    server_ready = ensure_stop_collection_server()
    disabled_attr = "" if server_ready else "disabled"
    components.html(
        f"""
        <style>
          html, body {{
            margin: 0;
            padding: 0;
            overflow: hidden;
          }}
        </style>
        <button id="stop-collection-btn" {disabled_attr} style="
            display: block;
            width: 100%;
            height: 42px;
            border: 1px solid #d94b4b;
            border-radius: 8px;
            background: #fff;
            color: #c73e3e;
            font-size: 15px;
            font-weight: 600;
            cursor: pointer;
            transition: background .15s ease, border-color .15s ease, color .15s ease;
        ">停止采集</button>
        <script>
        const btn = document.getElementById("stop-collection-btn");
        if (btn) {{
          btn.addEventListener("click", async () => {{
            btn.disabled = true;
            btn.textContent = "停止中...";
            try {{
              const res = await fetch("http://127.0.0.1:{STOP_COLLECTION_PORT}/stop", {{ method: "POST" }});
              if (res.ok) {{
                btn.textContent = "已请求停止";
              }} else if (res.status === 409) {{
                btn.textContent = "当前无采集";
              }} else {{
                btn.textContent = "停止失败";
              }}
            }} catch (error) {{
              btn.textContent = "停止失败";
            }}
          }});
        }}
        </script>
        """,
        height=42,
    )


def ensure_state():
    if "raw_products" not in st.session_state:
        st.session_state.raw_products = []
    if "collection_staged_raw_products" not in st.session_state:
        st.session_state.collection_staged_raw_products = []
    if "products" not in st.session_state:
        st.session_state.products = []
    if "run_log" not in st.session_state:
        st.session_state.run_log = ["Ready."]
    if "confirmed_category_paths" not in st.session_state:
        st.session_state.confirmed_category_paths = []
    if "show_category_dialog" not in st.session_state:
        st.session_state.show_category_dialog = False
    if "category_dialog_selected_paths" not in st.session_state:
        st.session_state.category_dialog_selected_paths = []
    if "last_cache_refresh_message" not in st.session_state:
        st.session_state.last_cache_refresh_message = ""
    if "last_category_mapping_message" not in st.session_state:
        st.session_state.last_category_mapping_message = ""
    if "last_collection_summary" not in st.session_state:
        st.session_state.last_collection_summary = ""
    if "last_raw_products_message" not in st.session_state:
        st.session_state.last_raw_products_message = ""
    if "category_search" not in st.session_state:
        st.session_state.category_search = ""
    if "filter_auto_apply_requested" not in st.session_state:
        st.session_state.filter_auto_apply_requested = False
    if "collection_in_progress" not in st.session_state:
        st.session_state.collection_in_progress = False
    if "collection_start_requested" not in st.session_state:
        st.session_state.collection_start_requested = False
    if "result_page_size" not in st.session_state:
        st.session_state.result_page_size = PAGE_SIZE_OPTIONS[0]
    if "result_current_page" not in st.session_state:
        st.session_state.result_current_page = 1
    if "result_page_jump" not in st.session_state:
        st.session_state.result_page_jump = 1
    if "result_export_scope" not in st.session_state:
        st.session_state.result_export_scope = "导出已勾选"
    if "result_view_mode" not in st.session_state:
        st.session_state.result_view_mode = "列表"
    if "select_current_page_products" not in st.session_state:
        st.session_state.select_current_page_products = False
    if "favorite_products" not in st.session_state:
        st.session_state.favorite_products = load_favorites(FAVORITES_PATH)
    if "product_notes" not in st.session_state:
        st.session_state.product_notes = load_notes(PRODUCT_NOTES_PATH)
    if "collection_total_raw_count" not in st.session_state:
        st.session_state.collection_total_raw_count = 0
    if "collection_completed_seed_count" not in st.session_state:
        st.session_state.collection_completed_seed_count = 0
    if "collection_failed_seed_count" not in st.session_state:
        st.session_state.collection_failed_seed_count = 0
    if "collection_warning_seed_count" not in st.session_state:
        st.session_state.collection_warning_seed_count = 0
    if "collection_empty_seed_count" not in st.session_state:
        st.session_state.collection_empty_seed_count = 0
    if "collection_failed_seed_details" not in st.session_state:
        persisted_failures = []
        if LAST_COLLECTION_REPORT.exists():
            try:
                previous_report = json.loads(LAST_COLLECTION_REPORT.read_text(encoding="utf-8"))
                persisted_failures = [
                    {
                        "label": item.get("label", "未知入口"),
                        "url": item.get("url", ""),
                        "error": item.get("error") or item.get("quality_message") or "未知错误",
                    }
                    for item in previous_report.get("entries", [])
                    if item.get("failed")
                ]
            except (json.JSONDecodeError, OSError):
                persisted_failures = []
        st.session_state.collection_failed_seed_details = persisted_failures
    if "collection_warning_seed_details" not in st.session_state:
        persisted_warnings = []
        if LAST_COLLECTION_REPORT.exists():
            try:
                previous_report = json.loads(LAST_COLLECTION_REPORT.read_text(encoding="utf-8"))
                persisted_warnings = [
                    {
                        "label": item.get("label", "未知入口"),
                        "url": item.get("url", ""),
                        "warning": item.get("quality_message") or "产品数量偏少",
                    }
                    for item in previous_report.get("entries", [])
                    if item.get("warning")
                ]
            except (json.JSONDecodeError, OSError):
                persisted_warnings = []
        st.session_state.collection_warning_seed_details = persisted_warnings
    if "collection_total_seed_count" not in st.session_state:
        st.session_state.collection_total_seed_count = 0
    if "collection_current_seed_index" not in st.session_state:
        st.session_state.collection_current_seed_index = 0
    if "collection_current_seed_label" not in st.session_state:
        st.session_state.collection_current_seed_label = ""


def log(message: str):
    st.session_state.run_log.append(f"{datetime.now().strftime('%H:%M:%S')}  {message}")


def readable_log_line(line: str) -> str:
    if line == "Ready.":
        return "准备就绪。"
    prefix = ""
    message = line
    if "  " in line:
        prefix, message = line.split("  ", 1)
        prefix = f"{prefix}  "
    translated = message
    if message.startswith("Stop requested while collection is idle"):
        translated = "点击了停止采集，但当前没有正在运行的采集。"
    elif message.startswith("Stop collection requested"):
        translated = "已请求停止采集，程序会在当前页面结束后保存已完成结果。"
    elif message.startswith("Loaded raw product pool from disk:"):
        count = message.split(":", 1)[1].strip().split(" ", 1)[0]
        translated = f"已载入最近一次原始采集池：{count} 条。"
    elif message.startswith("Loaded historical raw product pool:"):
        count = message.split(":", 1)[1].strip().split(" ", 1)[0]
        translated = f"已载入历史原始采集池：{count} 条。"
    elif message.startswith("Auto-applied filters after filter input change:"):
        translated = "筛选条件变更后，已自动重新计算当前结果。"
    elif message.startswith("Re-applied filters to raw product pool:"):
        translated = "已按当前筛选条件重新计算产品列表。"
    elif message.startswith("Filter widgets reset and filters re-applied"):
        translated = "已清空筛选条件，并重新计算产品列表。"
    elif message.startswith("Start batch category collection"):
        translated = "开始批量采集已选择的类目入口。"
    elif message.startswith("Open Amazon page and wait for SellerSprite plugin data:"):
        label = message.split(":", 1)[1].strip()
        translated = f"打开 Amazon 页面并等待卖家精灵加载：{label}。"
    elif message.startswith("SellerSprite plugin data refresh:"):
        translated = "卖家精灵页面数据读取完成，正在转成产品卡。"
    elif message.startswith("SellerSprite plugin collection finished."):
        translated = "本次卖家精灵采集已完成，产品已去重。"
    elif message.startswith("Filters applied:"):
        translated = "采集后已应用筛选条件，并更新产品列表。"
    elif message.startswith("Collection stopped by user."):
        translated = "采集已按用户请求停止，并保留已完成产品。"
    elif " collection failed:" in message:
        translated = f"采集失败：{message.split(' collection failed:', 1)[1].strip()}"
    elif message.startswith("Copied ") and " ASIN" in message:
        count = message.split(" ", 2)[1]
        translated = f"已复制 {count} 个 ASIN。"
    return prefix + translated


def request_filter_auto_apply() -> None:
    st.session_state.filter_auto_apply_requested = True


def sync_product_selection_from_widgets(products):
    for product in products:
        key = f"row_include_{product.asin}"
        if key in st.session_state:
            product.selected = bool(st.session_state[key])


def set_all_product_selection(products, selected: bool):
    for product in products:
        product.selected = selected
        st.session_state[f"row_include_{product.asin}"] = selected


def clear_product_selection_widget_state(products: list["Product"] | None = None) -> None:
    product_asins = {product.asin for product in products} if products is not None else None
    for key in list(st.session_state.keys()):
        if not key.startswith("row_include_"):
            continue
        if product_asins is None or key.removeprefix("row_include_") in product_asins:
            del st.session_state[key]


def select_all_filtered_products(products: list["Product"], visible_products: list["Product"]) -> None:
    for product in products:
        product.selected = True
    clear_product_selection_widget_state(products)
    for product in visible_products:
        st.session_state[f"row_include_{product.asin}"] = True


def handle_select_all_products_change():
    set_all_product_selection(
        st.session_state.products,
        bool(st.session_state.get("select_all_products", False)),
    )


def handle_select_current_page_products_change(products: list["Product"]):
    set_all_product_selection(
        products,
        bool(st.session_state.get("select_current_page_products", False)),
    )


def sync_product_annotations(products: list["Product"]) -> None:
    notes = st.session_state.get("product_notes", {})
    for product in products:
        if product.asin in notes:
            product.note = notes[product.asin]


def product_annotation_payload(product: "Product") -> dict:
    payload = asdict(product)
    payload["selected"] = False
    return payload


def toggle_product_favorite(product: "Product") -> None:
    asin = product.asin
    if not asin:
        return
    favorites = st.session_state.setdefault("favorite_products", {})
    if asin in favorites:
        remove_favorite(favorites, asin)
        log(f"Removed favorite: {asin}.")
    else:
        upsert_favorite(
            favorites,
            product_annotation_payload(product),
            source_label=product.category_path,
            source_url=product.amazon_url,
        )
        log(f"Added favorite: {asin}.")
    save_favorites(FAVORITES_PATH, favorites)


def apply_note_to_loaded_products(asin: str, note: str) -> None:
    for state_key in ("products", "raw_products", "collection_staged_raw_products"):
        for product in st.session_state.get(state_key, []):
            if getattr(product, "asin", "") == asin:
                product.note = note


def handle_product_note_change(asin: str, widget_key: str) -> None:
    notes = st.session_state.setdefault("product_notes", {})
    note = str(st.session_state.get(widget_key, "") or "").strip()
    set_product_note(notes, asin, note)
    save_notes(PRODUCT_NOTES_PATH, notes)
    apply_note_to_loaded_products(asin, note)
    st.session_state[f"product_note_editing_{asin}"] = False


def begin_list_note_edit(asin: str) -> None:
    note_key = f"product_note_list_{asin}"
    current_note = st.session_state.get("product_notes", {}).get(asin, "")
    if not current_note:
        for state_key in ("products", "raw_products", "collection_staged_raw_products"):
            for product in st.session_state.get(state_key, []):
                if getattr(product, "asin", "") == asin:
                    current_note = getattr(product, "note", "") or ""
                    break
            if current_note:
                break
    st.session_state[note_key] = current_note
    st.session_state[f"product_note_editing_{asin}"] = True


def render_list_product_favorite_button(product: "Product") -> None:
    asin = product.asin
    if not asin:
        return
    is_favorite = asin in st.session_state.get("favorite_products", {})
    active_class = " product-list-favorite-active-anchor" if is_favorite else ""
    st.markdown(
        f"<span class='product-list-favorite-anchor{active_class}' data-asin='{escape(asin)}'></span>",
        unsafe_allow_html=True,
    )
    st.button(
        "♥" if is_favorite else "♡",
        key=f"favorite_list_{asin}",
        help="取消收藏" if is_favorite else "收藏",
        on_click=toggle_product_favorite,
        args=(product,),
    )


def render_list_product_note_input(product: "Product") -> None:
    asin = product.asin
    if not asin:
        return
    note_key = f"product_note_list_{asin}"
    edit_key = f"product_note_editing_{asin}"
    current_note = st.session_state.get("product_notes", {}).get(asin, product.note or "")
    if note_key not in st.session_state:
        st.session_state[note_key] = current_note
    if st.session_state.get(edit_key):
        st.markdown(
            f"<span class='product-list-note-editor-anchor' data-asin='{escape(asin)}'></span>",
            unsafe_allow_html=True,
        )
        with st.form(key=f"product_note_form_{asin}", border=False):
            st.text_input(
                "备注",
                key=note_key,
                placeholder="备注：未备注",
                label_visibility="collapsed",
            )
            st.form_submit_button(
                "保存备注",
                on_click=handle_product_note_change,
                args=(asin, note_key),
            )
        return
    st.markdown(
        f"<span class='product-list-note-display-anchor' data-asin='{escape(asin)}'></span>",
        unsafe_allow_html=True,
    )
    note_label = f"备注：{current_note or '未备注'}"
    st.button(
        note_label,
        key=f"product_note_display_{asin}",
        help="点击编辑备注",
        on_click=begin_list_note_edit,
        args=(asin,),
    )


def render_product_annotation_controls(product: "Product", location: str) -> None:
    asin = product.asin
    if not asin:
        return
    note_key = f"product_note_{location}_{asin}"
    if note_key not in st.session_state:
        st.session_state[note_key] = st.session_state.get("product_notes", {}).get(asin, product.note or "")
    is_favorite = asin in st.session_state.get("favorite_products", {})
    st.markdown("<span class='product-annotation-anchor'></span>", unsafe_allow_html=True)
    annotation_cols = st.columns([0.58, 2.4], vertical_alignment="center")
    annotation_cols[0].button(
        "已收藏" if is_favorite else "收藏",
        key=f"favorite_{location}_{asin}",
        use_container_width=True,
        type="primary" if is_favorite else "secondary",
        on_click=toggle_product_favorite,
        args=(product,),
    )
    annotation_cols[1].text_input(
        "备注",
        key=note_key,
        placeholder="备注会按 ASIN 自动保存",
        label_visibility="collapsed",
        on_change=handle_product_note_change,
        args=(asin, note_key),
    )


def render_list_favorite_portal() -> None:
    components.html(
        """
        <script>
        window.parent.eval(`
          (() => {
            const clickRealControl = (selector) => {
              const real = document.querySelector(selector);
              if (real) real.click();
            };
            if (!window.__amazonSelectionListProxyBound) {
              window.__amazonSelectionListProxyBound = true;
              document.addEventListener('pointerdown', (event) => {
                const editorAnchor = document.querySelector('.product-list-note-editor-anchor[data-asin]');
                const asin = editorAnchor?.dataset.asin || '';
                if (!asin) return;
                const noteForm = editorAnchor
                  ?.closest('div[data-testid="stElementContainer"]')
                  ?.nextElementSibling
                  ?.querySelector('div[data-testid="stForm"]');
                const noteInput = noteForm?.querySelector('input');
                if (!noteInput) return;
                const clickedInsideEditor = noteForm.contains(event.target);
                const clickedNoteProxy = event.target.closest('.list-note-proxy[data-asin="' + asin + '"]');
                if (clickedInsideEditor || clickedNoteProxy) return;
                const tracker = noteInput._valueTracker;
                if (tracker) tracker.setValue(String(noteInput.value || '') + '__force_streamlit_sync__');
                noteInput.dispatchEvent(new Event('input', { bubbles: true }));
                noteInput.dispatchEvent(new Event('change', { bubbles: true }));
                noteInput.blur();
                setTimeout(() => {
                  noteForm.querySelector('button')?.click();
                }, 500);
              }, true);
              document.addEventListener('click', (event) => {
                const selectProxy = event.target.closest('.list-select-proxy[data-asin]');
                if (selectProxy) {
                  event.preventDefault();
                  const asin = selectProxy.dataset.asin || '';
                  clickRealControl('div.st-key-row_include_' + asin + ' input[type="checkbox"]');
                  return;
                }
                const favoriteProxy = event.target.closest('.list-favorite-proxy[data-asin]');
                if (favoriteProxy) {
                  event.preventDefault();
                  const asin = favoriteProxy.dataset.asin || '';
                  clickRealControl('div.st-key-favorite_list_' + asin + ' button');
                  return;
                }
                const noteProxy = event.target.closest('.list-note-proxy[data-asin]');
                if (noteProxy) {
                  event.preventDefault();
                  const asin = noteProxy.dataset.asin || '';
                  clickRealControl('div.st-key-product_note_display_' + asin + ' button');
                }
              });
            }
          })();
        `);
        </script>
        """,
        height=0,
    )


def export_products_for_scope(products: list["Product"], current_page_products: list["Product"], scope: str) -> list["Product"]:
    if scope in {"当前页", "导出当前页"}:
        return current_page_products
    if scope in {"全部筛选结果", "导出全部结果"}:
        return products
    return [product for product in products if product.selected]


def product_from_dict(data: dict) -> Product:
    product_fields = {field.name: field for field in fields(Product)}
    values = {}
    for name, field in product_fields.items():
        if name in data:
            values[name] = data[name]
        elif field.default is not MISSING:
            values[name] = field.default
        else:
            values[name] = None
    product = Product(**values)
    if "plugin_data_loaded" not in data:
        product.plugin_data_loaded = bool(product.monthly_bought)
    repair_product_price(product)
    product.selected = False
    return product


def repair_product_price(product: Product) -> None:
    if not product.sales_amount or not product.monthly_bought:
        return
    inferred_price = round(product.sales_amount / product.monthly_bought, 2)
    price_matches_fba = (
        product.price
        and product.fba_fee
        and abs(product.price - product.fba_fee) < 0.01
        and inferred_price > product.price * 1.5
    )
    if product.price <= 0 or price_matches_fba:
        product.price = inferred_price


def safe_slug(value: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "-", value or "").strip("-").lower()
    return slug[:60] or "collection"


def raw_products_payload(products: list[Product]) -> dict:
    return {
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "count": len(products),
        "products": [asdict(product) for product in products],
    }


def load_raw_products_payload(path: Path) -> tuple[list[Product], dict, str]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        return [], {}, f"读取采集结果失败：{exc}"
    products = [product_from_dict(item) for item in payload.get("products", []) if isinstance(item, dict)]
    return products, payload, ""


def load_raw_history_index() -> list[dict]:
    if not RAW_PRODUCTS_HISTORY_INDEX.exists():
        return []
    try:
        data = json.loads(RAW_PRODUCTS_HISTORY_INDEX.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []
    if isinstance(data, dict):
        data = data.get("records", [])
    return [record for record in data if isinstance(record, dict)]


def write_raw_history_index(records: list[dict]) -> None:
    RAW_PRODUCTS_HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    RAW_PRODUCTS_HISTORY_INDEX.write_text(
        json.dumps({"records": records[:RAW_PRODUCTS_HISTORY_LIMIT]}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def save_raw_products(products: list[Product], label: str = "", source_url: str = "") -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    RAW_PRODUCTS_HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    payload = raw_products_payload(products)
    payload["label"] = label or "未命名采集"
    payload["source_url"] = source_url or ""
    RAW_PRODUCTS_CACHE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    history_name = f"{timestamp}_{safe_slug(label)}.json"
    history_path = RAW_PRODUCTS_HISTORY_DIR / history_name
    history_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    records = load_raw_history_index()
    records.insert(0, {
        "saved_at": payload["saved_at"],
        "count": len(products),
        "label": payload["label"],
        "source_url": payload["source_url"],
        "file": str(history_path),
    })
    kept_records = records[:RAW_PRODUCTS_HISTORY_LIMIT]
    kept_files = {record.get("file") for record in kept_records}
    for old_record in records[RAW_PRODUCTS_HISTORY_LIMIT:]:
        old_file = old_record.get("file")
        if old_file and old_file not in kept_files:
            try:
                Path(old_file).unlink(missing_ok=True)
            except OSError:
                pass
    write_raw_history_index(kept_records)


def save_collection_report(
    seed_summaries: list[dict],
    total_raw_count: int,
    status: str,
    planned_seeds: list[tuple[str, str]] | None = None,
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "status": status,
        "planned_seed_count": len(planned_seeds or []),
        "completed_seed_count": len(seed_summaries),
        "failed_seed_count": sum(bool(item.get("failed")) for item in seed_summaries),
        "warning_seed_count": sum(bool(item.get("warning")) for item in seed_summaries),
        "empty_seed_count": sum(bool(item.get("empty")) for item in seed_summaries),
        "total_raw_count": total_raw_count,
        "planned_seeds": [
            {"label": label, "url": url}
            for label, url in (planned_seeds or [])
        ],
        "entries": seed_summaries,
    }
    LAST_COLLECTION_REPORT.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_raw_products() -> tuple[list[Product], str]:
    if not RAW_PRODUCTS_CACHE.exists():
        return [], "本地还没有最近一次原始采集池。请先完成一次采集。"
    products, payload, error = load_raw_products_payload(RAW_PRODUCTS_CACHE)
    if error:
        return [], error.replace("采集结果", "上次采集结果")
    saved_at = payload.get("saved_at", "-")
    return products, f"已载入最近一次原始采集池：{len(products)} 条，保存时间：{saved_at}。载入后已按当前筛选条件重新计算。"


def raw_history_options() -> list[tuple[str, Path]]:
    options = []
    for record in load_raw_history_index():
        file_path = Path(record.get("file", ""))
        if not file_path.exists():
            continue
        saved_at = str(record.get("saved_at", "-")).replace("T", " ")
        label = record.get("label") or "未命名采集"
        count = record.get("count", 0)
        options.append((f"{saved_at}｜{label}｜{count} 条", file_path))
    return options


def empty_products_message() -> str:
    if st.session_state.raw_products:
        return "原始采集池已有产品，但没有符合当前筛选条件的结果。可以放宽筛选条件后点击“应用筛选”，不需要重新采集。"
    if st.session_state.last_collection_summary:
        return "本次没有解析到产品，或当前类目确实没有可读取产品。请看上方采集提示和日志。"
    return "还没有产品数据。选择类目后点击“开始采集”。"


def clear_progress_bar(progress_bar, delay_seconds: float = 0.6) -> None:
    if not progress_bar:
        return
    time.sleep(delay_seconds)
    progress_bar.empty()


def reset_collection_run_messages() -> None:
    st.session_state.last_cache_refresh_message = ""
    st.session_state.last_collection_summary = ""
    st.session_state.last_raw_products_message = ""
    st.session_state.collection_staged_raw_products = []
    st.session_state.collection_total_raw_count = 0
    st.session_state.collection_completed_seed_count = 0
    st.session_state.collection_failed_seed_count = 0
    st.session_state.collection_warning_seed_count = 0
    st.session_state.collection_empty_seed_count = 0
    st.session_state.collection_failed_seed_details = []
    st.session_state.collection_warning_seed_details = []
    st.session_state.collection_total_seed_count = 0
    st.session_state.collection_current_seed_index = 0
    st.session_state.collection_current_seed_label = ""


def update_collection_total_status(placeholder=None) -> None:
    total_raw = int(st.session_state.get("collection_total_raw_count", 0) or 0)
    completed = int(st.session_state.get("collection_completed_seed_count", 0) or 0)
    failed = int(st.session_state.get("collection_failed_seed_count", 0) or 0)
    warning = int(st.session_state.get("collection_warning_seed_count", 0) or 0)
    empty = int(st.session_state.get("collection_empty_seed_count", 0) or 0)
    total_seeds = int(st.session_state.get("collection_total_seed_count", 0) or 0)
    current_seed = int(st.session_state.get("collection_current_seed_index", 0) or 0)
    current_label = str(st.session_state.get("collection_current_seed_label", "") or "")
    text = build_collection_total_status_text(
        total_raw=total_raw,
        completed=completed,
        total_seeds=total_seeds,
        current_seed=current_seed,
        current_label=current_label,
        empty=empty,
        warning=warning,
        failed=failed,
    )
    target = placeholder if placeholder is not None else st
    target.markdown(text)


def load_sellersprite_image_cache() -> dict[str, str]:
    if not SELLERSPRITE_IMAGE_CACHE.exists():
        return {}
    try:
        data = json.loads(SELLERSPRITE_IMAGE_CACHE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
    return {asin: url for asin, url in data.items() if asin and url}


def load_sellersprite_cache_meta() -> dict:
    if not SELLERSPRITE_META_CACHE.exists():
        return {}
    try:
        return json.loads(SELLERSPRITE_META_CACHE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}


def learned_category_links_mtime() -> float:
    try:
        return LEARNED_CATEGORY_LINKS.stat().st_mtime
    except OSError:
        return 0.0


@lru_cache(maxsize=4)
def load_learned_category_links_cached(mtime: float) -> dict:
    if not LEARNED_CATEGORY_LINKS.exists():
        return {}
    try:
        return json.loads(LEARNED_CATEGORY_LINKS.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}


def load_learned_category_links() -> dict:
    return deepcopy(load_learned_category_links_cached(learned_category_links_mtime()))


@lru_cache(maxsize=4)
def display_categories_cached(mtime: float) -> dict:
    categories = deepcopy(CATEGORIES)
    learned = load_learned_category_links_cached(mtime).get("categories", {})
    allowed_roots = set(categories)
    for path, payload in learned.items():
        parts = [part.strip() for part in path.split(" > ") if part.strip()]
        if not parts or parts[0] not in allowed_roots:
            continue
        if not category_url_matches_path(path, str(payload.get("url") or "")):
            continue
        if is_legacy_home_path(path):
            continue
        current = categories
        for part in parts:
            current.setdefault(part, {"zh": "", "count": 0, "children": {}})
            current[part].setdefault("children", {})
            node = current[part]
            current = node["children"]
    return categories


def display_categories() -> dict:
    return deepcopy(display_categories_cached(learned_category_links_mtime()))


def save_learned_category_links(seed_label: str, links) -> None:
    learned = load_learned_category_links()
    categories = learned.setdefault("categories", {})
    for link in links:
        title = str(getattr(link, "title", "") or "").strip()
        url = str(getattr(link, "url", "") or "").strip()
        if not title or not url:
            continue
        link_path = str(getattr(link, "path", "") or "").strip()
        storage_title = link_path or title
        if seed_label and not seed_label.startswith("http"):
            seed_parts = [part.strip() for part in seed_label.split(" > ") if part.strip()]
            if title in seed_parts:
                continue
            root = seed_parts[0] if seed_parts else ""
            root_children = set(CATEGORIES.get(root, {}).get("children", {}))
            if len(seed_parts) > 1 and title in root_children:
                continue
            if not storage_title.startswith(seed_label):
                storage_title = f"{seed_label} > {storage_title}"
        if not category_url_matches_path(storage_title, url):
            continue
        if is_legacy_home_path(storage_title):
            continue
        categories[storage_title] = {
            "title": title,
            "url": url,
            "source": seed_label,
            "is_leaf": bool(getattr(link, "is_leaf", True)),
            "node": str(getattr(link, "node", "") or ""),
            "depth": int(getattr(link, "depth", 0) or 0),
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
    learned["categories"], _ = clean_category_entries(categories)
    LEARNED_CATEGORY_LINKS.parent.mkdir(parents=True, exist_ok=True)
    LEARNED_CATEGORY_LINKS.write_text(json.dumps(learned, ensure_ascii=False, indent=2), encoding="utf-8")
    load_learned_category_links_cached.cache_clear()
    learned_category_lookup.cache_clear()
    display_categories_cached.cache_clear()


def normalize_category_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


@lru_cache(maxsize=4)
def learned_category_lookup(mtime: float) -> tuple[dict[str, str], tuple[tuple[str, str], ...]]:
    learned = load_learned_category_links_cached(mtime).get("categories", {})
    exact = {}
    normalized = []
    for title, payload in learned.items():
        url = str(payload.get("url") or "")
        if not title or not url:
            continue
        if not category_url_matches_path(title, url):
            continue
        if is_legacy_home_path(title):
            continue
        exact[title] = url
        normalized.append((normalize_category_text(title), url))
    return exact, tuple(normalized)


def find_learned_category_url(path: str) -> str:
    exact, normalized_learned = learned_category_lookup(learned_category_links_mtime())
    if path in exact:
        return exact[path]
    leaf = path.split(" > ")[-1]
    candidates = [leaf]
    candidates.extend(CATEGORY_TITLE_ALIASES.get(path, []))
    normalized_candidates = [normalize_category_text(candidate) for candidate in candidates if candidate]
    if path.count(" > ") >= 1:
        return ""
    for normalized_title, url in normalized_learned:
        if any(candidate and (candidate in normalized_title or normalized_title in candidate) for candidate in normalized_candidates):
            return url
    return ""


def find_exact_category_url(path: str) -> str:
    return CATEGORY_BESTSELLER_URLS.get(path) or find_learned_category_url(path)


def find_category_seed_url(path: str) -> tuple[str, str]:
    direct_url = find_exact_category_url(path)
    if direct_url:
        return path, direct_url
    parts = path.split(" > ")
    for index in range(len(parts) - 1, 0, -1):
        ancestor = " > ".join(parts[:index])
        ancestor_url = find_exact_category_url(ancestor)
        if ancestor_url:
            return ancestor, ancestor_url
    return "", ""


def amazon_url_for_list_type(url: str, list_type: str) -> str:
    if not url:
        return url
    target_kind = "new-releases" if list_type == "New Releases" else "bestsellers"
    return re.sub(r"/gp/(?:bestsellers|new-releases)/", f"/gp/{target_kind}/", url, count=1)


def refresh_category_link_mapping(selected_paths: list[str], custom_url: str = "", progress_bar=None) -> tuple[int, int]:
    seed_candidates: list[tuple[str, str]] = []
    seen: set[str] = set()

    if custom_url:
        seed_candidates.append(("自定义链接", custom_url))
        seen.add(custom_url)

    paths = selected_paths or list(CATEGORIES.keys())
    for path in paths:
        _, seed_url = find_category_seed_url(path)
        if not seed_url:
            top = path.split(" > ")[0]
            seed_url = CATEGORY_BESTSELLER_URLS.get(top, "")
        if seed_url and seed_url not in seen:
            seed_candidates.append((path, seed_url))
            seen.add(seed_url)

    if not seed_candidates:
        raise ValueError("没有可用于刷新类目映射的 Amazon 榜单入口。")

    learned_count = 0
    failed_count = 0
    queue = list(seed_candidates)
    processed_urls: set[str] = set()
    index = 0
    while index < len(queue):
        label, seed_url = queue[index]
        index += 1
        if seed_url in processed_urls:
            continue
        processed_urls.add(seed_url)
        if progress_bar:
            progress_bar.progress(
                min(95, int(((index - 1) / max(len(queue), 1)) * 100)),
                text=f"正在准备类目链接 {index}/{len(queue)}：{label}",
            )
        try:
            links = discover_bestseller_category_links(seed_url, max_links=120)
            save_learned_category_links(label, links)
            learned_count += len(links)
            log(f"Category mapping refreshed: {label}, learned {len(links)} links.")

            for path in selected_paths:
                if find_exact_category_url(path):
                    continue
                parts = path.split(" > ")
                for depth in range(len(parts) - 1, 0, -1):
                    ancestor = " > ".join(parts[:depth])
                    ancestor_url = find_exact_category_url(ancestor)
                    if ancestor_url and ancestor_url not in processed_urls and all(ancestor_url != queued_url for _, queued_url in queue):
                        queue.append((ancestor, ancestor_url))
                    if ancestor_url:
                        break
        except Exception as exc:
            failed_count += 1
            log(f"Category mapping refresh failed: {label}. {exc}")

    if progress_bar:
        progress_bar.progress(100, text=f"类目链接准备完成：发现/更新 {learned_count} 个类目入口，失败 {failed_count} 个入口。")
        time.sleep(0.4)
        progress_bar.empty()
    return learned_count, failed_count


def sellersprite_cache_summary() -> str:
    meta = load_sellersprite_cache_meta()
    captured_at = meta.get("captured_at", "未知时间")
    source_url = meta.get("source_url", "未知链接")
    product_count = meta.get("product_count", 0)
    if not SELLERSPRITE_DOM_CACHE.exists():
        return "还没有卖家精灵插件数据，请先采集。"
    updated_at = datetime.fromtimestamp(SELLERSPRITE_DOM_CACHE.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    return f"卖家精灵插件数据：{product_count} 条｜抓取时间：{captured_at or updated_at}｜来源：{source_url}"


def sellersprite_cache_hydration() -> tuple[int, int]:
    if not SELLERSPRITE_DOM_CACHE.exists():
        return 0, 0
    try:
        products = load_cached_sellersprite_products(limit=50)
    except Exception:
        return 0, 0
    hydrated = sum(1 for product in products if product.parent_monthly_sales or product.sales_amount or product.fba_fee)
    return len(products), hydrated


def sellersprite_cache_warning() -> str:
    if not SELLERSPRITE_DOM_CACHE.exists():
        return ""
    try:
        products = load_cached_sellersprite_products(limit=50)
    except Exception:
        return "卖家精灵插件数据存在，但解析失败。"
    if not products:
        return "卖家精灵插件数据里没有解析到产品。"
    hydrated = sum(1 for product in products if product.parent_monthly_sales or product.sales_amount or product.fba_fee)
    if hydrated < len(products):
        return f"卖家精灵插件字段未完全加载：共 {len(products)} 条产品，{hydrated} 条有销量/FBA/销售额字段。建议重新采集后再导出。"
    return ""


def sellersprite_cache_status_html(total: int, hydrated: int, chrome_ready: bool) -> str:
    meta = load_sellersprite_cache_meta()
    captured_at = str(meta.get("captured_at") or "未知时间").replace("T", " ")
    source_url = str(meta.get("source_url") or "")
    percent = int((hydrated / total) * 100) if total else 0
    status_text = "最近页面完整" if total and hydrated >= total else "等待插件加载"
    chrome_text = "Chrome 已连接" if chrome_ready else "Chrome 未连接"
    chrome_class = "ok" if chrome_ready else "warn"
    source_link = (
        f'<a href="{escape(source_url)}" target="_blank" rel="noopener noreferrer">来源链接</a>'
        if source_url
        else "<span>暂无来源</span>"
    )
    return f"""
    <div class="cache-card">
        <div class="cache-card-top">
            <div>
                <div class="cache-title">最近页面插件数据</div>
                <div class="cache-sub">抓取时间：{escape(captured_at)} · {source_link}</div>
            </div>
            <div class="cache-badges">
                <span class="cache-badge {chrome_class}">{chrome_text}</span>
                <span class="cache-badge">{status_text}</span>
            </div>
        </div>
        <div class="cache-progress"><span style="width:{percent}%"></span></div>
        <div class="cache-foot"><strong>{total}</strong> 条产品，<strong>{hydrated}</strong> 条卖家精灵字段完整</div>
    </div>
    """


def detect_risk(title: str, brand: str) -> str:
    text = f"{title} {brand}".lower()
    tags = [keyword for keyword in BLOCKED_KEYWORDS if keyword in text]
    return "; ".join(tags)


def score_product(rank, price, rating, reviews, bought, risks, min_price, max_price, max_reviews, min_bought):
    min_price = 0.0 if min_price is None else min_price
    max_price = float("inf") if max_price is None else max_price
    max_reviews = float("inf") if max_reviews is None else max_reviews
    min_bought = 0 if min_bought is None else min_bought
    score = 50
    reasons = []

    if rank <= 20:
        score += 12
        reasons.append("top rank")
    if reviews <= max_reviews:
        score += 12
        reasons.append("low reviews")
    if bought >= min_bought:
        score += 14
        reasons.append("monthly bought signal")
    if min_price <= price <= max_price:
        score += 8
        reasons.append("good price band")
    if rating >= 4.2:
        score += 6
        reasons.append("healthy rating")
    if reviews > 1000:
        score -= 12
        reasons.append("high competition")
    if price < min_price:
        score -= 10
        reasons.append("low price")
    if rating < 4.0:
        score -= 10
        reasons.append("rating risk")
    if risks:
        score -= 30
        reasons.append("blocked/risk keyword")

    score = max(0, min(100, score))
    if risks:
        level = "Risk"
    elif score >= 82:
        level = "A"
    elif score >= 68:
        level = "B"
    elif score >= 52:
        level = "C"
    else:
        level = "D"
    return score, level, ", ".join(reasons) or "no strong signal"


def parse_filter_number(value: str, default=None, as_int: bool = False):
    text = str(value or "").replace(",", "").replace("$", "").strip()
    if not text:
        return default
    try:
        number = float(text)
    except ValueError:
        return default
    return int(number) if as_int else number


def in_optional_range(value, min_value=None, max_value=None) -> bool:
    if min_value is not None and value < min_value:
        return False
    if max_value is not None and value > max_value:
        return False
    return True


def launched_at_matches(value: str, filter_value: str) -> bool:
    if filter_value in ("不限", "Any"):
        return True
    try:
        launched = datetime.strptime(str(value), "%Y-%m-%d").date()
    except ValueError:
        return False
    today = datetime.now().date()
    days = (today - launched).days
    ranges = {
        "近30天": (0, 30),
        "Last 30 days": (0, 30),
        "近60天": (0, 60),
        "Last 60 days": (0, 60),
        "近3个月": (0, 90),
        "Last 3 months": (0, 90),
        "近半年": (0, 183),
        "Last 6 months": (0, 183),
        "近1年": (0, 365),
        "Last year": (0, 365),
        "近2年": (0, 730),
        "Last 2 years": (0, 730),
        "近1~2年": (366, 730),
        "1-2 years": (366, 730),
    }
    min_days, max_days = ranges.get(filter_value, (0, 999999))
    return min_days <= days <= max_days


def product_matches_filters(product: Product, filters: dict) -> bool:
    return (
        in_optional_range(product.price, filters["min_price"], filters["max_price"])
        and in_optional_range(product.review_count, filters["min_reviews"], filters["max_reviews"])
        and in_optional_range(product.monthly_bought, filters["min_bought"], filters["max_bought"])
        and in_optional_range(product.child_monthly_sales, filters["min_child_sales"], filters["max_child_sales"])
        and in_optional_range(product.bsr_rank, filters["min_bsr"], filters["max_bsr"])
        and launched_at_matches(product.launched_at, filters["launch_window"])
    )


def apply_product_filters(products: list[Product], filters: dict) -> list[Product]:
    return [product for product in products if product_matches_filters(product, filters)]


def product_plugin_data_complete(product: Product) -> bool:
    return product.plugin_data_loaded


def product_plugin_completeness(product: Product) -> int:
    values = (
        product.monthly_bought,
        product.child_monthly_sales,
        product.sales_amount,
        product.fba_fee,
        product.bsr_rank,
        product.sub_rank,
        product.seller_count,
        product.variant_count,
        product.brand,
        product.seller_name,
        product.prime_fba not in ("", "Unknown"),
        product.margin_rate,
        product.launched_at not in ("", "Unknown"),
        product.package_dimensions not in ("", "Unknown"),
        product.package_weight_lb,
    )
    return sum(bool(value) for value in values)


def merge_product_records(existing: Product, candidate: Product) -> Product:
    """Combine repeated collection attempts without losing enriched fields."""
    existing_score = product_plugin_completeness(existing)
    candidate_score = product_plugin_completeness(candidate)
    primary, secondary = (
        (candidate, existing)
        if candidate_score >= existing_score
        else (existing, candidate)
    )
    values = asdict(primary)
    secondary_values = asdict(secondary)
    for name, value in values.items():
        if value in ("", 0, 0.0, None, "Unknown"):
            replacement = secondary_values.get(name)
            if replacement not in ("", 0, 0.0, None, "Unknown"):
                values[name] = replacement
    merged = Product(**values)
    merged.plugin_data_loaded = existing.plugin_data_loaded or candidate.plugin_data_loaded
    if product_plugin_data_complete(merged):
        merged.status = "OK"
        merged.error = ""
    else:
        merged.status = "PARTIAL"
        merged.error = "卖家精灵字段未完整加载"
    return merged


def build_collection_summary(raw_products: list[Product], filtered_products: list[Product], filters: dict) -> str:
    if not raw_products:
        return "当前没有原始采集结果。请先点击“开始采集”，或载入上次采集结果。"
    removed_count = len(raw_products) - len(filtered_products)
    summary = (
        f"原始采集池 {len(raw_products)} 条，当前筛选保留 {len(filtered_products)} 条，"
        f"筛掉 {removed_count} 条。"
    )
    incomplete_count = sum(not product_plugin_data_complete(product) for product in raw_products)
    if incomplete_count:
        summary += f" 其中 {incomplete_count} 条卖家精灵字段未完整加载，缺失值不代表真实为 0。"
    rejection_lines = filter_rejection_summary(raw_products, filters)[:5]
    if rejection_lines:
        summary += " 主要筛选原因：" + "；".join(rejection_lines)
    return summary


def format_filter_value(value, money: bool = False) -> str:
    if value is None:
        return "不限"
    if money:
        return f"${float(value):g}"
    return f"{int(value):,}" if isinstance(value, int) or float(value).is_integer() else f"{float(value):g}"


def format_filter_range(label: str, min_value, max_value, money: bool = False) -> str:
    if min_value is None and max_value is None:
        return ""
    if min_value is not None and max_value is not None:
        return f"{label} {format_filter_value(min_value, money)} ~ {format_filter_value(max_value, money)}"
    if min_value is not None:
        return f"{label} >= {format_filter_value(min_value, money)}"
    return f"{label} <= {format_filter_value(max_value, money)}"


def build_filter_summary(filters: dict) -> str:
    parts = [
        format_filter_range("价格", filters.get("min_price"), filters.get("max_price"), money=True),
        format_filter_range("评分数", filters.get("min_reviews"), filters.get("max_reviews")),
        format_filter_range("月销量", filters.get("min_bought"), filters.get("max_bought")),
        format_filter_range("子体销量", filters.get("min_child_sales"), filters.get("max_child_sales")),
        format_filter_range("BSR", filters.get("min_bsr"), filters.get("max_bsr")),
    ]
    launch_window = filters.get("launch_window")
    if launch_window and launch_window not in ("不限", "Any"):
        parts.append(f"上架时间：{launch_window}")
    parts = [part for part in parts if part]
    return "当前筛选条件：" + ("；".join(parts) if parts else "不限")


def build_collection_plan_text(selected_paths: list[str], custom_url: str, batch_collect: bool, seed_urls: list[tuple[str, str]]) -> str:
    if custom_url:
        return "采集计划：使用自定义 Amazon 榜单链接，读取第 1 页和第 2 页；采完后先进入原始池，再应用当前筛选条件。"
    if not selected_paths:
        return "采集计划：暂未选择类目。请选择类目，或填写一个具体 Amazon 榜单链接。"
    selected_seed_paths = compact_category_paths(selected_paths)
    if batch_collect or len(selected_seed_paths) > 1 or selection_contains_parent_category(selected_seed_paths):
        count = len(seed_urls)
        if count:
            suffix = "入口较多，预计耗时较长；建议先选择更小的类目试跑。" if count >= 20 else "每个入口读取第 1 页和第 2 页。"
            return f"采集计划：批量采集 {count} 个已映射小类入口；{suffix} 原始产品去重后再应用筛选。"
        return "采集计划：开始采集时展开当前选择，并准备已映射的小类入口；每个入口读取第 1 页和第 2 页。"
    return "采集计划：只采集当前选择中的第一个具体榜单入口，读取第 1 页和第 2 页；不会自动跳到其它类目。"


def apply_filters_to_raw_pool(filters: dict) -> None:
    for product in st.session_state.raw_products:
        repair_product_price(product)
    selected_by_asin = {product.asin: product.selected for product in st.session_state.products}
    filtered_products = apply_product_filters(st.session_state.raw_products, filters)
    for product in filtered_products:
        product.selected = bool(selected_by_asin.get(product.asin, product.selected))
    sync_product_annotations(filtered_products)
    st.session_state.products = filtered_products
    st.session_state.last_collection_summary = build_collection_summary(
        st.session_state.raw_products,
        st.session_state.products,
        filters,
    )
    reset_result_page()


def stage_raw_products(products: list[Product]) -> None:
    if not products:
        return
    staged_by_asin = {
        product.asin: product
        for product in st.session_state.get("collection_staged_raw_products", [])
        if product.asin
    }
    for product in products:
        if not product.asin:
            continue
        product.selected = False
        existing = staged_by_asin.get(product.asin)
        staged_by_asin[product.asin] = merge_product_records(existing, product) if existing else product
    staged_products = list(staged_by_asin.values())
    st.session_state.collection_staged_raw_products = staged_products
    st.session_state.raw_products = staged_products
    st.session_state.last_raw_products_message = (
        f"采集中暂存：当前已拿到原始产品 {len(staged_products)} 条。"
        "如果中途停止，会保存这些已完成产品；采集完成后会再按当前筛选条件计算列表。"
    )


def filter_rejection_summary(products: list[Product], filters: dict) -> list[str]:
    reasons = {
        "卖家精灵字段未完整加载": 0,
        "价格低于最低值": 0,
        "价格高于最高值": 0,
        "评分数低于最低值": 0,
        "评分数高于最高值": 0,
        "月销量低于最低值": 0,
        "月销量高于最高值": 0,
        "子体销量不在范围": 0,
        "BSR 不在范围": 0,
        "上架时间不符合": 0,
    }
    for product in products:
        plugin_missing = not product_plugin_data_complete(product)
        plugin_filter_active = any(
            filters[key] is not None
            for key in ("min_bought", "max_bought", "min_child_sales", "max_child_sales", "min_bsr", "max_bsr")
        )
        if plugin_missing and plugin_filter_active:
            reasons["卖家精灵字段未完整加载"] += 1
        if filters["min_price"] is not None and product.price < filters["min_price"]:
            reasons["价格低于最低值"] += 1
        if filters["max_price"] is not None and product.price > filters["max_price"]:
            reasons["价格高于最高值"] += 1
        if filters["min_reviews"] is not None and product.review_count < filters["min_reviews"]:
            reasons["评分数低于最低值"] += 1
        if filters["max_reviews"] is not None and product.review_count > filters["max_reviews"]:
            reasons["评分数高于最高值"] += 1
        if not plugin_missing and filters["min_bought"] is not None and product.monthly_bought < filters["min_bought"]:
            reasons["月销量低于最低值"] += 1
        if not plugin_missing and filters["max_bought"] is not None and product.monthly_bought > filters["max_bought"]:
            reasons["月销量高于最高值"] += 1
        if not plugin_missing and not in_optional_range(product.child_monthly_sales, filters["min_child_sales"], filters["max_child_sales"]):
            reasons["子体销量不在范围"] += 1
        if not plugin_missing and not in_optional_range(product.bsr_rank, filters["min_bsr"], filters["max_bsr"]):
            reasons["BSR 不在范围"] += 1
        if not launched_at_matches(product.launched_at, filters["launch_window"]):
            reasons["上架时间不符合"] += 1
    return [f"{name}：{count} 条" for name, count in sorted(reasons.items(), key=lambda item: item[1], reverse=True) if count]


def render_range_filter(
    title: str,
    key_prefix: str,
    min_default: str = "",
    max_default: str = "",
    money: bool = False,
    disabled: bool = False,
):
    st.markdown(f"<div class='filter-label'>{escape(title)} <span>?</span></div>", unsafe_allow_html=True)
    st.markdown("<div class='range-filter-anchor'></div>", unsafe_allow_html=True)
    cols = st.columns([1, 0.18, 1], vertical_alignment="center")
    with cols[0]:
        min_value = st.text_input(
            f"{title} 最小值",
            value=min_default,
            placeholder="最小值",
            key=f"{key_prefix}_min",
            on_change=request_filter_auto_apply,
            label_visibility="collapsed",
            disabled=disabled,
        )
    with cols[1]:
        st.markdown("<div class='filter-range-sep'>~</div>", unsafe_allow_html=True)
    with cols[2]:
        max_value = st.text_input(
            f"{title} 最大值",
            value=max_default,
            placeholder="最大值",
            key=f"{key_prefix}_max",
            on_change=request_filter_auto_apply,
            label_visibility="collapsed",
            disabled=disabled,
        )
    if money:
        st.markdown("<div class='filter-money-hint'>$</div>", unsafe_allow_html=True)
    return min_value, max_value


def reset_filter_widgets() -> None:
    defaults = {
        "filter_price_min": "",
        "filter_price_max": "",
        "filter_reviews_min": "",
        "filter_reviews_max": "",
        "filter_monthly_sales_min": "",
        "filter_monthly_sales_max": "",
        "filter_child_sales_min": "",
        "filter_child_sales_max": "",
        "filter_bsr_min": "",
        "filter_bsr_max": "",
        "filter_launch_window": "不限" if st.session_state.get("ui_lang", "中文") == "中文" else "Any",
    }
    for key, value in defaults.items():
        st.session_state[key] = value


def fake_image(seed: str) -> str:
    colors = ["f4a261", "2a9d8f", "e76f51", "8ab17d", "577590", "e9c46a"]
    color = random.choice(colors)
    return f"https://placehold.co/320x320/{color}/ffffff?text={seed}"


def make_fake_product(category_path, list_type, rank, filters) -> Product:
    asin = "B0" + "".join(random.choices("ABCDEFGHJKLMNPQRSTUVWXYZ23456789", k=8))
    leaf = category_path.split(" > ")[-1]
    title = random.choice(
        [
            f"Upgraded {leaf} with Easy Clean Design",
            f"Portable {leaf} Kit for Small Spaces",
            f"Compact {leaf} Set for Families",
            f"Smart {leaf} Accessory for Daily Use",
            f"Foldable {leaf} Organizer with Travel Case",
        ]
    )
    brand = random.choice(["Novaly", "Petuno", "Auralis", "Mavento", "NorthPeak", "Amazon Basics"])
    price = round(random.uniform(12.99, 229.99), 2)
    rating = round(random.uniform(3.6, 4.9), 1)
    reviews = random.choice([12, 28, 43, 76, 128, 216, 304, 640, 1210])
    bought = random.choice([50, 100, 200, 300, 500, 1000, 2000])
    risks = detect_risk(title, brand)
    score, level, reason = score_product(
        rank,
        price,
        rating,
        reviews,
        bought,
        risks,
        filters["min_price"],
        filters["max_price"],
        filters["max_reviews"],
        filters["min_bought"],
    )
    return Product(
        selected=False,
        review_status="Pending",
        note="",
        list_type=list_type,
        category_path=category_path,
        rank=rank,
        title=title,
        asin=asin,
        amazon_url=f"https://www.amazon.com/dp/{asin}",
        image_url=fake_image(asin),
        price=price,
        rating=rating,
        review_count=reviews,
        monthly_bought=bought,
        brand=brand,
        prime_fba=random.choice(["Prime/FBA", "Prime", "Unknown"]),
        delivery=random.choice(["Free delivery", "Ships from Amazon", "Unknown"]),
        variant_count=random.randint(1, 8),
        launched_at=random.choice(["2026-04-18", "2026-02-09", "2025-11-22", "Unknown"]),
        package_dimensions=random.choice(["8.3 x 5.4 x 2.1 in", "12.0 x 9.2 x 4.8 in", "15.7 x 11.0 x 6.5 in"]),
        package_weight_lb=round(random.uniform(0.4, 8.5), 2),
        fba_fee=round(random.uniform(3.28, 18.75), 2),
        potential_score=score,
        potential_level=level,
        reason=reason,
        risk_tags=risks,
        scraped_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        status="OK",
        error="",
    )


def generate_products(list_type, selected_paths, pages, filters):
    products = []
    for path in selected_paths:
        log(f"Generate demo data: {list_type} / {path}")
        for page in range(1, pages + 1):
            for rank in range(1, 13):
                products.append(make_fake_product(path, list_type, (page - 1) * 12 + rank, filters))
    return products


def make_product_from_scraped(scraped, list_type, category_path, filters) -> Product:
    risks = detect_risk(scraped.title, scraped.brand)
    price = scraped.price or 0.0
    rating = scraped.rating or 0.0
    reviews = scraped.review_count or 0
    bought = scraped.monthly_bought or 0
    score, level, reason = score_product(
        scraped.rank,
        price,
        rating,
        reviews,
        bought,
        risks,
        filters["min_price"],
        filters["max_price"],
        filters["max_reviews"],
        filters["min_bought"],
    )
    return Product(
        selected=False,
        review_status="Pending",
        note="",
        list_type=list_type,
        category_path=category_path,
        rank=scraped.rank,
        title=scraped.title,
        asin=scraped.asin,
        amazon_url=scraped.amazon_url,
        image_url=scraped.image_url or fake_image(scraped.asin),
        price=price,
        rating=rating,
        review_count=reviews,
        monthly_bought=bought,
        brand=scraped.brand,
        prime_fba="Unknown",
        delivery="Unknown",
        variant_count=0,
        launched_at="Unknown",
        package_dimensions="Unknown",
        package_weight_lb=0.0,
        fba_fee=0.0,
        potential_score=score,
        potential_level=level,
        reason=reason,
        risk_tags=risks,
        scraped_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        status="OK",
        error="",
    )


def collect_real_products(list_type, url, filters) -> list[Product]:
    scraped_products = fetch_best_sellers(url or DEFAULT_TEST_URL, limit=30)
    return [
        make_product_from_scraped(
            scraped,
            list_type,
            "Pet Supplies > Dogs > Dog Feeding & Watering Supplies",
            filters,
        )
        for scraped in scraped_products
    ]


def export_rows_signature(products: list[Product]) -> tuple[tuple, ...]:
    fields = [field for field, _ in SELLERSPRITE_EXPORT_COLUMNS]
    rows = []
    for product in products:
        product_row = vars(product)
        row = {field: product_row.get(field, "") for field, _ in SELLERSPRITE_EXPORT_COLUMNS}
        row["image_preview_formula"] = image_formula(str(row.get("image_url", "")))
        rows.append(tuple(row.get(field, "") for field in fields))
    return tuple(rows)


def csv_bytes(products):
    return cached_csv_bytes(export_rows_signature(products))


@st.cache_data(show_spinner=False)
def cached_csv_bytes(rows: tuple[tuple, ...]):
    if not rows:
        return b""
    output = StringIO()
    headers = [header for _, header in SELLERSPRITE_EXPORT_COLUMNS]
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    for values in rows:
        writer.writerow({header: values[index] for index, (_, header) in enumerate(SELLERSPRITE_EXPORT_COLUMNS)})
    return output.getvalue().encode("utf-8-sig")


def table_rows(products: list[Product], display_start: int = 0) -> list[dict]:
    rows = []
    for index, product in enumerate(products, start=1):
        row = asdict(product)
        display_number = display_start + index
        row["rank"] = display_number
        rows.append({header: row.get(field, "") for field, header in SELLERSPRITE_TABLE_COLUMNS})
    return rows


def image_formula(image_url: str) -> str:
    if not image_url:
        return ""
    escaped_url = str(image_url).replace('"', '""')
    return f'=IMAGE("{escaped_url}","",3,50,50)'


def excel_bytes(products):
    return cached_excel_bytes(export_rows_signature(products))


@st.cache_data(show_spinner=False)
def cached_excel_bytes(rows: tuple[tuple, ...]):
    if not rows:
        return b""
    fields = [field for field, _ in SELLERSPRITE_EXPORT_COLUMNS]
    headers = [header for _, header in SELLERSPRITE_EXPORT_COLUMNS]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "卖家精灵采集"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="F6F7F9")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="5F6673")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_number, row_values in enumerate(rows, start=2):
        sheet.append(list(row_values))
        sheet.row_dimensions[row_number].height = 42

    width_by_field = {
        "image_preview_formula": 10,
        "rank": 6,
        "title": 34,
        "asin": 16,
        "brand": 14,
        "seller_name": 18,
        "bsr_category": 18,
        "sub_category": 22,
        "category_path": 42,
        "amazon_url": 36,
        "image_url": 36,
        "reason": 42,
        "risk_tags": 20,
        "package_dimensions": 28,
        "scraped_at": 20,
        "error": 24,
    }
    money_fields = {"sales_amount", "price", "fba_fee"}
    integer_fields = {"rank", "seller_count", "bsr_rank", "sub_rank", "monthly_bought", "variant_count", "review_count", "potential_score"}
    for column_index, (field, header) in enumerate(SELLERSPRITE_EXPORT_COLUMNS, start=1):
        width = width_by_field.get(field, min(max(len(header) + 4, 12), 20))
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = width
        for cell in sheet[letter][1:]:
            cell.alignment = Alignment(vertical="center", wrap_text=field in {"title", "reason", "category_path", "package_dimensions"})
            if field in money_fields:
                cell.number_format = '$#,##0.00'
            elif field in integer_fields:
                cell.number_format = '#,##0'
    for column_index, (field, _) in enumerate(SELLERSPRITE_EXPORT_COLUMNS, start=1):
        if field in {"amazon_url", "image_url"}:
            for row_number in range(2, len(rows) + 2):
                cell = sheet.cell(row=row_number, column=column_index)
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_identity(products: list[Product]) -> tuple[tuple, ...]:
    return tuple(
        (
            product.asin,
            product.scraped_at,
            product.price,
            product.review_count,
            product.monthly_bought,
            product.child_monthly_sales,
            product.bsr_rank,
            product.sales_amount,
        )
        for product in products
    )


def prepared_export_signature(products: list[Product], file_name: str, file_format: str) -> tuple:
    return (file_name, file_format, export_identity(products))


def export_bytes_for_format(products: list[Product], file_format: str) -> bytes:
    if file_format == "csv":
        return csv_bytes(products)
    return excel_bytes(products)


def export_mime_for_format(file_format: str) -> str:
    if file_format == "csv":
        return "text/csv"
    return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def render_lazy_export_button(
    container,
    label: str,
    products: list[Product],
    file_name: str,
    *,
    file_format: str = "xlsx",
    disabled: bool = False,
    use_container_width: bool = False,
    key: str,
) -> None:
    signature = prepared_export_signature(products, file_name, file_format)
    state_key = f"prepared_export_{key}"
    prepared = st.session_state.get(state_key)
    if prepared and prepared.get("signature") == signature:
        download_label = f"下载{label.replace('生成', '')}" if label.startswith("生成") else label.replace("Prepare", "Download")
        container.download_button(
            download_label,
            data=prepared["data"],
            file_name=file_name,
            mime=export_mime_for_format(file_format),
            use_container_width=use_container_width,
            key=f"download_export_{key}",
        )
    else:
        prepare_clicked = container.button(
            label,
            key=f"prepare_export_{key}",
            disabled=disabled,
            use_container_width=use_container_width,
        )
    if not (prepared and prepared.get("signature") == signature) and prepare_clicked:
        prepared = {
            "signature": signature,
            "data": export_bytes_for_format(products, file_format),
        }
        st.session_state[state_key] = prepared


def level_badge(level: str):
    colors = {
        "A": "#0f766e",
        "B": "#2563eb",
        "C": "#a16207",
        "D": "#6b7280",
        "Risk": "#b91c1c",
    }
    return f"<span style='background:{colors.get(level, '#6b7280')}; color:white; padding:3px 8px; border-radius:6px; font-size:12px;'>{level}</span>"


def render_product_identity(product: Product):
    return None


def category_label(name: str, node: dict, indent: int = 0) -> str:
    zh = f" ({node.get('zh', '')})" if node.get("zh") else ""
    return f"{name}{zh}"


def leaf_label(name: str, node: dict, indent: int = 0) -> str:
    prefix = "- " if indent else ""
    return prefix + category_label(name, node, indent)


def category_matches_filter(name: str, node: dict, query: str) -> bool:
    if not query:
        return True
    haystack = f"{name} {node.get('zh', '')}".lower()
    if query in haystack:
        return True
    return any(category_matches_filter(child_name, child_node, query) for child_name, child_node in node.get("children", {}).items())


def category_widget_key(prefix: str, path: str) -> str:
    return f"{prefix}_{re.sub(r'[^A-Za-z0-9]+', '_', path).strip('_')}"


def handle_category_select_all_change():
    selected = bool(st.session_state.get("category_select_all", False))
    category_tree = display_categories()
    st.session_state.category_dialog_selected_paths = list(category_tree) if selected else []


def category_path_parts(path: str) -> list[str]:
    return [part.strip() for part in path.split(">") if part.strip()]


def compact_category_paths(paths: list[str]) -> list[str]:
    unique_paths = list(dict.fromkeys(path for path in paths if path))
    order_by_path = {path: index for index, path in enumerate(unique_paths)}
    ordered_paths = sorted(unique_paths, key=lambda path: (path.count(" > "), order_by_path[path]))
    compact: list[str] = []
    compact_set: set[str] = set()
    for path in ordered_paths:
        parts = path.split(" > ")
        if any(" > ".join(parts[:depth]) in compact_set for depth in range(1, len(parts))):
            continue
        compact.append(path)
        compact_set.add(path)
    return compact


def get_category_node(path: str, tree: dict | None = None) -> dict:
    node = None
    current = display_categories() if tree is None else tree
    for part in category_path_parts(path):
        node = current.get(part, {})
        current = node.get("children", {}) if isinstance(node, dict) else {}
    return node if isinstance(node, dict) else {}


def iter_child_paths(path: str, node: dict):
    for child_name, child_node in node.get("children", {}).items():
        child_path = f"{path} > {child_name}"
        yield child_path, child_node


def iter_category_leaf_paths(path: str, node: dict):
    children = list(iter_child_paths(path, node))
    if not children:
        yield path
        return
    for child_path, child_node in children:
        yield from iter_category_leaf_paths(child_path, child_node)


def handle_category_row_select_change(path: str, node: dict):
    selected = bool(st.session_state.get(category_widget_key("cat_sel", path), False))
    st.session_state.category_dialog_selected_paths = toggle_compact_category_selection(
        display_categories(),
        st.session_state.get("category_dialog_selected_paths", []),
        path,
        selected,
    )


def selected_category_paths_from_state(tree: dict | None = None, prefix: str = "", seen: set[int] | None = None) -> list[str]:
    return compact_category_paths(st.session_state.get("category_dialog_selected_paths", []))


def render_category_row(path: str, name: str, node: dict, depth: int, selected_paths: list[str], all_categories: bool, confirmed: set[str], query: str):
    children = node.get("children", {})
    has_children = bool(children)
    expand_key = category_widget_key("cat_exp", path)
    select_key = category_widget_key("cat_sel", path)
    if expand_key not in st.session_state:
        st.session_state[expand_key] = False
    selected = category_path_selected(
        path,
        st.session_state.get("category_dialog_selected_paths", []),
    )
    st.session_state[select_key] = selected

    indent = max(0.001, min(depth, 4) * 0.055)
    label_width = max(0.52, 0.72 - indent)
    cols = st.columns([indent, 0.04, 0.045, label_width, 0.19], gap=None, vertical_alignment="center")
    with cols[0]:
        if depth:
            st.markdown(f"<span class='category-tree-indent depth-{depth}'></span>", unsafe_allow_html=True)
        else:
            st.markdown("<span class='category-tree-root-indent'></span>", unsafe_allow_html=True)
    with cols[1]:
        if has_children:
            arrow = "▾" if st.session_state[expand_key] else "▸"
            if st.button(arrow, key=f"toggle_{expand_key}", use_container_width=True):
                st.session_state[expand_key] = not st.session_state[expand_key]
        else:
            st.markdown("<span class='category-tree-spacer'></span>", unsafe_allow_html=True)
    with cols[2]:
        selected = st.checkbox(
            f"选择 {path}",
            key=select_key,
            on_change=handle_category_row_select_change,
            args=(path, node),
            label_visibility="collapsed",
        )
    with cols[3]:
        label = category_label(name, node, depth)
        st.markdown(f"<div class='category-tree-label'>{escape(label)}</div>", unsafe_allow_html=True)
    with cols[4]:
        count = int(node.get("count") or 0)
        if count:
            st.markdown(f"<div class='category-count-badge'>{count:,}</div>", unsafe_allow_html=True)
        else:
            st.markdown("<div></div>", unsafe_allow_html=True)

    if selected:
        selected_paths.append(path)

    if has_children and (st.session_state[expand_key] or query):
        for child_name, child_node in children.items():
            if query and not category_matches_filter(child_name, child_node, query):
                continue
            render_category_row(
                f"{path} > {child_name}",
                child_name,
                child_node,
                depth + 1,
                selected_paths,
                all_categories,
                confirmed,
                query,
            )


def render_category_tree(query: str = ""):
    selected_paths = []
    category_tree = display_categories()
    root_paths = list(category_tree)
    st.session_state["category_select_all"] = bool(root_paths) and all(
        category_path_selected(path, st.session_state.get("category_dialog_selected_paths", []))
        for path in root_paths
    )
    all_categories = st.checkbox(
        T["all_categories"],
        key="category_select_all",
        on_change=handle_category_select_all_change,
    )
    confirmed = set(st.session_state.confirmed_category_paths)
    for main, main_node in category_tree.items():
        if not category_matches_filter(main, main_node, query):
            continue
        render_category_row(main, main, main_node, 0, selected_paths, all_categories, confirmed, query)
    return selected_category_paths_from_state(category_tree)


@st.dialog("选择类目", width="large", on_dismiss=close_category_dialog_state)
def render_category_dialog():
    if st.session_state.get("category_clear_requested"):
        st.session_state.confirmed_category_paths = []
        st.session_state.category_dialog_selected_paths = []
        for key in list(st.session_state.keys()):
            if key.startswith(("main_", "mid_", "leaf_", "cat_sel_", "cat_exp_")) or key == "category_select_all":
                del st.session_state[key]
        st.session_state.category_clear_requested = False

    st.markdown("<span class='category-dialog-grid-anchor'></span>", unsafe_allow_html=True)
    left_panel, right_panel = st.columns([0.58, 0.42], gap="medium", vertical_alignment="top")
    with left_panel:
        query = st.text_input(
            "搜索类目",
            key="category_search",
            placeholder="请输入 Node ID / 类目关键词",
            label_visibility="collapsed",
        ).strip().lower()
        with st.container(border=False):
            st.markdown("<span class='category-tree-scroll-anchor'></span>", unsafe_allow_html=True)
            selected_paths = render_category_tree(query)
    with right_panel:
        title_col, clear_col = st.columns([0.72, 0.28], vertical_alignment="center")
        title_col.markdown(f"<div class='category-selected-title'>已选（{len(selected_paths)}）</div>", unsafe_allow_html=True)
        clear_clicked = clear_col.button("清空", use_container_width=True)
        if selected_paths:
            selected_html = "".join(f"<span class='selected-pill'>{escape(path)}</span>" for path in selected_paths[:24])
            if len(selected_paths) > 24:
                selected_html += f"<span class='selected-pill'>+{len(selected_paths) - 24}</span>"
            st.markdown(f"<div class='category-selected-panel'>{selected_html}</div>", unsafe_allow_html=True)
        else:
            st.markdown("<div class='category-selected-empty'>暂未选择类目</div>", unsafe_allow_html=True)
        if clear_clicked:
            st.session_state.category_clear_requested = True
            st.rerun()
    st.markdown("<span class='category-footer-anchor'></span>", unsafe_allow_html=True)
    spacer_left, cancel_col, spacer_mid, confirm_col = st.columns([0.72, 0.10, 0.02, 0.16])
    if cancel_col.button("取消", use_container_width=True):
        st.session_state.category_dialog_selected_paths = list(st.session_state.confirmed_category_paths)
        st.session_state.show_category_dialog = False
        st.rerun()
    if confirm_col.button("确认选择", type="primary", use_container_width=True):
        st.session_state.confirmed_category_paths = compact_category_paths(st.session_state.category_dialog_selected_paths)
        st.session_state.show_category_dialog = False
        st.rerun()


def _display_dash(value, suffix=""):
    if value in (None, "", "Unknown"):
        return "-"
    return f"{value}{suffix}"


def _display_money(value):
    return "$0.00" if not value else f"${value:,.2f}"


def _display_int(value, suffix=""):
    return "0" if not value else f"{value:,}{suffix}"


def _copy_js(value: str) -> str:
    payload = json.dumps(value or "")
    return (
        f"navigator.clipboard.writeText({payload}).then(()=>{{"
        "this.textContent='✓';this.classList.add('copied');"
        "setTimeout(()=>{this.textContent='⧉';this.classList.remove('copied');},1200);"
        "});"
    )


def render_clipboard_bridge():
    components.html(
        """
        <script>
        window.parent.eval(`
          (() => {
            document.querySelectorAll('.copy-icon[data-copy-value]').forEach((button) => {
              if (button.dataset.copyBound === '1') return;
              button.dataset.copyBound = '1';
              button.addEventListener('click', async (event) => {
                event.preventDefault();
                event.stopPropagation();
                const value = button.dataset.copyValue || '';
                if (!value) return;
                const oldText = button.textContent;
                const oldTitle = button.title;
                button.textContent = '✓';
                button.classList.add('copied');
                button.title = '已复制';
                setTimeout(() => {
                  button.textContent = oldText || '⧉';
                  button.classList.remove('copied');
                  button.title = oldTitle || '复制';
                }, 1200);
                try {
                  await navigator.clipboard.writeText(value);
                } catch (error) {
                  const input = document.createElement('textarea');
                  input.value = value;
                  input.style.position = 'fixed';
                  input.style.left = '-9999px';
                  document.body.appendChild(input);
                  input.select();
                  document.execCommand('copy');
                  input.remove();
                }
              });
            });
          })();
        `);
        </script>
        """,
        height=0,
    )


def seller_product_html(product: Product, display_number: int | None = None) -> str:
    title = escape(product.title)
    asin = escape(product.asin)
    brand = escape(product.brand or "-")
    image_url = escape(product.image_url or fake_image(product.asin))
    amazon_url = escape(product.amazon_url or f"https://www.amazon.com/dp/{product.asin}")
    category_path = escape(product.category_path)
    leaf_category = escape(product.category_path.split(" > ")[-1])
    monthly_bought = _display_int(product.monthly_bought, "+")
    sales_amount_value = getattr(product, "sales_amount", 0)
    sales_amount = "-" if not sales_amount_value else f"${sales_amount_value:,.0f}"
    child_sales_label = getattr(product, "child_monthly_sales_label", "")
    child_sales = child_sales_label or _display_int(getattr(product, "child_monthly_sales", 0), "+")
    bsr_rank = getattr(product, "bsr_rank", 0) or product.rank
    sub_rank = getattr(product, "sub_rank", 0)
    bsr_category = getattr(product, "bsr_category", "")
    sub_category = getattr(product, "sub_category", "")
    qa_count = "-" if not product.review_count else f"{max(12, product.review_count // 3):,}"
    review_new = "-" if not product.review_count else f"{max(8, product.review_count // 4):,}"
    margin = product.margin_rate or ("0" if not product.fba_fee else f"{min(45, max(12, int(product.potential_score / 3)))}%")
    delivery = product.delivery or product.prime_fba or "0"
    package_weight = _display_dash(product.package_weight_lb, " pounds")
    package_dimensions = _display_dash(product.package_dimensions)
    row_number = display_number if display_number is not None else product.rank
    note_preview = escape(product.note or "未备注")
    select_class = " is-selected" if product.selected else ""
    favorite_class = " is-favorite" if product.asin in st.session_state.get("favorite_products", {}) else ""
    favorite_label = "♥" if favorite_class else "♡"
    note_mode = "edit" if st.session_state.get(f"product_note_editing_{product.asin}") else "display"
    return f"""
    <div class="seller-row">
        <div class="list-select-host" data-asin="{asin}">
            <button type="button" class="list-select-proxy{select_class}" data-asin="{asin}" title="选择产品"></button>
        </div>
        <div class="seller-main">
            <div class="seller-rank">{row_number}</div>
            <div class="seller-product">
                <div class="seller-image-wrap">
                    <span class="level-corner">{escape(product.potential_level)}</span>
                    <img src="{image_url}" alt="{title}" />
                </div>
                <div class="seller-info">
                    <div class="seller-title">{title}</div>
                    <div class="meta-line">ASIN: <strong>{asin}</strong><button type="button" class="copy-icon" data-copy-value="{asin}" title="复制 ASIN">⧉</button><a class="mini-link" href="{amazon_url}" target="_blank" rel="noopener noreferrer" title="打开 Amazon DP 链接">↗</a></div>
                    <div class="meta-line muted">父ASIN：- <button type="button" class="copy-icon disabled-icon" title="暂无父ASIN">⧉</button><a class="mini-link" href="{amazon_url}" target="_blank" rel="noopener noreferrer" title="打开 Amazon DP 链接">↗</a></div>
                    <div class="meta-line">品牌:<strong>{brand}</strong><button type="button" class="copy-icon" data-copy-value="{brand}" title="复制品牌">⧉</button><a class="mini-link" href="{amazon_url}" target="_blank" rel="noopener noreferrer" title="打开 Amazon DP 链接">↗</a></div>
                </div>
            </div>
            <div class="cell bsr-cell"><strong>{_display_int(bsr_rank)}</strong><span class="green">↑ -</span><span class="green">↑ -</span></div>
            <div class="trend-cell"><svg viewBox="0 0 150 64" preserveAspectRatio="none"><path d="M0 42 L18 42 L32 36 L44 22 L58 34 L72 24 L88 39 L104 32 L122 20 L150 24" fill="none" stroke="#ff8a1f" stroke-width="2"/><path d="M0 42 L18 42 L32 36 L44 22 L58 34 L72 24 L88 39 L104 32 L122 20 L150 24 L150 64 L0 64 Z" fill="#fff1df"/></svg></div>
            <div class="cell"><strong>{monthly_bought}</strong><span class="muted">0</span></div>
            <div class="cell"><strong>{sales_amount}</strong></div>
            <div class="cell"><strong>{child_sales}</strong><span class="muted">0</span></div>
            <div class="cell"><strong>{_display_dash(product.variant_count)}</strong></div>
            <div class="cell"><strong>{_display_money(product.price)}</strong><span class="muted">{qa_count}</span></div>
            <div class="cell"><strong>{_display_int(product.review_count)}</strong><span class="muted">{review_new}</span></div>
            <div class="cell"><strong>{_display_dash(product.rating)}</strong><span class="muted">-</span></div>
            <div class="cell"><strong>{_display_money(product.fba_fee)}</strong><span class="muted">{margin}</span></div>
            <div class="cell"><strong>{escape(str(_display_dash(product.launched_at)))}</strong><span class="muted">-</span></div>
            <div class="cell"><strong>{delivery}</strong><span class="muted">-</span></div>
            <div class="ops-cell">
                <div class="list-favorite-host" data-asin="{asin}">
                    <button type="button" class="list-favorite-proxy{favorite_class}" data-asin="{asin}" title="收藏">{favorite_label}</button>
                </div>
            </div>
        </div>
        <div class="seller-detail">
            <div>浏览同类目: <span class="orange">{escape(bsr_category or category_path)}</span> <span class="pill orange-pill">BS榜单</span> <span class="pill orange-pill">新品榜</span> <span class="pill orange-pill">市场分析</span> <span class="pill orange-pill">找相似</span></div>
            <div>中文类目名: - <span class="rank-pill">#{_display_int(sub_rank) if sub_rank else 1}</span> in {escape(sub_category or leaf_category)}</div>
            <div>LQS: <strong>0</strong>　卖家: <strong>{escape(product.seller_name or '0')}</strong>　BuyBox卖家: <strong>{escape(product.seller_name or '0')}</strong>　商品重量: <strong>{package_weight}</strong>　商品尺寸: <strong>{escape(str(package_dimensions))}</strong>　包装重量: <strong>{package_weight}</strong>　包装尺寸: <strong>{escape(str(package_dimensions))}</strong></div>
            <div class="seller-note-preview list-note-host list-note-text-button" data-asin="{asin}" data-note-mode="{note_mode}">
                <button type="button" class="list-note-proxy" data-asin="{asin}" title="点击编辑备注">备注：{note_preview}</button>
            </div>
        </div>
    </div>
    """


def product_tile_html(product: Product) -> str:
    title = escape(product.title)
    asin = escape(product.asin)
    brand = escape(product.brand or "-")
    seller_name = escape(product.seller_name or product.brand or "-")
    image_url = escape(product.image_url or fake_image(product.asin))
    amazon_url = escape(product.amazon_url or f"https://www.amazon.com/dp/{product.asin}")
    bsr_rank = getattr(product, "bsr_rank", 0) or product.rank
    sub_rank = getattr(product, "sub_rank", 0)
    bsr_category = escape(getattr(product, "bsr_category", "") or product.category_path.split(" > ")[0])
    sub_category = escape(getattr(product, "sub_category", "") or product.category_path.split(" > ")[-1])
    monthly_bought = _display_int(product.monthly_bought, "+")
    child_sales_label = getattr(product, "child_monthly_sales_label", "")
    child_sales = child_sales_label or _display_int(getattr(product, "child_monthly_sales", 0), "+")
    sales_amount_value = getattr(product, "sales_amount", 0)
    sales_amount = "-" if not sales_amount_value else f"${sales_amount_value:,.0f}"
    delivery = escape(product.delivery or product.prime_fba or "-")
    seller_count = _display_int(getattr(product, "seller_count", 0))
    rating = "-" if not product.rating else f"{product.rating:.1f}"
    review_count = "-" if not product.review_count else f"{product.review_count:,}"
    rating_review_label = f"{rating}/{review_count}"
    launched_at = escape(str(_display_dash(product.launched_at)))
    note_preview = escape(product.note or "未备注")
    return f"""
    <div class="tile-card">
        <div class="tile-image-wrap">
            <img src="{image_url}" alt="{title}" />
        </div>
        <div class="tile-title" title="{title}">{title}</div>
        <div class="tile-asin">ASIN: <strong>{asin}</strong><button type="button" class="copy-icon" data-copy-value="{asin}" title="复制 ASIN">⧉</button><a class="mini-link" href="{amazon_url}" target="_blank" rel="noopener noreferrer" title="打开 Amazon DP 链接">↗</a></div>
        <div class="tile-line">卖家: <strong>{seller_name}</strong> <span class="tile-fulfillment">{delivery}</span><span class="tile-seller-count">卖家: {seller_count}</span></div>
        <div class="tile-line">品牌: <strong>{brand}</strong><button type="button" class="copy-icon" data-copy-value="{brand}" title="复制品牌">⧉</button></div>
        <div class="tile-rank-block">
            <div><span class="tile-rank-pill">#{_display_int(bsr_rank)}</span> in {bsr_category}</div>
            <div><span class="tile-rank-pill">#{_display_int(sub_rank) if sub_rank else "-"}</span> in {sub_category}</div>
        </div>
        <div class="tile-stats">
            <div>销量(父): <strong>{monthly_bought}</strong></div>
            <div>子体销量: <strong>{child_sales}</strong></div>
            <div>销售额: <strong>{sales_amount}</strong></div>
            <div>变体数: <strong>{_display_dash(product.variant_count)}</strong></div>
            <div>价格: <strong>{_display_money(product.price)}</strong></div>
            <div>评分/评分数: <strong>{rating_review_label}</strong></div>
            <div class="tile-wide">上架时间: <strong>{launched_at}</strong></div>
        </div>
        <div class="tile-note-preview">备注：{note_preview}</div>
    </div>
    """


def render_tile_cards(products):
    st.markdown("<div class='tile-grid-frame'>", unsafe_allow_html=True)
    for start in range(0, len(products), 4):
        columns = st.columns(4, gap="medium")
        for column, product in zip(columns, products[start:start + 4]):
            with column:
                st.markdown("<span class='tile-card-select-anchor'></span>", unsafe_allow_html=True)
                st.checkbox(
                    "选择产品",
                    key=f"row_include_{product.asin}",
                    value=product.selected,
                    label_visibility="collapsed",
                )
                st.markdown(product_tile_html(product), unsafe_allow_html=True)
                render_product_annotation_controls(product, "tile")
    st.markdown("</div>", unsafe_allow_html=True)


def render_favorites_panel() -> None:
    favorites = st.session_state.get("favorite_products", {})
    if not favorites:
        st.info("暂未收藏产品。")
        return
    favorite_products = [
        product_from_dict(record.get("product", {}))
        for record in favorites.values()
        if isinstance(record, dict) and isinstance(record.get("product", {}), dict)
    ]
    sync_product_annotations(favorite_products)
    st.caption(f"已收藏 {len(favorite_products)} 个 ASIN，收藏记录按 ASIN 去重，备注会在不同采集批次之间保留。")
    for start in range(0, len(favorite_products), 4):
        columns = st.columns(4, gap="medium")
        for column, product in zip(columns, favorite_products[start:start + 4]):
            with column:
                st.markdown(product_tile_html(product), unsafe_allow_html=True)
                render_product_annotation_controls(product, "favorite")


def collect_sellersprite_products(list_type, filters, category_path: str = "") -> list[Product]:
    scraped_products = load_cached_sellersprite_products(SELLERSPRITE_DOM_CACHE, limit=200)
    image_cache = load_sellersprite_image_cache()
    products = []
    for scraped in scraped_products:
        plugin_complete = sellersprite_product_hydrated(scraped)
        risks = detect_risk(scraped.title, scraped.brand)
        score, level, reason = score_product(
            scraped.rank,
            scraped.price,
            scraped.rating,
            scraped.review_count,
            scraped.parent_monthly_sales,
            risks,
            filters["min_price"],
            filters["max_price"],
            filters["max_reviews"],
            filters["min_bought"],
        )
        product = Product(
            selected=False,
            review_status="待审核",
            note="",
            list_type=list_type,
            category_path=scraped.sub_category or scraped.bsr_category or category_path or "未识别类目",
            rank=scraped.rank,
            title=scraped.title,
            asin=scraped.asin,
            amazon_url=f"https://www.amazon.com/dp/{scraped.asin}",
            image_url=image_cache.get(scraped.asin, fake_image(scraped.asin)),
            price=scraped.price,
            rating=scraped.rating,
            review_count=scraped.review_count,
            monthly_bought=scraped.parent_monthly_sales,
            brand=scraped.brand,
            prime_fba=scraped.fulfillment or "Unknown",
            delivery=scraped.fulfillment or "Unknown",
            variant_count=scraped.variant_count,
            launched_at=scraped.launched_at or "Unknown",
            package_dimensions=scraped.package_dimensions or "Unknown",
            package_weight_lb=scraped.package_weight_lb,
            fba_fee=scraped.fba_fee,
            potential_score=score,
            potential_level=level,
            reason=reason,
            risk_tags=risks,
            scraped_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            status="OK" if plugin_complete else "PARTIAL",
            error="" if plugin_complete else "卖家精灵字段未完整加载",
            seller_name=scraped.seller,
            seller_count=scraped.seller_count,
            sales_amount=scraped.sales_amount,
            child_monthly_sales=scraped.child_monthly_sales,
            child_monthly_sales_label=scraped.child_monthly_sales_label,
            bsr_rank=scraped.bsr_rank,
            bsr_category=scraped.bsr_category,
            sub_rank=scraped.sub_rank,
            sub_category=scraped.sub_category,
            margin_rate=scraped.margin_rate,
            plugin_data_loaded=plugin_complete,
        )
        products.append(product)
    return products


def collect_sellersprite_entry(
    target_url: str,
    list_type: str,
    filters: dict,
    progress=None,
    page_count: int = 2,
    progress_label: str = "",
    category_path: str = "",
) -> tuple[list[Product], list]:
    products_by_asin: dict[str, Product] = {}
    duplicate_pages = 0
    refresh_results = []

    def update_progress(percent: int, message: str):
        raise_if_stop_requested()
        if progress:
            progress(percent, f"{message}｜当前入口已去重 {len(products_by_asin)} 条")

    def collect_page_products(page: int, refresh_result):
        nonlocal duplicate_pages
        raise_if_stop_requested()
        refresh_results.append(refresh_result)
        if not refresh_result.ok:
            cached_text = ""
            if SELLERSPRITE_DOM_CACHE.exists():
                cached_text = SELLERSPRITE_DOM_CACHE.read_text(encoding="utf-8", errors="replace")
            if "there are no hot new releases available in this category" in cached_text.lower():
                refresh_result.ok = True
                refresh_result.message = EMPTY_NEW_RELEASES_MESSAGE
            elif refresh_result.message != SELLERSPRITE_INCOMPLETE_MESSAGE:
                raise RuntimeError(refresh_result.message)
        page_name = "第一页" if page == 1 else "第二页" if page == 2 else f"第 {page} 页"
        parsed_products = collect_sellersprite_products(
            list_type,
            filters,
            category_path or progress_label,
        )
        before_count = len(products_by_asin)
        for product in parsed_products:
            existing = products_by_asin.get(product.asin)
            products_by_asin[product.asin] = merge_product_records(existing, product) if existing else product
        added_count = len(products_by_asin) - before_count
        stage_raw_products(list(products_by_asin.values()))
        if progress:
            progress(
                99,
                f"{page_name}转成产品卡：本页 {len(parsed_products)} 条，新增 ASIN {added_count} 条，当前入口已去重 {len(products_by_asin)} 条",
            )
        if page > 1 and parsed_products and added_count == 0:
            duplicate_pages += 1
            log(f"{progress_label or target_url}: {page_name} parsed {len(parsed_products)} rows but added 0 unique ASINs. Pagination may be duplicated.")

    refresh_sellersprite_cache_pages(
        target_url,
        SELLERSPRITE_DOM_CACHE,
        SELLERSPRITE_IMAGE_CACHE,
        SELLERSPRITE_META_CACHE,
        expected_products=SELLERSPRITE_EXPECTED_PRODUCTS_PER_PAGE,
        page_count=page_count,
        progress=update_progress,
        page_callback=collect_page_products,
        stop_check=raise_if_stop_requested,
    )
    if duplicate_pages:
        log(f"{progress_label or target_url}: {duplicate_pages} duplicated page(s) detected during collection.")
    return list(products_by_asin.values()), refresh_results


def sellersprite_collection_quality(
    products: list[Product],
    refresh_results: list,
    list_type: str,
) -> tuple[bool, str]:
    return evaluate_sellersprite_collection_quality(
        len(products),
        refresh_results,
        empty_message=EMPTY_NEW_RELEASES_MESSAGE,
        expected_products_per_page=SELLERSPRITE_EXPECTED_PRODUCTS_PER_PAGE,
        min_products_per_page=SELLERSPRITE_MIN_PRODUCTS_PER_PAGE,
        min_products_two_pages=SELLERSPRITE_MIN_PRODUCTS_TWO_PAGES,
        list_type=list_type,
    )


def collect_sellersprite_entry_with_quality_retry(
    target_url: str,
    list_type: str,
    filters: dict,
    progress=None,
    page_count: int = 2,
    progress_label: str = "",
    category_path: str = "",
    retry_limit: int = SELLERSPRITE_CATEGORY_RETRY_LIMIT,
) -> tuple[list[Product], list, bool, str]:
    stable_category_path = category_path or progress_label
    best_products, best_results = collect_sellersprite_entry(
        target_url,
        list_type,
        filters,
        progress=progress,
        page_count=page_count,
        progress_label=progress_label,
        category_path=stable_category_path,
    )
    best_ok, best_message = sellersprite_collection_quality(best_products, best_results, list_type)
    label = progress_label or target_url
    retry_index = 0
    while not best_ok and retry_index < retry_limit:
        raise_if_stop_requested()
        retry_index += 1
        if progress:
            progress(99, f"{best_message}。正在自动补采第 {retry_index}/{retry_limit} 次。")
        log(f"{label}: {best_message}. Retrying collection {retry_index}/{retry_limit}.")
        retry_products, retry_results = collect_sellersprite_entry(
            target_url,
            list_type,
            filters,
            progress=progress,
            page_count=page_count,
            progress_label=f"{label} 重试{retry_index}",
            category_path=stable_category_path,
        )
        retry_ok, retry_message = sellersprite_collection_quality(retry_products, retry_results, list_type)
        merged_by_asin = {product.asin: product for product in best_products if product.asin}
        for product in retry_products:
            existing = merged_by_asin.get(product.asin)
            merged_by_asin[product.asin] = merge_product_records(existing, product) if existing else product
        best_products = list(merged_by_asin.values())
        best_result_score = (
            sum(int(getattr(result, "hydrated_count", 0) or 0) for result in best_results),
            sum(int(getattr(result, "product_count", 0) or 0) for result in best_results),
        )
        retry_result_score = (
            sum(int(getattr(result, "hydrated_count", 0) or 0) for result in retry_results),
            sum(int(getattr(result, "product_count", 0) or 0) for result in retry_results),
        )
        if retry_ok or retry_result_score > best_result_score:
            best_results = retry_results
        best_ok, best_message = sellersprite_collection_quality(best_products, best_results, list_type)
        log(f"{label}: retry {retry_index}/{retry_limit} result: {retry_message}. best: {best_message}.")
    if progress and not best_ok:
        progress(99, f"{best_message}。已保留当前能解析到的产品。")
    return best_products, best_results, best_ok, best_message


def collect_sellersprite_batch(
    seed_url: str,
    list_type: str,
    filters: dict,
    progress_bar,
    seed_label: str = "",
    progress_start: int = 0,
    progress_end: int = 100,
    progress_prefix: str = "",
    total_status_placeholder=None,
) -> tuple[list[Product], bool, str, int]:
    progress_span = max(1, progress_end - progress_start)

    def set_batch_progress(local_percent: int, text: str):
        raise_if_stop_requested()
        global_percent = min(100, progress_start + int((max(0, min(100, local_percent)) / 100) * progress_span))
        progress_bar.progress(global_percent, text=text)

    if not is_rank_category_url(seed_url):
        raise ValueError("当前链接不是具体榜单类目页。为避免自动跳类目，本次不会打开父类页继续发现链接。")

    set_batch_progress(0, "当前入口开始采集：只打开已映射的具体 Amazon 榜单页。")

    def update_entry_progress(percent: int, message: str):
        st.session_state.collection_total_raw_count = len(
            st.session_state.get("collection_staged_raw_products", [])
        )
        update_collection_total_status(total_status_placeholder)
        set_batch_progress(percent, message)

    products, refresh_results, quality_ok, quality_message = collect_sellersprite_entry_with_quality_retry(
        seed_url,
        list_type,
        filters,
        progress=update_entry_progress,
        page_count=2,
        progress_label=seed_label or "当前类目",
        category_path=seed_label,
    )
    if not quality_ok:
        set_batch_progress(99, f"{quality_message}。疑似漏采，已保留当前能解析到的产品。")
        if is_sellersprite_load_failure(quality_message):
            raise SellerSpriteLoadTimeout(
                f"当前入口卖家精灵加载超时：{quality_message}",
                products,
                len(refresh_results),
            )
    completed_pages = completed_collection_page_count(refresh_results)
    set_batch_progress(100, f"当前入口完成：读取 {completed_pages} 页，原始去重 {len(products)} 条。")
    return products, quality_ok, quality_message, completed_pages


def resolve_category_seed_urls(selected_paths: list[str], custom_url: str = "") -> list[tuple[str, str]]:
    seeds: list[tuple[str, str]] = []
    seen: set[str] = set()
    for path in collection_seed_paths(selected_paths):
        url = find_exact_category_url(path)
        seed_label = path
        if not url or url in seen:
            continue
        label = path if seed_label == path else f"{path}（通过 {seed_label} 自动发现）"
        seeds.append((label, url))
        seen.add(url)
    if not seeds and custom_url:
        seeds.append(("自定义链接", custom_url))
    return seeds


def collection_seed_paths(selected_paths: list[str]) -> list[str]:
    selected_paths = compact_category_paths(selected_paths)
    leaf_paths: list[str] = []
    seen: set[str] = set()
    for path in selected_paths:
        node = get_category_node(path)
        expanded_paths = list(iter_category_leaf_paths(path, node)) if node else [path]
        for leaf_path in expanded_paths:
            if leaf_path not in seen:
                leaf_paths.append(leaf_path)
                seen.add(leaf_path)
    return leaf_paths or selected_paths


def selection_contains_parent_category(selected_paths: list[str]) -> bool:
    selected_set = set(selected_paths)
    has_selected_descendants = any(
        any(" > ".join(path.split(" > ")[:depth]) in selected_set for depth in range(1, len(path.split(" > "))))
        for path in selected_set
    )
    category_tree = display_categories()
    has_children = any(bool(get_category_node(path, category_tree).get("children")) for path in selected_paths)
    return has_selected_descendants or has_children


def resolve_primary_collection_url(selected_paths: list[str], custom_url: str = "") -> tuple[str, str]:
    if custom_url:
        return "自定义链接", custom_url
    seeds = resolve_category_seed_urls(selected_paths, "")
    if seeds:
        return seeds[0]
    return "默认测试链接", DEFAULT_TEST_URL


def collect_sellersprite_batch_from_seeds(
    seed_urls: list[tuple[str, str]],
    list_type: str,
    filters: dict,
    progress_bar,
    total_status_placeholder=None,
) -> list[Product]:
    if not seed_urls:
        raise ValueError("没有可采集的类目链接。请先选择已映射的类目，或填写自定义 Amazon Best Sellers 链接。")
    raw_by_asin: dict[str, Product] = {}
    seed_summaries: list[dict] = []
    deferred_seeds = CollectionRetryQueue()

    def merge_seed_products(products: list[Product]) -> int:
        added = 0
        for product in products:
            existing = raw_by_asin.get(product.asin)
            if not existing:
                added += 1
                raw_by_asin[product.asin] = product
            else:
                raw_by_asin[product.asin] = merge_product_records(existing, product)
        stage_raw_products(list(raw_by_asin.values()))
        st.session_state.collection_total_raw_count = len(raw_by_asin)
        return added

    def append_failed_summary(
        seed_label: str,
        seed_url: str,
        products: list[Product],
        page_count: int,
        error_message: str,
        detail_error: str | None = None,
    ) -> None:
        added = merge_seed_products(products)
        detail_message = detail_error or error_message
        st.session_state.collection_failed_seed_count += 1
        st.session_state.collection_failed_seed_details.append({
            "label": seed_label,
            "url": seed_url,
            "error": detail_message,
        })
        seed_summaries.append({
            "label": seed_label,
            "url": seed_url,
            "added": added,
            "raw": len(products),
            "quality_ok": False,
            "quality_message": error_message,
            "pages": page_count,
            "failed": True,
            "warning": False,
            "empty": False,
            "error": detail_message,
        })

    def append_success_summary(
        seed_label: str,
        seed_url: str,
        products: list[Product],
        quality_ok: bool,
        quality_message: str,
        page_count: int,
    ) -> int:
        added = merge_seed_products(products)
        warning = is_collection_quality_warning(quality_message)
        empty = quality_message == EMPTY_NEW_RELEASES_MESSAGE
        failed = not quality_ok and not empty
        seed_summaries.append({
            "label": seed_label,
            "url": seed_url,
            "added": added,
            "raw": len(products),
            "quality_ok": quality_ok,
            "quality_message": quality_message,
            "pages": page_count,
            "failed": failed,
            "warning": warning,
            "empty": empty,
            "error": "" if quality_ok else quality_message,
        })
        if failed:
            st.session_state.collection_failed_seed_count += 1
            st.session_state.collection_failed_seed_details.append({
                "label": seed_label,
                "url": seed_url,
                "error": quality_message,
            })
        if warning:
            st.session_state.collection_warning_seed_count += 1
            st.session_state.collection_warning_seed_details.append({
                "label": seed_label,
                "url": seed_url,
                "warning": quality_message,
            })
        if empty:
            st.session_state.collection_empty_seed_count += 1
        return added

    st.session_state.collection_total_seed_count = len(seed_urls)
    st.session_state.collection_completed_seed_count = 0
    st.session_state.collection_failed_seed_count = 0
    st.session_state.collection_warning_seed_count = 0
    st.session_state.collection_empty_seed_count = 0
    st.session_state.collection_failed_seed_details = []
    st.session_state.collection_warning_seed_details = []
    st.session_state.collection_total_raw_count = 0
    st.session_state.collection_current_seed_index = 0
    st.session_state.collection_current_seed_label = ""
    save_collection_report([], 0, "running", seed_urls)
    update_collection_total_status(total_status_placeholder)
    for seed_index, (seed_label, seed_url) in enumerate(seed_urls, start=1):
        raise_if_stop_requested()
        seed_finished = False
        st.session_state.collection_current_seed_index = seed_index
        st.session_state.collection_current_seed_label = seed_label
        update_collection_total_status(total_status_placeholder)
        seed_start = int(((seed_index - 1) / len(seed_urls)) * 90)
        seed_end = int((seed_index / len(seed_urls)) * 90)
        progress_bar.progress(
            seed_start,
            text=f"准备采集小类 {seed_index}/{len(seed_urls)}：{seed_label}｜总原始去重 {len(raw_by_asin)} 条",
        )
        try:
            seed_products, quality_ok, quality_message, page_read_count = collect_sellersprite_batch(
                amazon_url_for_list_type(seed_url, list_type),
                list_type,
                filters,
                progress_bar,
                seed_label=seed_label,
                progress_start=seed_start,
                progress_end=seed_end,
                progress_prefix="",
                total_status_placeholder=total_status_placeholder,
            )
        except SellerSpriteLoadTimeout as exc:
            added = merge_seed_products(exc.products)
            deferred_seeds.defer(
                label=seed_label,
                url=seed_url,
                pages=exc.page_count,
                products=exc.products,
                error=str(exc),
            )
            save_collection_report(seed_summaries, len(raw_by_asin), "running", seed_urls)
            progress_bar.progress(
                seed_end,
                text=(
                    f"小类 {seed_index}/{len(seed_urls)} 插件加载超时，已加入末尾补采队列：{seed_label}｜"
                    f"保留新增 ASIN {added} 条｜待补采 {len(deferred_seeds)} 个"
                ),
            )
            log(
                f"Seed {seed_index}/{len(seed_urls)} deferred after SellerSprite timeout: "
                f"{seed_label}. Kept {added} partial products and continued."
            )
        except CollectionStopped:
            for product in st.session_state.get("collection_staged_raw_products", []):
                existing = raw_by_asin.get(product.asin)
                raw_by_asin[product.asin] = merge_product_records(existing, product) if existing else product
            stage_raw_products(list(raw_by_asin.values()))
            save_collection_report(seed_summaries, len(raw_by_asin), "stopped", seed_urls)
            raise
        except Exception as exc:
            partial_products = st.session_state.get("collection_staged_raw_products", [])
            error_message = str(exc).strip() or exc.__class__.__name__
            append_failed_summary(
                seed_label,
                seed_url,
                partial_products,
                0,
                f"入口采集失败：{error_message}",
                detail_error=error_message,
            )
            added = seed_summaries[-1]["added"]
            save_collection_report(seed_summaries, len(raw_by_asin), "running", seed_urls)
            progress_bar.progress(
                seed_end,
                text=(
                    f"小类 {seed_index}/{len(seed_urls)} 失败，已保留部分数据并继续：{seed_label}｜"
                    f"新增 ASIN {added} 条｜总原始去重 {len(raw_by_asin)} 条"
                ),
            )
            log(
                f"Seed {seed_index}/{len(seed_urls)} failed: {seed_label}. "
                f"Kept {added} partial products. Error: {error_message}"
            )
            seed_finished = True
        else:
            added = append_success_summary(
                seed_label,
                seed_url,
                seed_products,
                quality_ok,
                quality_message,
                page_read_count,
            )
            progress_bar.progress(
                seed_end,
                text=f"小类 {seed_index}/{len(seed_urls)} 完成：{seed_label}｜新增 ASIN {added} 条｜总原始去重 {len(raw_by_asin)} 条",
            )
            save_collection_report(seed_summaries, len(raw_by_asin), "running", seed_urls)
            log(f"Seed {seed_index}/{len(seed_urls)} finished: {seed_label}. added raw {added}, total raw {len(raw_by_asin)}.")
            seed_finished = True
        finally:
            st.session_state.collection_total_raw_count = len(raw_by_asin)
            if seed_finished:
                st.session_state.collection_completed_seed_count = len(seed_summaries)
            st.session_state.collection_current_seed_index = 0
            st.session_state.collection_current_seed_label = ""
            update_collection_total_status(total_status_placeholder)
    raise_if_stop_requested()

    retry_total = len(deferred_seeds)
    for retry_index, deferred in enumerate(deferred_seeds, start=1):
        raise_if_stop_requested()
        seed_label = deferred.label
        seed_url = deferred.url
        retry_start = 90 + int(((retry_index - 1) / max(retry_total, 1)) * 10)
        retry_end = 90 + int((retry_index / max(retry_total, 1)) * 10)
        st.session_state.collection_current_seed_index = min(
            len(seed_summaries) + 1,
            len(seed_urls),
        )
        st.session_state.collection_current_seed_label = f"补采：{seed_label}"
        update_collection_total_status(total_status_placeholder)
        progress_bar.progress(
            retry_start,
            text=f"末尾补采 {retry_index}/{retry_total}：{seed_label}｜总原始去重 {len(raw_by_asin)} 条",
        )
        try:
            seed_products, quality_ok, quality_message, page_read_count = collect_sellersprite_batch(
                amazon_url_for_list_type(seed_url, list_type),
                list_type,
                filters,
                progress_bar,
                seed_label=f"{seed_label}（末尾补采）",
                progress_start=retry_start,
                progress_end=retry_end,
                progress_prefix="",
                total_status_placeholder=total_status_placeholder,
            )
        except SellerSpriteLoadTimeout as exc:
            append_failed_summary(
                seed_label,
                seed_url,
                exc.products,
                exc.page_count,
                f"末尾补采仍失败：{exc}",
            )
            progress_bar.progress(
                retry_end,
                text=f"末尾补采 {retry_index}/{retry_total} 仍失败，已记录并继续：{seed_label}",
            )
            log(f"Deferred seed retry still failed: {seed_label}. Error: {exc}")
        except CollectionStopped:
            stage_raw_products(list(raw_by_asin.values()))
            save_collection_report(seed_summaries, len(raw_by_asin), "stopped", seed_urls)
            raise
        except Exception as exc:
            error_message = str(exc).strip() or exc.__class__.__name__
            append_failed_summary(
                seed_label,
                seed_url,
                deferred.products,
                deferred.pages,
                f"末尾补采异常：{error_message}",
            )
            progress_bar.progress(
                retry_end,
                text=f"末尾补采 {retry_index}/{retry_total} 异常，已记录并继续：{seed_label}",
            )
            log(f"Deferred seed retry raised an exception: {seed_label}. Error: {error_message}")
        else:
            added = append_success_summary(
                seed_label,
                seed_url,
                seed_products,
                quality_ok,
                quality_message,
                page_read_count,
            )
            progress_bar.progress(
                retry_end,
                text=(
                    f"末尾补采 {retry_index}/{retry_total} 完成：{seed_label}｜"
                    f"新增 ASIN {added} 条｜总原始去重 {len(raw_by_asin)} 条"
                ),
            )
            log(f"Deferred seed retry finished: {seed_label}. added raw {added}.")
        finally:
            st.session_state.collection_completed_seed_count = len(seed_summaries)
            st.session_state.collection_total_raw_count = len(raw_by_asin)
            st.session_state.collection_current_seed_index = 0
            st.session_state.collection_current_seed_label = ""
            save_collection_report(seed_summaries, len(raw_by_asin), "running", seed_urls)
            update_collection_total_status(total_status_placeholder)

    progress_bar.progress(100, text=f"全部小类采集完成：原始去重合计 {len(raw_by_asin)} 条。现在可以应用筛选或查看产品列表。")
    ok_count = sum(1 for item in seed_summaries if item["quality_ok"])
    failed_items = [item for item in seed_summaries if item.get("failed")]
    warning_items = [item for item in seed_summaries if item.get("warning")]
    empty_items = [item for item in seed_summaries if item.get("empty")]
    weak_items = [item for item in seed_summaries if not item["quality_ok"]]
    weak_preview = "；".join(f"{item['label']}（{item['quality_message']}）" for item in weak_items[:3])
    weak_suffix = f" 疑似漏采小类：{weak_preview}" if weak_preview else ""
    if len(weak_items) > 3:
        weak_suffix += f"；另有 {len(weak_items) - 3} 个小类请查看日志。"
    st.session_state.last_cache_refresh_message = (
        f"批量采集完成：计划入口 {len(seed_urls)} 个，实际处理 {len(seed_summaries)} 个；"
        f"质量通过 {ok_count} 个（其中数量偏少警告 {len(warning_items)} 个、空榜 {len(empty_items)} 个），"
        f"失败 {len(failed_items)} 个，"
        f"疑似漏采或失败 {len(weak_items)} 个；原始去重合计 {len(raw_by_asin)} 条。"
        f"{weak_suffix}"
    )
    save_collection_report(seed_summaries, len(raw_by_asin), "completed", seed_urls)
    return list(raw_by_asin.values())


def render_cards(products, display_start: int = 0):
    st.markdown("<span class='seller-list-frame seller-table-header-anchor'></span>", unsafe_allow_html=True)
    st.markdown(
        """
        <div class="seller-header">
            <div>#</div><div>产品信息 <span>数据解释</span></div><div>大类BSR</div><div>销量趋势(父)</div>
            <div>销量(父)<br>增长率</div><div>销售额</div><div>子体销量<br>子体销售额</div><div>变体数</div>
            <div>价格<br>Q&A</div><div>评分数<br>月新增</div><div>评分<br>留评率</div><div>FBA<br>毛利率</div>
            <div>上架时间</div><div>配送<br>买家运费</div><div>操作</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    for index, product in enumerate(products, start=1):
        display_number = display_start + index
        st.markdown(
            f"<span class='product-card-select-anchor' data-asin='{escape(product.asin)}'></span>",
            unsafe_allow_html=True,
        )
        product.selected = st.checkbox(
            "Include in export",
            value=product.selected,
            key=f"row_include_{product.asin}",
            label_visibility="collapsed",
        )
        render_list_product_favorite_button(product)
        st.markdown(seller_product_html(product, display_number), unsafe_allow_html=True)
        render_list_product_note_input(product)
        st.markdown("<div class='seller-row-space'></div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)


def set_result_page(page: int, total_items: int) -> None:
    st.session_state.result_current_page = clamp_page(
        page,
        total_items,
        st.session_state.get("result_page_size", PAGE_SIZE_OPTIONS[0]),
    )


def render_result_pagination_controls(total_items: int, location: str) -> None:
    page_size = normalize_page_size(st.session_state.get("result_page_size"))
    current_page = clamp_page(st.session_state.get("result_current_page", 1), total_items, page_size)
    st.session_state.result_current_page = current_page
    total_pages = page_count(total_items, page_size)
    st.markdown(f"<span class='result-pagination-anchor result-pagination-{location}'></span>", unsafe_allow_html=True)
    if location == "top":
        cols = st.columns([1.0, 0.95, 0.8, 1.2, 0.95, 0.95, 1.2], vertical_alignment="center")
        cols[0].markdown(
            f"<div class='pagination-meta'>{page_range_label(total_items, current_page, page_size)}</div>",
            unsafe_allow_html=True,
        )
        cols[1].selectbox(
            "每页",
            PAGE_SIZE_OPTIONS,
            key="result_page_size",
            label_visibility="collapsed",
            on_change=handle_result_page_size_change,
        )
        if cols[2].button("上一页", key="result_prev_top", use_container_width=True, disabled=current_page <= 1):
            set_result_page(current_page - 1, total_items)
            st.rerun()
        cols[3].markdown(
            f"<div class='pagination-meta page-count'>第 <strong>{current_page}</strong> / {total_pages} 页</div>",
            unsafe_allow_html=True,
        )
        if cols[4].button("下一页", key="result_next_top", use_container_width=True, disabled=current_page >= total_pages):
            set_result_page(current_page + 1, total_items)
            st.rerun()
        st.session_state.result_page_jump = current_page
        cols[5].number_input(
            "跳转页",
            min_value=1,
            max_value=total_pages,
            step=1,
            key="result_page_jump",
            label_visibility="collapsed",
            on_change=handle_result_page_jump_change,
        )
        if cols[6].button("跳转", key="result_jump_top", use_container_width=True):
            handle_result_page_jump_change()
            st.rerun()
    else:
        cols = st.columns([1.25, 0.85, 1.2, 0.85, 3.2], vertical_alignment="center")
        cols[0].markdown(
            f"<div class='pagination-meta'>{page_range_label(total_items, current_page, page_size)}</div>",
            unsafe_allow_html=True,
        )
        if cols[1].button("上一页", key=f"result_prev_{location}", use_container_width=True, disabled=current_page <= 1):
            set_result_page(current_page - 1, total_items)
            st.rerun()
        cols[2].markdown(
            f"<div class='pagination-meta page-count'>第 <strong>{current_page}</strong> / {total_pages} 页</div>",
            unsafe_allow_html=True,
        )
        if cols[3].button("下一页", key=f"result_next_{location}", use_container_width=True, disabled=current_page >= total_pages):
            set_result_page(current_page + 1, total_items)
            st.rerun()

ensure_state()
collection_locked = bool(st.session_state.collection_in_progress)

st.markdown("<span class='page-header-anchor'></span>", unsafe_allow_html=True)
header_left, header_right = st.columns([4.5, 1.15], vertical_alignment="center")
with header_right:
    UI_LANG = st.radio(
        "Language / 语言",
        ["中文", "English"],
        horizontal=True,
        key="ui_language_selector",
        on_change=close_category_dialog_state,
        label_visibility="collapsed",
        disabled=collection_locked,
    )
T = TEXT[UI_LANG]
with header_left:
    st.title(T["title"])
    st.caption(T["caption"])

st.markdown(
    """
    <style>
    html, body, [data-testid="stAppViewContainer"] {
        background: #f4f5f7;
        overflow-x: auto;
    }
    [data-testid="stHeader"] {
        background: transparent;
    }
    .block-container {
        padding-top: 0.55rem;
        padding-left: 1.25rem;
        padding-right: 1.25rem;
        max-width: none;
        background: transparent;
    }
    h1 {
        color: #222733;
        font-size: clamp(28px, 2.8vw, 40px) !important;
        letter-spacing: 0 !important;
        line-height: 1.05 !important;
        margin-bottom: 2px !important;
    }
    div[data-testid="stCaptionContainer"] {
        color: #8b94a3;
        font-size: 13px;
    }
    div[data-testid="stElementContainer"]:has(.page-header-anchor) {
        display: none;
    }
    div[data-testid="stElementContainer"]:has(.page-header-anchor) + div[data-testid="stHorizontalBlock"] {
        align-items: center !important;
        margin-bottom: 0.35rem;
    }
    div[data-testid="stElementContainer"]:has(.page-header-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stVerticalBlock"] {
        gap: 0.15rem !important;
    }
    div[data-testid="stElementContainer"]:has(.page-header-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stRadio"] {
        margin-top: 0 !important;
    }
    div[data-testid="stElementContainer"]:has(.setup-panel-anchor) {
        display: none;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:has(.setup-panel-anchor) {
        padding: 10px 12px 9px !important;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:has(.setup-panel-anchor) > div > div[data-testid="stVerticalBlock"] {
        gap: 0.45rem !important;
    }
    div[data-testid="stElementContainer"]:has(.setup-controls-anchor) {
        display: none;
    }
    div[data-testid="stElementContainer"]:has(.setup-controls-anchor) + div[data-testid="stHorizontalBlock"] {
        align-items: end !important;
        gap: 14px !important;
    }
    div[data-testid="stElementContainer"]:has(.setup-controls-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stVerticalBlock"] {
        gap: 0.3rem !important;
    }
    div[data-testid="stElementContainer"]:has(.setup-controls-anchor) + div[data-testid="stHorizontalBlock"] label {
        color: #4f5867 !important;
        font-size: 13px !important;
        font-weight: 600 !important;
        min-height: 20px !important;
    }
    div[data-testid="stElementContainer"]:has(.category-control-anchor) {
        display: none;
    }
    div[data-testid="stElementContainer"]:has(.category-control-anchor) + div[data-testid="stHorizontalBlock"] {
        gap: 8px !important;
    }
    div[data-testid="stElementContainer"]:has(.category-control-anchor) + div[data-testid="stHorizontalBlock"] > div:first-child {
        flex: 1 1 auto !important;
        min-width: 150px !important;
    }
    div[data-testid="stElementContainer"]:has(.category-control-anchor) + div[data-testid="stHorizontalBlock"] > div:last-child {
        flex: 0 0 auto !important;
        min-width: 82px !important;
    }
    div[data-testid="stElementContainer"]:has(.category-control-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stCaptionContainer"] {
        line-height: 42px;
        white-space: nowrap;
    }
    div[data-testid="stElementContainer"]:has(.setup-selection-anchor) {
        display: none;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:has(.setup-panel-anchor) div[data-testid="stAlert"] {
        margin: 0 !important;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:has(.setup-panel-anchor) div[data-testid="stAlert"] > div {
        min-height: 38px !important;
        padding: 8px 12px !important;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:has(.setup-panel-anchor) div[data-testid="stAlert"] p {
        font-size: 13px !important;
        line-height: 1.35 !important;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:has(.setup-panel-anchor) hr {
        margin: 0.35rem 0 !important;
    }
    body:not(:has(div[data-testid="stDialog"])) div[data-testid="stApp"],
    body:not(:has(div[data-testid="stDialog"])) [data-testid="stAppViewContainer"] {
        height: auto !important;
        min-height: 100vh !important;
        overflow-y: visible !important;
    }
    body:not(:has(div[data-testid="stDialog"])) section[data-testid="stMain"] {
        height: auto !important;
        min-height: 100vh !important;
        overflow-y: visible !important;
    }
    html:has(div[data-testid="stDialog"]),
    body:has(div[data-testid="stDialog"]) {
        overflow: hidden !important;
    }
    body:has(div[data-testid="stDialog"]) [data-testid="stAppViewContainer"] {
        height: 100vh;
        overflow: hidden !important;
    }
    div[data-testid="stAppViewContainer"]:has(div[data-testid="stDialog"])::before {
        content: "";
        position: fixed;
        inset: 0;
        background: rgba(17, 24, 39, 0.58);
        z-index: 999;
        pointer-events: none;
    }
    div[data-testid="stVerticalBlockBorderWrapper"] {
        background: #ffffff;
        border: 1px solid #e7ebf1;
        border-radius: 6px;
        box-shadow: 0 1px 2px rgba(15, 23, 42, .03);
        padding: 8px 10px 10px;
    }
    div[data-testid="stMetric"] {
        background: #ffffff;
        border: 1px solid #eef0f4;
        border-radius: 6px;
        box-shadow: 0 1px 2px rgba(15, 23, 42, .025);
        padding: 10px 12px;
    }
    div[data-testid="stMetric"] label {
        color: #7b8491 !important;
        font-size: 13px !important;
    }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] {
        color: #2b303b;
        font-size: 28px;
    }
    div[data-testid="stTabs"] {
        background: transparent;
    }
    div[data-testid="stTabs"] [role="tablist"] {
        background: transparent;
        border-bottom: 1px solid #e2e6ee;
    }
    div[data-testid="stTabs"] [role="tabpanel"] {
        padding-top: 4px;
    }
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) {
        display: none;
    }
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stHorizontalBlock"] {
        background: #f4f5f7 !important;
        box-shadow: 0 1px 0 rgba(15, 23, 42, .08);
        margin-top: 0 !important;
        margin-bottom: 0 !important;
        padding: 5px 0 4px !important;
        position: sticky !important;
        top: 0 !important;
        z-index: 80 !important;
    }
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor),
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor),
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor),
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) {
        display: none;
    }
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stHorizontalBlock"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stHorizontalBlock"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stHorizontalBlock"] {
        background: #f4f5f7 !important;
        margin-bottom: 0 !important;
        margin-top: 0 !important;
        padding-left: 0 !important;
        padding-right: 0 !important;
        position: sticky !important;
        z-index: 80 !important;
    }
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stHorizontalBlock"] {
        padding-top: 4px !important;
        padding-bottom: 0 !important;
        top: 0 !important;
    }
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stHorizontalBlock"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stHorizontalBlock"] {
        padding-top: 0 !important;
        padding-bottom: 4px !important;
        top: 40px !important;
    }
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div {
        min-height: 36px !important;
        height: 36px !important;
    }
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button {
        border-radius: 7px !important;
        padding: 0 10px !important;
        white-space: nowrap !important;
    }
    div[data-testid="stElementContainer"]:has(.result-view-toggle-anchor) {
        display: none;
    }
    div[data-testid="stElementContainer"]:has(.result-view-toggle-anchor) + div[data-testid="stHorizontalBlock"] {
        align-items: center !important;
        background: #ffffff;
        border: 1px solid #d8dee8;
        border-radius: 8px;
        gap: 2px !important;
        min-width: 92px;
        padding: 2px !important;
    }
    div[data-testid="stElementContainer"]:has(.result-view-toggle-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="column"] {
        padding: 0 !important;
    }
    div[data-testid="stElementContainer"]:has(.result-view-toggle-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button {
        border: 0 !important;
        border-radius: 6px !important;
        box-shadow: none !important;
        font-size: 13px !important;
        font-weight: 700 !important;
        height: 30px !important;
        min-height: 30px !important;
        padding: 0 6px !important;
    }
    div[data-testid="stElementContainer"]:has(.toolbar-spacer) {
        display: none;
    }
    div[data-testid="stElementContainer"]:has(.seller-list-frame) {
        height: 0 !important;
        margin: 0 !important;
        min-height: 0 !important;
        overflow: hidden !important;
        padding: 0 !important;
    }
    div[data-testid="stElementContainer"]:has(.seller-table-header-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.seller-table-header-anchor) + div[data-testid="stHorizontalBlock"],
    div[data-testid="stElementContainer"]:has(.seller-table-header-anchor) + div[data-testid="stElementContainer"] {
        background: #f6f7f9 !important;
        box-shadow: 0 2px 4px rgba(15, 23, 42, .08);
        margin-top: 0 !important;
        position: sticky !important;
        top: 80px !important;
        z-index: 79 !important;
    }
    div[data-testid="stPopover"] button {
        border: 1px solid #d9dee8;
        border-radius: 6px;
        min-height: 44px;
        justify-content: flex-start;
    }
    div[data-testid="stTextInput"] input,
    div[data-testid="stNumberInput"] input,
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stButton"] button {
        min-height: 42px;
    }
    div[data-testid="stTextInput"] div[data-baseweb="input"],
    div[data-testid="stNumberInput"] div[data-baseweb="input"],
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div {
        background: #ffffff !important;
        border: 1px solid #cfd6e1 !important;
        box-shadow: 0 1px 2px rgba(15, 23, 42, .035) !important;
        transition: border-color .15s ease, box-shadow .15s ease, background .15s ease;
    }
    div[data-testid="stTextInput"] div[data-baseweb="input"]:hover,
    div[data-testid="stNumberInput"] div[data-baseweb="input"]:hover,
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div:hover {
        border-color: #adb7c5 !important;
    }
    div[data-testid="stTextInput"] div[data-baseweb="input"]:focus-within,
    div[data-testid="stNumberInput"] div[data-baseweb="input"]:focus-within,
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div:focus-within {
        border-color: #ff6b6d !important;
        box-shadow: 0 0 0 2px rgba(255, 75, 75, .12) !important;
    }
    div[data-testid="stTextInput"] input,
    div[data-testid="stNumberInput"] input {
        background: #ffffff !important;
        color: #2f3642 !important;
    }
    div[data-testid="stTextInput"] input::placeholder,
    div[data-testid="stNumberInput"] input::placeholder {
        color: #929baa !important;
        opacity: 1 !important;
    }
    div[data-testid="stButton"] button {
        align-items: center;
        display: inline-flex;
        justify-content: center;
    }
    div[data-testid="stElementContainer"]:has(.collection-action-toolbar) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button {
        min-height: 40px;
        height: 40px;
        border-radius: 8px;
        font-size: 14px;
        padding: 0 14px;
        white-space: nowrap;
    }
    div[data-testid="stElementContainer"]:has(.collection-action-toolbar) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button p {
        white-space: nowrap;
    }
    div[data-testid="stElementContainer"]:has(.collection-action-toolbar) + div[data-testid="stHorizontalBlock"] > div:nth-child(2) iframe {
        display: block !important;
        height: 42px !important;
        margin-top: 1px !important;
        width: 100% !important;
    }
    div[data-testid="stElementContainer"]:has(.collection-action-toolbar) + div[data-testid="stHorizontalBlock"] > div:nth-child(2) button {
        background: #ffffff !important;
        border: 1px solid #ff4b4b !important;
        color: #ff4b4b !important;
        font-weight: 700 !important;
    }
    div[data-testid="stElementContainer"]:has(.collection-action-toolbar) + div[data-testid="stHorizontalBlock"] > div:nth-child(2) button:hover {
        background: #fff1f1 !important;
        border-color: #ff4b4b !important;
        color: #ff4b4b !important;
    }
    div[data-testid="stElementContainer"]:has(.collection-action-toolbar) + div[data-testid="stHorizontalBlock"] {
        flex-wrap: nowrap !important;
        gap: 8px !important;
        overflow-x: auto;
        padding-bottom: 2px;
    }
    div[data-testid="stElementContainer"]:has(.collection-action-toolbar) + div[data-testid="stHorizontalBlock"] > div {
        flex: 0 0 108px !important;
        min-width: 108px !important;
    }
    div[data-testid="stElementContainer"]:has(.collection-action-toolbar) + div[data-testid="stHorizontalBlock"] > div:last-child {
        flex-basis: 132px !important;
        min-width: 132px !important;
    }
    .filter-label {
        color: #5f6673;
        font-size: 15px;
        font-weight: 700;
        margin: 6px 0 8px;
    }
    .filter-label span {
        align-items: center;
        border: 1px solid #ff7a1a;
        border-radius: 50%;
        color: #ff7a1a;
        display: inline-flex;
        font-size: 12px;
        font-weight: 700;
        height: 18px;
        justify-content: center;
        margin-left: 8px;
        width: 18px;
    }
    .filter-range-sep {
        color: #b8bfca;
        font-size: 18px;
        font-weight: 700;
        line-height: 42px;
        text-align: center;
    }
    div[data-testid="stElementContainer"]:has(.range-filter-anchor) + div[data-testid="stHorizontalBlock"] {
        flex-wrap: nowrap !important;
        gap: 8px !important;
        overflow-x: auto;
        padding-bottom: 2px;
    }
    div[data-testid="stColumn"] div[data-testid="stVerticalBlock"]:has(.range-filter-anchor) div[data-testid="stHorizontalBlock"] {
        flex-wrap: nowrap !important;
        gap: 8px !important;
        overflow-x: auto;
        padding-bottom: 2px;
    }
    div[data-testid="stElementContainer"]:has(.range-filter-anchor) + div[data-testid="stHorizontalBlock"] > div:nth-child(1),
    div[data-testid="stElementContainer"]:has(.range-filter-anchor) + div[data-testid="stHorizontalBlock"] > div:nth-child(3),
    div[data-testid="stColumn"] div[data-testid="stVerticalBlock"]:has(.range-filter-anchor) div[data-testid="stHorizontalBlock"] > div:nth-child(1),
    div[data-testid="stColumn"] div[data-testid="stVerticalBlock"]:has(.range-filter-anchor) div[data-testid="stHorizontalBlock"] > div:nth-child(3) {
        flex: 1 0 132px !important;
        min-width: 132px !important;
    }
    div[data-testid="stElementContainer"]:has(.range-filter-anchor) + div[data-testid="stHorizontalBlock"] > div:nth-child(2) {
        flex: 0 0 24px !important;
        min-width: 24px !important;
    }
    div[data-testid="stColumn"] div[data-testid="stVerticalBlock"]:has(.range-filter-anchor) div[data-testid="stHorizontalBlock"] > div:nth-child(2) {
        flex: 0 0 24px !important;
        min-width: 24px !important;
    }
    .filter-money-hint {
        display: none;
    }
    .source-static {
        align-items: center;
        background: #f1f4f8;
        border: 1px solid #e7ebf1;
        border-radius: 6px;
        color: #1f2937;
        display: flex;
        font-size: 15px;
        font-weight: 600;
        min-height: 42px;
        padding: 0 14px;
        white-space: nowrap;
    }
    div[data-testid="stButton"] button p {
        white-space: nowrap;
    }
    div[data-testid="stRadio"] > label,
    div[data-testid="stSelectbox"] > label,
    div[data-testid="stTextInput"] > label,
    div[data-testid="stNumberInput"] > label {
        min-height: 24px;
    }
    div[data-testid="stDialog"] {
        max-height: 100vh;
        overflow: hidden !important;
        z-index: 1000;
    }
    div[data-testid="stDialog"] [role="dialog"] {
        border-radius: 4px !important;
        height: auto !important;
        max-height: calc(100dvh - 16px) !important;
        display: flex !important;
        flex-direction: column !important;
        overflow: hidden !important;
        width: min(92vw, 1320px) !important;
    }
    div[data-testid="stDialog"] section {
        height: auto !important;
        padding-bottom: 0 !important;
        max-height: calc(100vh - 56px) !important;
        overflow: hidden !important;
    }
    div[data-testid="stDialog"] section > div {
        height: auto !important;
        overflow: hidden !important;
    }
    div[data-testid="stDialog"] div[data-testid="stVerticalBlock"] {
        gap: 0.38rem;
    }
    div[data-testid="stDialog"] input {
        background: #ffffff !important;
        box-shadow: none !important;
        font-size: 16px !important;
    }
    div[data-testid="stDialog"] div[data-baseweb="input"] {
        border-color: #d8dee8 !important;
        box-shadow: none !important;
    }
    div[data-testid="stDialog"] div[data-baseweb="input"]:focus-within {
        border-color: #d8dee8 !important;
        box-shadow: none !important;
        outline: none !important;
    }
    div[data-testid="stDialog"] input:focus {
        box-shadow: none !important;
        outline: none !important;
    }
    div[data-testid="stDialog"] div[data-testid="stCheckbox"] {
        min-height: 32px;
    }
    div[data-testid="stDialog"] div[data-testid="stCheckbox"] label {
        align-items: center;
    }
    div[data-testid="stDialog"] div[data-testid="stHorizontalBlock"] > div:first-child button[kind="secondary"] {
        background: transparent !important;
        border: 0 !important;
        box-shadow: none !important;
        color: #b8c0cc !important;
        font-size: 17px !important;
        height: 28px !important;
        min-height: 28px !important;
        padding: 0 !important;
    }
    div[data-testid="stDialog"] div[data-testid="stHorizontalBlock"] > div:first-child button[kind="secondary"]:hover {
        background: transparent !important;
        color: #ff7a1a !important;
    }
    .category-tree-label {
        color: #4b5563;
        font-size: 14px;
        line-height: 30px;
        overflow: hidden;
        text-overflow: ellipsis;
        white-space: nowrap;
    }
    .category-tree-indent {
        display: block;
        height: 30px;
        position: relative;
    }
    .category-tree-indent::after {
        background: #edf0f5;
        content: "";
        height: 30px;
        position: absolute;
        right: 8px;
        top: 0;
        width: 1px;
    }
    .category-tree-root-indent {
        display: block;
        height: 30px;
        width: 1px;
    }
    .category-count-badge {
        background: #fff4e5;
        border-radius: 5px;
        color: #ff991f;
        display: inline-block;
        float: right;
        font-size: 13px;
        line-height: 24px;
        min-width: 44px;
        padding: 0 6px;
        text-align: center;
    }
    .category-tree-spacer {
        display: inline-block;
        height: 30px;
        width: 100%;
    }
    div[data-testid="stDialog"] div[data-testid="stLayoutWrapper"]:has(> div[data-testid="stVerticalBlock"] > div[data-testid="stElementContainer"] .category-tree-scroll-anchor),
    div[data-testid="stDialog"] div[data-testid="stVerticalBlock"]:has(> div[data-testid="stElementContainer"] .category-tree-scroll-anchor) {
        height: auto !important;
        max-height: max(180px, calc(100dvh - 300px)) !important;
        min-height: 0 !important;
        overflow-y: auto !important;
        scrollbar-gutter: stable;
    }
    .category-tree-scroll-anchor {
        display: none;
    }
    .category-dialog-body {
        max-height: min(62vh, 620px);
        overflow-y: auto;
        padding: 2px 4px 10px 0;
        scrollbar-gutter: stable;
    }
    .category-selected-title {
        color: #2f3642;
        font-size: 18px;
        font-weight: 700;
        margin: 0;
        line-height: 42px;
    }
    .category-selected-panel {
        background: #ffffff;
        border: 1px solid #e7ebf1;
        border-radius: 6px;
        height: min(240px, calc(100dvh - 330px));
        max-height: min(240px, calc(100dvh - 330px));
        min-height: 100px;
        overflow-y: auto;
        padding: 12px;
    }
    .category-selected-empty {
        color: #8b94a3;
        font-size: 15px;
        line-height: 1.5;
        min-height: 92px;
        padding: 8px 0 12px;
    }
    div[data-testid="stDialog"] div[data-testid="stElementContainer"]:has(.category-dialog-grid-anchor) {
        display: none;
    }
    div[data-testid="stDialog"] div[data-testid="stElementContainer"]:has(.category-dialog-grid-anchor) + div[data-testid="stHorizontalBlock"] {
        align-items: flex-start !important;
        max-height: max(220px, calc(100dvh - 250px)) !important;
        min-height: 0 !important;
        overflow-y: auto !important;
        scrollbar-gutter: stable;
    }
    .category-dialog-footer {
        background: #ffffff;
        border-top: 1px solid #e7ebf1;
        bottom: 0;
        flex: 0 0 auto;
        margin: 8px -4px 0;
        padding: 12px 4px 0;
        position: sticky;
        z-index: 10;
    }
    div[data-testid="stDialog"] div[data-testid="stElementContainer"]:has(.category-footer-anchor) {
        margin-top: 6px !important;
        padding-top: 6px !important;
        border-top: 1px solid #e7ebf1 !important;
    }
    div[data-testid="stDialog"] div[data-testid="stElementContainer"]:has(.category-footer-anchor) + div[data-testid="stHorizontalBlock"] {
        background: #ffffff !important;
        padding: 0 !important;
        position: sticky !important;
        bottom: 0 !important;
        z-index: 20 !important;
    }
    div[data-testid="stDialog"] div[data-testid="stElementContainer"]:has(.category-footer-anchor) + div[data-testid="stHorizontalBlock"] button {
        min-height: 34px !important;
        height: 34px !important;
        border-radius: 4px !important;
        font-size: 14px !important;
    }
    .cache-card {
        background: #ffffff;
        border: 1px solid #e7ebf1;
        border-radius: 6px;
        box-shadow: 0 1px 2px rgba(15, 23, 42, .03);
        margin: 12px 0 4px;
        padding: 12px 14px;
    }
    .cache-card-top {
        align-items: center;
        display: flex;
        gap: 12px;
        justify-content: space-between;
    }
    .cache-title {
        color: #1f2937;
        font-size: 14px;
        font-weight: 700;
        line-height: 1.35;
    }
    .cache-sub {
        color: #8a94a3;
        font-size: 12px;
        line-height: 1.45;
        margin-top: 2px;
    }
    .cache-sub a {
        color: #ff7a1a;
        text-decoration: none;
    }
    .cache-badges {
        display: flex;
        flex-wrap: wrap;
        gap: 6px;
        justify-content: flex-end;
    }
    .cache-badge {
        background: #f3f5f8;
        border-radius: 999px;
        color: #657080;
        display: inline-flex;
        font-size: 12px;
        padding: 3px 9px;
        white-space: nowrap;
    }
    .cache-badge.ok {
        background: #ecfdf5;
        color: #059669;
    }
    .cache-badge.warn {
        background: #fff7ed;
        color: #ea580c;
    }
    .cache-progress {
        background: #eef1f5;
        border-radius: 999px;
        height: 6px;
        margin-top: 10px;
        overflow: hidden;
    }
    .cache-progress span {
        background: #ff4d4f;
        border-radius: inherit;
        display: block;
        height: 100%;
    }
    .cache-foot {
        color: #8a94a3;
        font-size: 12px;
        margin-top: 6px;
    }
    .cache-foot strong {
        color: #ff7a1a;
    }
    .cache-source-note {
        color: #6b7280;
        font-size: 12px;
        line-height: 1.5;
        margin-top: 6px;
    }
    .selected-pill {
        display: inline-block;
        background: #fff4e5;
        color: #f28c18;
        border-radius: 6px;
        padding: 4px 8px;
        margin: 2px 4px 2px 0;
        font-size: 12px;
        max-width: 100%;
        overflow: hidden;
        text-overflow: ellipsis;
        vertical-align: middle;
        white-space: nowrap;
    }
    .product-row-gap {
        height: 14px;
    }
    .cards-scroll-title {
        color: #5f6b7a;
        font-size: 13px;
        margin: 8px 0 4px;
    }
    .toolbar-meta {
        color: #7a8491;
        font-size: 14px;
        line-height: 40px;
        white-space: nowrap;
    }
    .toolbar-meta strong {
        color: #ff7a1a;
    }
    .toolbar-spacer {
        height: 0;
    }
    div[data-testid="stElementContainer"]:has(.result-pagination-anchor) {
        display: none;
    }
    div[data-testid="stElementContainer"]:has(.result-pagination-anchor) + div[data-testid="stHorizontalBlock"] {
        align-items: center !important;
        margin: 8px 0 10px !important;
        padding: 8px 10px !important;
        background: #ffffff;
        border: 1px solid #e6ebf1;
        border-radius: 8px;
    }
    .pagination-meta {
        color: #667383;
        font-size: 13px;
        line-height: 36px;
        white-space: nowrap;
    }
    .pagination-meta strong {
        color: #ff7a1a;
    }
    .pagination-meta.page-count {
        text-align: center;
    }
    .seller-list-frame {
        background: #f1f3f6;
        border-radius: 4px;
        overflow-x: auto;
        padding: 0 0 6px;
        position: relative;
        scrollbar-gutter: auto;
    }
    .seller-list-frame::-webkit-scrollbar {
        height: 10px;
    }
    .seller-list-frame::-webkit-scrollbar-thumb {
        background: #d9dee7;
        border-radius: 999px;
    }
    .seller-row-space {
        height: 10px;
    }
    .seller-header {
        display: grid;
        grid-template-columns: 30px 254px 56px 92px 68px 78px 82px 46px 62px 64px 52px 58px 70px 56px 32px;
        align-items: center;
        gap: 7px;
        min-width: 1190px;
        width: max(100%, 1190px);
        background: #f6f7f9;
        border: 1px solid #eef0f4;
        border-left: 0;
        border-right: 0;
        color: #6f7782;
        font-size: 12px;
        font-weight: 700;
        min-height: 58px;
        padding: 8px 10px;
        position: static;
        box-shadow: none;
    }
    .seller-header span {
        color: #ff7a1a;
        font-weight: 500;
        margin-left: 8px;
    }
    .seller-header > div:nth-child(2) {
        padding-left: 96px;
    }
    .seller-row {
        min-width: 1190px;
        width: max(100%, 1190px);
        background: #ffffff;
        border: 1px solid #e8edf3;
        border-radius: 3px;
        box-shadow: 0 1px 1px rgba(15, 23, 42, .02);
        padding: 12px 10px 9px;
        position: relative;
        transition: border-color .15s ease, box-shadow .15s ease, background .15s ease;
    }
    .seller-row:hover {
        border-color: #e2e7ef;
        box-shadow: 0 4px 12px rgba(15, 23, 42, .04);
    }
    .tile-grid-frame {
        margin-top: 4px;
    }
    .tile-card {
        background: #ffffff;
        border: 1px solid #e7ebf1;
        border-radius: 4px;
        box-shadow: 0 1px 2px rgba(15, 23, 42, .03);
        color: #2a303b;
        min-height: 510px;
        padding: 12px 12px 14px;
        position: relative;
        transition: border-color .15s ease, box-shadow .15s ease, transform .15s ease;
    }
    .tile-card:hover {
        border-color: #dce3ec;
        box-shadow: 0 8px 18px rgba(15, 23, 42, .07);
        transform: translateY(-1px);
    }
    div[data-testid="stElementContainer"]:has(.product-card-select-anchor),
    div[data-testid="stElementContainer"]:has(.tile-card-select-anchor) {
        display: none;
    }
    div[data-testid="stElementContainer"]:has(.product-card-select-anchor) + div[data-testid="stCheckbox"],
    div[data-testid="stElementContainer"]:has(.product-card-select-anchor) + div[data-testid="stElementContainer"]:has(div[data-testid="stCheckbox"]) {
        height: 1px !important;
        margin: 0 !important;
        opacity: 0 !important;
        overflow: hidden !important;
        pointer-events: none !important;
        position: absolute !important;
        width: 1px !important;
        z-index: -1 !important;
    }
    div[data-testid="stElementContainer"]:has(.tile-card-select-anchor) + div[data-testid="stCheckbox"],
    div[data-testid="stElementContainer"]:has(.tile-card-select-anchor) + div[data-testid="stElementContainer"]:has(div[data-testid="stCheckbox"]) {
        margin: 0 0 -30px 10px !important;
        position: relative !important;
        transform: translateY(11px);
        width: 28px !important;
        z-index: 12 !important;
    }
    div[data-testid="stElementContainer"]:has(.product-card-select-anchor) + div[data-testid="stCheckbox"] label,
    div[data-testid="stElementContainer"]:has(.product-card-select-anchor) + div[data-testid="stElementContainer"] div[data-testid="stCheckbox"] label,
    div[data-testid="stElementContainer"]:has(.tile-card-select-anchor) + div[data-testid="stCheckbox"] label,
    div[data-testid="stElementContainer"]:has(.tile-card-select-anchor) + div[data-testid="stElementContainer"] div[data-testid="stCheckbox"] label {
        min-height: 24px !important;
    }
    div[data-testid="stElementContainer"]:has(.product-card-select-anchor) + div[data-testid="stCheckbox"] label > div,
    div[data-testid="stElementContainer"]:has(.product-card-select-anchor) + div[data-testid="stElementContainer"] div[data-testid="stCheckbox"] label > div,
    div[data-testid="stElementContainer"]:has(.tile-card-select-anchor) + div[data-testid="stCheckbox"] label > div,
    div[data-testid="stElementContainer"]:has(.tile-card-select-anchor) + div[data-testid="stElementContainer"] div[data-testid="stCheckbox"] label > div {
        margin: 0 !important;
    }
    .list-select-host {
        left: 14px;
        position: absolute;
        top: 14px;
        z-index: 18;
    }
    .list-select-proxy {
        align-items: center;
        background: #ffffff;
        border: 1px solid #d8dee8;
        border-radius: 4px;
        color: #ffffff;
        cursor: pointer;
        display: flex;
        height: 18px;
        justify-content: center;
        padding: 0;
        transition: background .12s ease, border-color .12s ease, box-shadow .12s ease;
        width: 18px;
    }
    .list-select-proxy.is-selected {
        background: var(--brand);
        border-color: var(--brand);
    }
    .list-select-proxy.is-selected::after {
        content: "✓";
        font-size: 13px;
        font-weight: 900;
        line-height: 1;
    }
    .list-select-proxy:hover {
        border-color: var(--brand);
        box-shadow: 0 0 0 3px var(--brand-soft);
    }
    .tile-image-wrap {
        align-items: center;
        display: flex;
        height: 210px;
        justify-content: center;
        margin-bottom: 8px;
    }
    .tile-image-wrap img {
        height: 200px;
        max-width: 100%;
        object-fit: contain;
    }
    .tile-title {
        color: #172033;
        display: -webkit-box;
        font-size: 14px;
        font-weight: 600;
        line-height: 1.35;
        min-height: 38px;
        overflow: hidden;
        -webkit-box-orient: vertical;
        -webkit-line-clamp: 2;
    }
    .tile-asin,
    .tile-line {
        color: #707987;
        font-size: 13px;
        line-height: 1.65;
        overflow: hidden;
        text-overflow: ellipsis;
        white-space: nowrap;
    }
    .tile-asin strong,
    .tile-line strong,
    .tile-stats strong {
        color: #222938;
        font-weight: 700;
    }
    .tile-fulfillment,
    .tile-seller-count {
        background: #fff1dc;
        border-radius: 5px;
        color: #ff7a1a;
        display: inline-block;
        font-size: 12px;
        font-weight: 700;
        line-height: 1;
        margin-left: 6px;
        padding: 5px 6px;
    }
    .tile-rank-block {
        border-bottom: 1px dashed #e8edf3;
        border-top: 1px solid #eef2f6;
        color: #4f5b6b;
        font-size: 13px;
        line-height: 1.8;
        margin: 10px -12px 0;
        padding: 9px 12px;
    }
    .tile-rank-pill {
        background: #ff8617;
        border-radius: 5px;
        color: #ffffff;
        display: inline-block;
        font-size: 12px;
        font-weight: 800;
        line-height: 1;
        margin-right: 6px;
        padding: 5px 7px;
    }
    .tile-stats {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 7px 14px;
        color: #7a828e;
        font-size: 13px;
        line-height: 1.45;
        padding-top: 11px;
    }
    .tile-stats .tile-wide {
        grid-column: 1 / -1;
    }
    .tile-note-preview,
    .seller-note-preview {
        background: #f7f9fc;
        border: 1px solid #edf1f6;
        border-radius: 5px;
        color: #5f6875;
        font-size: 12px;
        line-height: 1.45;
        margin-top: 8px;
        overflow: hidden;
        padding: 6px 8px;
        text-overflow: ellipsis;
        white-space: nowrap;
    }
    .seller-note-preview {
        margin-right: 14px;
        max-width: 860px;
    }
    div[data-testid="stElementContainer"]:has(.product-list-favorite-anchor),
    div[data-testid="stElementContainer"]:has(.product-list-note-display-anchor),
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) {
        display: none;
    }
    div[data-testid="stElementContainer"]:has(.product-list-favorite-anchor) + div[data-testid="stButton"],
    div[data-testid="stElementContainer"]:has(.product-list-favorite-anchor) + div[data-testid="stElementContainer"]:has(div[data-testid="stButton"]) {
        height: 1px !important;
        margin: 0 !important;
        opacity: 0 !important;
        overflow: hidden !important;
        pointer-events: none !important;
        position: absolute !important;
        width: 1px !important;
        z-index: -1 !important;
    }
    .list-favorite-host {
        align-items: center;
        display: flex;
        justify-content: center;
        min-height: 34px;
        width: 100%;
    }
    .list-favorite-proxy {
        background: transparent !important;
        border: none !important;
        border-radius: 999px !important;
        box-shadow: none !important;
        color: var(--muted-light) !important;
        cursor: pointer;
        font-size: 30px !important;
        font-weight: 700 !important;
        height: 34px !important;
        line-height: 1 !important;
        min-height: 34px !important;
        min-width: 34px !important;
        padding: 0 !important;
        width: 34px !important;
    }
    .list-favorite-proxy.is-favorite {
        background: transparent !important;
        color: var(--brand) !important;
    }
    .list-favorite-proxy:hover {
        background: var(--brand-soft) !important;
        color: var(--brand) !important;
    }
    div[data-testid="stElementContainer"]:has(.product-list-note-display-anchor) + div[data-testid="stButton"],
    div[data-testid="stElementContainer"]:has(.product-list-note-display-anchor) + div[data-testid="stElementContainer"]:has(div[data-testid="stButton"]) {
        height: 1px !important;
        margin: 0 !important;
        opacity: 0 !important;
        overflow: hidden !important;
        pointer-events: none !important;
        position: absolute !important;
        width: 1px !important;
        z-index: -1 !important;
    }
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stTextInput"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"]:has(div[data-testid="stTextInput"]),
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stForm"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"]:has(div[data-testid="stForm"]),
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stLayoutWrapper"] {
        display: block !important;
        margin: -30px 0 6px 44px !important;
        max-width: 860px !important;
        position: relative !important;
        width: calc(100% - 92px) !important;
        z-index: 24 !important;
    }
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stForm"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"] div[data-testid="stForm"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stForm"] {
        background: transparent !important;
        border: 0 !important;
        box-shadow: none !important;
        height: 24px !important;
        min-height: 24px !important;
        overflow: hidden !important;
        padding: 0 !important;
    }
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stTextInput"] + div[data-testid="stButton"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"]:has(div[data-testid="stTextInput"]) + div[data-testid="stElementContainer"]:has(div[data-testid="stButton"]),
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stForm"] div[data-testid="stButton"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"] div[data-testid="stForm"] div[data-testid="stButton"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stForm"] div[data-testid="stButton"] {
        height: 1px !important;
        margin: 0 !important;
        opacity: 0 !important;
        overflow: hidden !important;
        pointer-events: none !important;
        position: absolute !important;
        width: 1px !important;
        z-index: -1 !important;
    }
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stTextInput"] label,
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"] div[data-testid="stTextInput"] label,
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stForm"] label,
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"] div[data-testid="stForm"] label,
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stForm"] label {
        display: none !important;
    }
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stTextInput"] [data-testid="stTextInputRootElement"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stTextInput"] [data-baseweb="input"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stTextInput"] [data-baseweb="base-input"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"] div[data-testid="stTextInput"] [data-testid="stTextInputRootElement"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"] div[data-testid="stTextInput"] [data-baseweb="input"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"] div[data-testid="stTextInput"] [data-baseweb="base-input"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stForm"] [data-testid="stTextInputRootElement"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stForm"] [data-baseweb="input"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stForm"] [data-baseweb="base-input"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"] div[data-testid="stForm"] [data-testid="stTextInputRootElement"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"] div[data-testid="stForm"] [data-baseweb="input"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"] div[data-testid="stForm"] [data-baseweb="base-input"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stForm"] [data-testid="stTextInputRootElement"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stForm"] [data-baseweb="input"],
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stForm"] [data-baseweb="base-input"] {
        background: transparent !important;
        border: 0 !important;
        box-shadow: none !important;
        height: 24px !important;
        min-height: 24px !important;
    }
    .list-note-host {
        min-height: 24px;
        padding: 0 !important;
    }
    .list-note-host[data-note-mode="display"] {
        background: transparent;
        border: 0;
        margin-top: 5px;
    }
    .list-note-text-button[data-note-mode="display"] {
        box-shadow: none;
    }
    .list-note-host[data-note-mode="edit"] .list-note-proxy {
        visibility: hidden;
    }
    .list-note-proxy {
        background: transparent !important;
        border: 0 !important;
        box-shadow: none !important;
        color: #6b7686 !important;
        cursor: pointer;
        display: block;
        font-size: 12px !important;
        font-weight: 600 !important;
        min-height: 24px !important;
        padding: 0 !important;
        text-align: left !important;
        width: fit-content !important;
    }
    .list-note-proxy:hover {
        color: var(--brand) !important;
    }
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stTextInput"] input,
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"] div[data-testid="stTextInput"] input,
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stForm"] input,
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"] div[data-testid="stForm"] input,
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stForm"] input {
        background: transparent !important;
        border: 0 !important;
        border-radius: 0 !important;
        box-shadow: none !important;
        color: #4f5b6b !important;
        font-size: 12px !important;
        font-weight: 600 !important;
        height: 24px !important;
        line-height: 24px !important;
        max-height: 24px !important;
        min-height: 24px !important;
        padding: 0 !important;
    }
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stTextInput"] input:focus,
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"] div[data-testid="stTextInput"] input:focus,
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stForm"] input:focus,
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stElementContainer"] div[data-testid="stForm"] input:focus,
    div[data-testid="stElementContainer"]:has(.product-list-note-editor-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stForm"] input:focus {
        box-shadow: none !important;
        outline: 0 !important;
    }
    div[data-testid="stElementContainer"]:has(.product-annotation-anchor) {
        display: none;
    }
    div[data-testid="stElementContainer"]:has(.product-annotation-anchor) + div[data-testid="stHorizontalBlock"] {
        background: #ffffff;
        border: 1px solid #e8edf3;
        border-radius: 0 0 5px 5px;
        margin: -7px 0 9px 46px !important;
        max-width: 620px;
        padding: 7px 8px 8px !important;
    }
    div[data-testid="stElementContainer"]:has(.product-annotation-anchor) + div[data-testid="stHorizontalBlock"] button {
        min-height: 38px !important;
    }
    div[data-testid="stElementContainer"]:has(.product-annotation-anchor) + div[data-testid="stHorizontalBlock"] input {
        min-height: 38px !important;
    }
    .seller-main {
        display: grid;
        grid-template-columns: 30px 254px 56px 92px 68px 78px 82px 46px 62px 64px 52px 58px 70px 56px 32px;
        align-items: center;
        gap: 7px;
    }
    .seller-rank {
        color: #a6aeb9;
        font-size: 16px;
        text-align: center;
    }
    .seller-product {
        display: grid;
        grid-template-columns: 124px minmax(0, 1fr);
        gap: 12px;
        align-items: center;
    }
    .seller-image-wrap {
        position: relative;
        width: 124px;
        min-height: 136px;
    }
    .seller-image-wrap img {
        width: 116px;
        height: 116px;
        object-fit: contain;
        margin-top: 20px;
    }
    .seller-info {
        min-width: 0;
    }
    .seller-title {
        color: #111827;
        font-size: 14px;
        font-weight: 500;
        line-height: 1.3;
        overflow: hidden;
        text-overflow: ellipsis;
        white-space: nowrap;
        margin-bottom: 6px;
    }
    .cell {
        color: #151a24;
        font-size: 12px;
        line-height: 1.55;
        text-align: center;
    }
    .cell strong {
        display: block;
        font-size: 14px;
        font-weight: 500;
    }
    .muted {
        color: #9aa2ae;
    }
    .green {
        color: #43af52;
        display: block;
    }
    .trend-cell svg {
        width: 100px;
        height: 52px;
        display: block;
    }
    .ops-cell {
        align-self: center;
        color: #ff7a1a;
        display: grid;
        gap: 7px;
        justify-items: center;
        font-size: 15px;
        padding-top: 0;
    }
    .seller-detail {
        border-top: 1px solid #eef0f4;
        color: #7a828e;
        font-size: 13px;
        line-height: 1.75;
        margin: 9px 0 0 30px;
        padding: 8px 0 0 0;
    }
    .orange {
        color: #ff7a1a;
    }
    .pill {
        border-radius: 4px;
        color: #ffffff;
        display: inline-block;
        font-size: 13px;
        font-weight: 600;
        line-height: 1;
        margin-left: 6px;
        padding: 4px 8px;
    }
    .orange-pill { background: #ff8617; }
    .blue-pill { background: #4b6ff3; }
    .green-pill { background: #35c28d; }
    .rank-pill {
        background: #ff8617;
        border-radius: 10px;
        color: #ffffff;
        display: inline-block;
        font-size: 12px;
        margin: 0 6px;
        padding: 1px 7px;
    }
    .product-identity {
        display: grid;
        grid-template-columns: 42px 150px minmax(0, 1fr);
        gap: 22px;
        align-items: start;
        min-height: 170px;
        padding-top: 8px;
    }
    div[data-testid="stCheckbox"]:has(input[id*="image_include_"]) {
        margin-bottom: -34px;
        position: relative;
        z-index: 3;
        width: 28px;
    }
    .product-rank {
        color: #a3abb7;
        font-size: 22px;
        text-align: center;
        padding-top: 82px;
    }
    .product-media {
        position: relative;
        width: 150px;
        height: 165px;
        display: flex;
        flex-direction: column;
        align-items: flex-start;
        justify-content: flex-start;
    }
    .product-media img {
        width: 128px;
        height: 128px;
        object-fit: contain;
        margin-left: 4px;
        margin-top: 22px;
    }
    .level-corner {
        position: absolute;
        top: 0;
        left: 0;
        background: #ef2b13;
        color: #ffffff;
        border-radius: 5px;
        padding: 2px 6px;
        font-weight: 700;
        font-size: 15px;
        line-height: 1.1;
    }
    .product-info {
        min-width: 0;
        padding-top: 44px;
    }
    .product-title {
        color: #111827;
        font-size: 21px;
        font-weight: 500;
        line-height: 1.3;
        max-width: 760px;
        overflow: hidden;
        text-overflow: ellipsis;
        white-space: nowrap;
        margin-bottom: 6px;
    }
    .meta-line {
        color: #7b8491;
        font-size: 15px;
        line-height: 1.5;
        white-space: nowrap;
    }
    .meta-line strong {
        color: #111827;
        font-weight: 500;
    }
    .meta-line.muted {
        color: #a6adb8;
    }
    .mini-icon {
        color: #c7ced8;
        display: inline-block;
        font-size: 14px;
        margin-left: 5px;
        vertical-align: 1px;
    }
    .copy-icon {
        background: transparent;
        border: 1px solid transparent;
        border-radius: 4px;
        color: #aeb7c3;
        cursor: pointer;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        font-size: 16px;
        height: 22px;
        margin-left: 7px;
        min-width: 22px;
        padding: 0 4px;
        position: relative;
        transition: all .15s ease;
        vertical-align: 0;
    }
    .copy-icon:hover,
    .copy-icon.copied {
        background: #fff4e8;
        border-color: #ffb26f;
        color: #ff7a1a;
    }
    .copy-icon:active {
        background: #ff7a1a;
        border-color: #ff7a1a;
        color: #ffffff;
    }
    .copy-icon.disabled-icon {
        cursor: default;
        opacity: .45;
    }
    .copy-icon.disabled-icon:hover {
        background: transparent;
        border-color: transparent;
        color: #aeb7c3;
    }
    .mini-link {
        border: 1px solid transparent;
        border-radius: 4px;
        color: #aeb7c3 !important;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        font-size: 17px;
        height: 22px;
        margin-left: 5px;
        min-width: 22px;
        padding: 0 4px;
        position: relative;
        text-decoration: none !important;
        transition: all .15s ease;
        vertical-align: 0;
    }
    .mini-link:hover {
        background: #fff4e8;
        border-color: #ffb26f;
        color: #ff7a1a !important;
    }
    div[data-testid="stExpander"] label p,
    div[data-testid="stCheckbox"] label p {
        font-family: inherit;
        white-space: normal;
        line-height: 1.35;
    }

    /* Visual refresh: presentation only. Core collection and filtering logic is unchanged. */
    :root {
        --app-bg: #f2f4f7;
        --surface: #ffffff;
        --surface-soft: #f8fafc;
        --surface-raised: #fcfcfd;
        --ink: #182230;
        --ink-soft: #344054;
        --muted: #667085;
        --muted-light: #98a2b3;
        --line: #e1e6ed;
        --line-strong: #cfd6df;
        --brand: #d94b4b;
        --brand-deep: #b93636;
        --brand-soft: #fff3f2;
        --accent: #0f766e;
        --accent-soft: #ecfdf8;
        --warning: #c66a16;
        --shadow-soft: 0 1px 2px rgba(16, 24, 40, .035), 0 8px 24px rgba(16, 24, 40, .025);
    }
    html,
    body,
    [data-testid="stAppViewContainer"] {
        background: var(--app-bg) !important;
        color: var(--ink);
    }
    body:not(:has(div[data-testid="stDialog"])) section[data-testid="stMain"] {
        overflow-x: visible !important;
        overflow-y: visible !important;
    }
    [data-testid="stHeader"] {
        background: transparent !important;
        border-bottom: 0;
        backdrop-filter: none;
    }
    .block-container {
        padding: 0.7rem 1.4rem 2.5rem !important;
    }
    h1 {
        border-left: 0;
        color: var(--ink) !important;
        font-size: clamp(28px, 2.45vw, 38px) !important;
        font-weight: 760 !important;
        line-height: 1.08 !important;
        margin: 2px 0 5px !important;
        position: relative;
        padding-left: 13px;
    }
    h1::before {
        background: var(--brand);
        border-radius: 2px;
        content: "";
        height: 31px;
        left: 0;
        position: absolute;
        top: 50%;
        transform: translateY(-50%);
        width: 4px;
    }
    div[data-testid="stCaptionContainer"] {
        color: var(--muted-light) !important;
        font-size: 12.5px;
    }
    div[data-testid="stElementContainer"]:has(.page-header-anchor) + div[data-testid="stHorizontalBlock"] {
        margin-bottom: 0.7rem !important;
    }
    div[data-testid="stElementContainer"]:has(.page-header-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stRadio"] {
        background: rgba(255, 255, 255, .72);
        border: 1px solid var(--line);
        border-radius: 8px;
        padding: 5px 10px;
    }
    div[data-testid="stVerticalBlockBorderWrapper"] {
        background: var(--surface) !important;
        border: 1px solid var(--line) !important;
        border-radius: 8px !important;
        box-shadow: var(--shadow-soft) !important;
    }
    div[data-testid="stVerticalBlockBorderWrapper"] > div {
        background: var(--surface) !important;
        border-radius: inherit;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:has(.setup-panel-anchor) {
        padding: 15px 16px 14px !important;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:has(.setup-panel-anchor) > div > div[data-testid="stVerticalBlock"] {
        gap: 0.58rem !important;
    }
    div[data-testid="stElementContainer"]:has(.setup-controls-anchor) + div[data-testid="stHorizontalBlock"] {
        gap: 18px !important;
    }
    div[data-testid="stElementContainer"]:has(.setup-controls-anchor) + div[data-testid="stHorizontalBlock"] label {
        color: var(--ink-soft) !important;
        font-size: 12px !important;
        font-weight: 700 !important;
        letter-spacing: 0 !important;
    }
    .section-heading {
        align-items: center;
        color: var(--ink);
        display: flex;
        font-size: 15px;
        font-weight: 750;
        gap: 9px;
        margin: 2px 0 0;
    }
    .section-heading::before {
        background: var(--brand);
        border-radius: 2px;
        content: "";
        height: 15px;
        width: 3px;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:has(.setup-panel-anchor) hr {
        border-color: var(--line) !important;
        margin: 0.7rem 0 0.65rem !important;
    }
    div[data-testid="stTextInput"] div[data-baseweb="input"],
    div[data-testid="stNumberInput"] div[data-baseweb="input"],
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stPopover"] button {
        background: var(--surface-raised) !important;
        border: 1px solid var(--line-strong) !important;
        border-radius: 7px !important;
        box-shadow: 0 1px 2px rgba(16, 24, 40, .025) !important;
    }
    div[data-testid="stTextInput"] div[data-baseweb="input"]:hover,
    div[data-testid="stNumberInput"] div[data-baseweb="input"]:hover,
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div:hover,
    div[data-testid="stPopover"] button:hover {
        background: var(--surface) !important;
        border-color: #aeb8c5 !important;
    }
    div[data-testid="stTextInput"] div[data-baseweb="input"]:focus-within,
    div[data-testid="stNumberInput"] div[data-baseweb="input"]:focus-within,
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div:focus-within {
        background: var(--surface) !important;
        border-color: var(--brand) !important;
        box-shadow: 0 0 0 3px rgba(217, 75, 75, .1) !important;
    }
    div[data-testid="stTextInput"] input,
    div[data-testid="stNumberInput"] input {
        background: transparent !important;
        color: var(--ink) !important;
        font-weight: 500;
    }
    div[data-testid="stTextInput"] input::placeholder,
    div[data-testid="stNumberInput"] input::placeholder {
        color: #8994a5 !important;
    }
    div[data-testid="stButton"] button,
    div[data-testid="stDownloadButton"] button {
        border-color: var(--line-strong);
        border-radius: 7px !important;
        box-shadow: 0 1px 2px rgba(16, 24, 40, .025);
        color: var(--ink-soft);
        font-weight: 650;
        transition: background .15s ease, border-color .15s ease, box-shadow .15s ease, color .15s ease;
    }
    div[data-testid="stButton"] button:hover,
    div[data-testid="stDownloadButton"] button:hover {
        background: var(--surface-soft);
        border-color: #aeb8c5;
        color: var(--ink);
        box-shadow: 0 2px 5px rgba(16, 24, 40, .06);
    }
    div[data-testid="stButton"] button[kind="primary"],
    div[data-testid="stDownloadButton"] button[kind="primary"] {
        background: var(--brand) !important;
        border-color: var(--brand) !important;
        box-shadow: 0 2px 5px rgba(185, 54, 54, .18) !important;
        color: #ffffff !important;
    }
    div[data-testid="stButton"] button[kind="primary"]:hover,
    div[data-testid="stDownloadButton"] button[kind="primary"]:hover {
        background: var(--brand-deep) !important;
        border-color: var(--brand-deep) !important;
    }
    div[data-testid="stButton"] button:disabled,
    div[data-testid="stDownloadButton"] button:disabled {
        background: #f4f6f8 !important;
        border-color: #e4e8ee !important;
        box-shadow: none !important;
        color: #a6afbc !important;
        opacity: 1 !important;
    }
    div[data-testid="stAlert"] {
        background: #f6f8fb !important;
        border: 1px solid #dfe6ef !important;
        color: #36536f !important;
    }
    div[data-testid="stAlert"] svg {
        color: #557a9d !important;
        fill: #557a9d !important;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:has(.setup-panel-anchor) div[data-testid="stAlert"] {
        border-radius: 7px;
        border: 1px solid #dbe4ef;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:has(.setup-panel-anchor) div[data-testid="stAlert"][data-baseweb="notification"] {
        background: #f5f8fc;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]:has(.setup-panel-anchor) div[data-testid="stAlert"] p {
        color: #36536f;
    }
    .filter-label {
        color: var(--ink-soft);
        font-size: 13px;
        font-weight: 720;
        margin: 5px 0 7px;
    }
    .filter-label span {
        border-color: #d78a42;
        color: #bd681c;
        height: 17px;
        margin-left: 6px;
        width: 17px;
    }
    .filter-range-sep {
        color: #aeb7c4;
        font-size: 15px;
        font-weight: 600;
    }
    .selected-pill {
        background: var(--brand-soft);
        border: 1px solid #f5d4d1;
        color: #a83b3b;
        font-weight: 600;
    }
    div[data-testid="stElementContainer"]:has(.collection-action-toolbar) + div[data-testid="stHorizontalBlock"] {
        background: var(--surface-soft);
        border: 1px solid var(--line);
        border-radius: 8px;
        gap: 8px !important;
        overflow-x: auto;
        padding: 8px !important;
    }
    div[data-testid="stElementContainer"]:has(.collection-action-toolbar) + div[data-testid="stHorizontalBlock"] > div:nth-child(2) iframe {
        border-radius: 7px;
        height: 40px !important;
        margin-top: 0 !important;
    }
    div[data-testid="stMetric"] {
        background: var(--surface) !important;
        border-color: var(--line) !important;
        border-radius: 7px !important;
        box-shadow: 0 1px 2px rgba(16, 24, 40, .025) !important;
    }
    div[data-testid="stMetric"] label {
        color: var(--muted) !important;
        font-weight: 650;
    }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] {
        color: var(--ink);
        font-size: 25px;
        font-weight: 650;
    }
    div[data-testid="stTabs"] [role="tablist"] {
        border-bottom-color: var(--line);
        gap: 20px;
    }
    div[data-testid="stTabs"] button[role="tab"] {
        color: var(--muted);
        font-weight: 650;
        padding-left: 0;
        padding-right: 0;
    }
    div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
        color: var(--brand);
    }
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stHorizontalBlock"] {
        background: rgba(242, 244, 247, .96) !important;
        border-bottom: 1px solid var(--line);
        box-shadow: 0 5px 14px rgba(16, 24, 40, .055);
        padding: 7px 0 6px !important;
    }
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor),
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor),
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) {
        display: none;
    }
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stHorizontalBlock"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stHorizontalBlock"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stHorizontalBlock"] {
        background: rgba(242, 244, 247, .96) !important;
        margin-bottom: 0 !important;
        margin-top: 0 !important;
        padding-left: 0 !important;
        padding-right: 0 !important;
        position: sticky !important;
        z-index: 80 !important;
    }
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stHorizontalBlock"] {
        padding-top: 4px !important;
        padding-bottom: 0 !important;
        top: 0 !important;
    }
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stHorizontalBlock"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stHorizontalBlock"] {
        border-bottom: 1px solid var(--line);
        box-shadow: 0 5px 14px rgba(16, 24, 40, .055);
        padding-top: 0 !important;
        padding-bottom: 4px !important;
        top: 40px !important;
    }
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-actions-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-controls-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stLayoutWrapper"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stElementContainer"]:has(.cards-toolbar-sort-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stSelectbox"] div[data-baseweb="select"] > div {
        min-height: 36px !important;
        height: 36px !important;
    }
    div[data-testid="stElementContainer"]:has(.result-view-toggle-anchor) {
        display: none;
    }
    div[data-testid="stElementContainer"]:has(.result-view-toggle-anchor) + div[data-testid="stHorizontalBlock"] {
        align-items: center !important;
        background: #ffffff;
        border: 1px solid var(--line-strong);
        border-radius: 8px;
        gap: 2px !important;
        min-width: 92px;
        padding: 2px !important;
    }
    div[data-testid="stElementContainer"]:has(.result-view-toggle-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="column"] {
        padding: 0 !important;
    }
    div[data-testid="stElementContainer"]:has(.result-view-toggle-anchor) + div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button {
        border: 0 !important;
        border-radius: 6px !important;
        box-shadow: none !important;
        font-size: 13px !important;
        font-weight: 700 !important;
        height: 30px !important;
        min-height: 30px !important;
        padding: 0 6px !important;
    }
    div[data-testid="stElementContainer"]:has(.seller-table-header-anchor) + div[data-testid="stLayoutWrapper"],
    div[data-testid="stElementContainer"]:has(.seller-table-header-anchor) + div[data-testid="stHorizontalBlock"],
    div[data-testid="stElementContainer"]:has(.seller-table-header-anchor) + div[data-testid="stElementContainer"] {
        background: #eef1f5 !important;
        border-bottom: 1px solid #dce2ea;
        box-shadow: 0 4px 12px rgba(16, 24, 40, .045);
        top: 80px !important;
    }
    .seller-header {
        background: #eef1f5;
        border-color: #dfe4eb;
        color: #596474;
    }
    .seller-header span {
        color: var(--brand);
    }
    .seller-row {
        border-color: var(--line);
        border-radius: 6px;
        box-shadow: 0 1px 2px rgba(16, 24, 40, .025);
    }
    .seller-row:hover {
        border-color: #c9d1dc;
        box-shadow: 0 8px 20px rgba(16, 24, 40, .055);
    }
    .seller-title,
    .product-title,
    .cell strong,
    .meta-line strong {
        color: var(--ink);
    }
    .seller-detail {
        border-top-color: #e9edf2;
        color: var(--muted);
    }
    .orange,
    .toolbar-meta strong,
    .cache-foot strong {
        color: var(--brand);
    }
    .orange-pill {
        background: #bd671d;
    }
    .rank-pill {
        background: var(--brand);
    }
    .green-pill {
        background: var(--accent);
    }
    .cache-card {
        border-color: var(--line);
        border-radius: 8px;
        box-shadow: var(--shadow-soft);
    }
    .cache-progress span {
        background: var(--accent);
    }
    .cache-badge.ok {
        background: var(--accent-soft);
        color: var(--accent);
    }
    .category-count-badge {
        background: #fff7ed;
        border: 1px solid #f8dfbf;
        color: var(--warning);
        font-weight: 650;
    }
    .category-selected-title {
        color: var(--ink);
    }
    .category-selected-panel {
        background: var(--surface-soft);
        border-color: var(--line);
        border-radius: 8px;
    }
    div[data-testid="stDialog"] [role="dialog"] {
        border: 1px solid rgba(255, 255, 255, .5);
        border-radius: 8px !important;
        box-shadow: 0 24px 70px rgba(16, 24, 40, .24);
    }
    div[data-testid="stDialog"] section {
        background: var(--surface);
    }
    div[data-testid="stDialog"] div[data-baseweb="input"] {
        background: var(--surface-soft) !important;
        border-color: var(--line-strong) !important;
    }
    div[data-testid="stDialog"] div[data-baseweb="input"]:focus-within {
        background: var(--surface) !important;
        border-color: var(--brand) !important;
        box-shadow: 0 0 0 3px rgba(217, 75, 75, .1) !important;
    }
    div[data-testid="stDialog"] div[data-testid="stElementContainer"]:has(.category-footer-anchor) + div[data-testid="stHorizontalBlock"] {
        border-top: 1px solid var(--line);
        justify-content: flex-end !important;
        padding-top: 8px !important;
    }
    div[data-testid="stDialog"] div[data-testid="stElementContainer"]:has(.category-footer-anchor) + div[data-testid="stHorizontalBlock"] > div:nth-child(1),
    div[data-testid="stDialog"] div[data-testid="stElementContainer"]:has(.category-footer-anchor) + div[data-testid="stHorizontalBlock"] > div:nth-child(3) {
        display: none !important;
    }
    div[data-testid="stDialog"] div[data-testid="stElementContainer"]:has(.category-footer-anchor) + div[data-testid="stHorizontalBlock"] > div:nth-child(2) {
        flex: 0 0 116px !important;
        min-width: 116px !important;
    }
    div[data-testid="stDialog"] div[data-testid="stElementContainer"]:has(.category-footer-anchor) + div[data-testid="stHorizontalBlock"] > div:nth-child(4) {
        flex: 0 0 168px !important;
        min-width: 168px !important;
    }
    div[data-testid="stDialog"] div[data-testid="stElementContainer"]:has(.category-footer-anchor) + div[data-testid="stHorizontalBlock"] button {
        border-radius: 7px !important;
    }
    .copy-icon:hover,
    .copy-icon.copied,
    .mini-link:hover {
        background: var(--brand-soft);
        border-color: #edb8b4;
        color: var(--brand) !important;
    }
    .copy-icon:active {
        background: var(--brand);
        border-color: var(--brand);
    }
    @media (max-width: 760px) {
        .block-container {
            padding-left: 0.85rem;
            padding-right: 0.85rem;
        }
        h1 {
            font-size: 29px !important;
            max-width: 12em;
        }
        div[data-testid="stElementContainer"]:has(.setup-controls-anchor) + div[data-testid="stHorizontalBlock"] {
            align-items: stretch !important;
            flex-direction: column !important;
            gap: 8px !important;
        }
        div[data-testid="stVerticalBlockBorderWrapper"] {
            padding: 6px;
        }
        .filter-label {
            font-size: 14px;
        }
        .source-static,
        div[data-testid="stTextInput"] input,
        div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
        div[data-testid="stButton"] button {
            min-height: 40px;
        }
        div[data-testid="stColumn"] div[data-testid="stVerticalBlock"]:has(.range-filter-anchor) div[data-testid="stHorizontalBlock"] {
            gap: 6px !important;
            overflow-x: visible;
        }
        div[data-testid="stColumn"] div[data-testid="stVerticalBlock"]:has(.range-filter-anchor) div[data-testid="stHorizontalBlock"] > div:nth-child(1),
        div[data-testid="stColumn"] div[data-testid="stVerticalBlock"]:has(.range-filter-anchor) div[data-testid="stHorizontalBlock"] > div:nth-child(3) {
            flex: 1 1 76px !important;
            min-width: 76px !important;
        }
        div[data-testid="stColumn"] div[data-testid="stVerticalBlock"]:has(.range-filter-anchor) div[data-testid="stHorizontalBlock"] > div:nth-child(2) {
            flex: 0 0 18px !important;
            min-width: 18px !important;
        }
        div[data-testid="stColumn"] div[data-testid="stVerticalBlock"]:has(.range-filter-anchor) input {
            font-size: 13px;
            padding-left: 10px;
            padding-right: 10px;
        }
        .toolbar-meta {
            line-height: 32px;
        }
        div[data-testid="stDialog"] [role="dialog"] {
            height: auto !important;
            max-height: calc(100dvh - 12px) !important;
            width: min(96vw, 1320px) !important;
        }
        div[data-testid="stDialog"] div[data-testid="stElementContainer"]:has(.category-footer-anchor) + div[data-testid="stHorizontalBlock"] {
            flex-direction: row !important;
            flex-wrap: nowrap !important;
            gap: 8px !important;
        }
        div[data-testid="stDialog"] div[data-testid="stElementContainer"]:has(.category-footer-anchor) + div[data-testid="stHorizontalBlock"] > div {
            min-width: 0 !important;
        }
        div[data-testid="stDialog"] div[data-testid="stElementContainer"]:has(.category-footer-anchor) + div[data-testid="stHorizontalBlock"] button {
            min-width: 0 !important;
        }
        div[data-testid="stDialog"] .category-tree-label {
            font-size: 13px;
        }
        div[data-testid="stDialog"] .category-count-badge {
            font-size: 12px;
            min-width: 38px;
            padding: 0 4px;
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

compacted_confirmed_paths = compact_category_paths(st.session_state.confirmed_category_paths)
if compacted_confirmed_paths != st.session_state.confirmed_category_paths:
    st.session_state.confirmed_category_paths = compacted_confirmed_paths

if st.session_state.show_category_dialog:
    render_category_dialog()

with st.container(border=True):
    st.markdown("<span class='setup-panel-anchor'></span>", unsafe_allow_html=True)
    data_source = "卖家精灵插件"
    custom_url = ""
    batch_category_collect = False
    st.markdown("<span class='setup-controls-anchor'></span>", unsafe_allow_html=True)
    setup_list, setup_category, setup_site = st.columns([1.1, 1.65, 1], vertical_alignment="bottom")
    with setup_list:
        list_type = st.radio(
            T["list_type"],
            ["New Releases", "Best Sellers"],
            horizontal=True,
            key="list_type_selector",
            on_change=close_category_dialog_state,
            disabled=collection_locked,
        )
    with setup_category:
        st.markdown(f"<div class='filter-label'>{escape(T['categories'])}</div>", unsafe_allow_html=True)
        st.markdown("<span class='category-control-anchor'></span>", unsafe_allow_html=True)
        category_button_col, category_count_col = st.columns([1.5, 0.7], vertical_alignment="center")
        with category_button_col:
            if st.button(
                "选择类目",
                key="open_category_dialog_button",
                use_container_width=True,
                disabled=collection_locked,
            ):
                st.session_state.category_dialog_selected_paths = list(st.session_state.confirmed_category_paths)
                st.session_state.show_category_dialog = True
                st.rerun()
        with category_count_col:
            st.caption(f"已选 {len(st.session_state.confirmed_category_paths)} 个类目")
    with setup_site:
        marketplace = st.selectbox(
            "站点",
            ["美国站"],
            index=0,
            key="marketplace_selector",
            on_change=close_category_dialog_state,
            disabled=collection_locked,
        )

    selected_paths = st.session_state.confirmed_category_paths
    if selected_paths:
        st.markdown("<span class='setup-selection-anchor'></span>", unsafe_allow_html=True)
        preview = selected_paths[:6]
        pills = "".join(f"<span class='selected-pill'>{escape(path)}</span>" for path in preview)
        more = f" +{len(selected_paths) - len(preview)}" if len(selected_paths) > len(preview) else ""
        st.markdown(pills + more, unsafe_allow_html=True)
    # Preparing every leaf URL can involve thousands of category paths.
    # Defer that work until collection starts so confirming the dialog stays responsive.
    mapped_seed_urls: list[tuple[str, str]] = []
    seller_cache_total = 0
    seller_cache_hydrated = 0
    chrome_ready = False
    if data_source == "卖家精灵插件":
        seller_cache_total, seller_cache_hydrated = sellersprite_cache_hydration()
        chrome_ready = chrome_debugger_available()
        cache_warning = sellersprite_cache_warning()
        if cache_warning and st.session_state.last_cache_refresh_message:
            st.warning(cache_warning)
        if not chrome_ready:
            st.warning("实时采集需要连接采集 Chrome。未连接时不会使用旧产品缓存替代，请先双击“一键启动工具.bat”。")
        else:
            st.info("采集 Chrome 已连接。点击“开始采集”后会打开 Amazon 页面，并等待卖家精灵插件加载后再读取产品。")

    if st.session_state.last_cache_refresh_message:
        st.info(st.session_state.last_cache_refresh_message)
    if st.session_state.last_category_mapping_message:
        st.info(st.session_state.last_category_mapping_message)

    st.divider()
    history_options = raw_history_options()
    if history_options:
        st.markdown("<div class='section-heading'>历史原始采集池</div>", unsafe_allow_html=True)
        st.caption("本地只保留最近 5 次原始采集池；载入后会立即按当前筛选条件重新计算下方结果。")
        history_cols = st.columns([3, 1], vertical_alignment="bottom")
        history_labels = [label for label, _ in history_options]
        selected_history_label = history_cols[0].selectbox(
            "历史原始采集池",
            history_labels,
            label_visibility="collapsed",
            disabled=collection_locked,
        )
        load_history = history_cols[1].button(
            "载入选中记录",
            key="load_history_record_button",
            use_container_width=True,
            disabled=collection_locked,
            on_click=close_category_dialog_state,
        )
    else:
        selected_history_label = ""
        load_history = False

    st.markdown(f"<div class='section-heading'>{escape(T['filters'])}</div>", unsafe_allow_html=True)
    filter_top = st.columns(3)
    with filter_top[0]:
        min_price_raw, max_price_raw = render_range_filter(
            T["price"], "filter_price", "24.99", "", money=True, disabled=collection_locked
        )
    with filter_top[1]:
        min_reviews_raw, max_reviews_raw = render_range_filter(
            T["reviews"], "filter_reviews", "", "300", disabled=collection_locked
        )
    with filter_top[2]:
        min_bought_raw, max_bought_raw = render_range_filter(
            T["monthly_sales"], "filter_monthly_sales", "100", "", disabled=collection_locked
        )

    filter_bottom = st.columns([1, 1, 1], vertical_alignment="bottom")
    with filter_bottom[0]:
        min_child_sales_raw, max_child_sales_raw = render_range_filter(
            T["child_sales"], "filter_child_sales", disabled=collection_locked
        )
    with filter_bottom[1]:
        min_bsr_raw, max_bsr_raw = render_range_filter(
            T["bsr"], "filter_bsr", disabled=collection_locked
        )
    with filter_bottom[2]:
        st.markdown(f"<div class='filter-label'>{escape(T['launched_at'])} <span>?</span></div>", unsafe_allow_html=True)
        launch_options = ["不限", "近30天", "近60天", "近3个月", "近半年", "近1年", "近2年", "近1~2年"] if UI_LANG == "中文" else ["Any", "Last 30 days", "Last 60 days", "Last 3 months", "Last 6 months", "Last year", "Last 2 years", "1-2 years"]
        launch_window = st.selectbox(
            "上架时间",
            launch_options,
            key="filter_launch_window",
            on_change=request_filter_auto_apply,
            label_visibility="collapsed",
            disabled=collection_locked,
        )
    current_filters = {
        "min_price": parse_filter_number(min_price_raw, None),
        "max_price": parse_filter_number(max_price_raw, None),
        "min_reviews": parse_filter_number(min_reviews_raw, None, as_int=True),
        "max_reviews": parse_filter_number(max_reviews_raw, None, as_int=True),
        "min_bought": parse_filter_number(min_bought_raw, None, as_int=True),
        "max_bought": parse_filter_number(max_bought_raw, None, as_int=True),
        "min_child_sales": parse_filter_number(min_child_sales_raw, None, as_int=True),
        "max_child_sales": parse_filter_number(max_child_sales_raw, None, as_int=True),
        "min_bsr": parse_filter_number(min_bsr_raw, None, as_int=True),
        "max_bsr": parse_filter_number(max_bsr_raw, None, as_int=True),
        "launch_window": launch_window,
    }
    st.caption(build_filter_summary(current_filters))

    if st.session_state.filter_auto_apply_requested and st.session_state.raw_products:
        apply_filters_to_raw_pool(current_filters)
        st.session_state.filter_auto_apply_requested = False
        log(f"Auto-applied filters after filter input change: kept {len(st.session_state.products)}/{len(st.session_state.raw_products)}.")
    elif st.session_state.filter_auto_apply_requested:
        st.session_state.filter_auto_apply_requested = False

    st.info(build_collection_plan_text(selected_paths, custom_url, batch_category_collect, mapped_seed_urls))
    st.markdown("<div class='collection-action-toolbar'></div>", unsafe_allow_html=True)
    action_cols = st.columns([1.05, 1.05, 1.05, 1.05, 1.2], vertical_alignment="center")
    seller_cache_can_run = data_source != "卖家精灵插件" or chrome_ready
    run_clicked = action_cols[0].button(
        T["run"],
        key="run_collection_button",
        type="primary",
        use_container_width=True,
        disabled=not seller_cache_can_run or collection_locked,
        on_click=prepare_collection_run,
    )
    run = bool(run_clicked or st.session_state.collection_start_requested)
    st.session_state.collection_start_requested = False
    with action_cols[1]:
        render_stop_collection_button()
    apply_filter = action_cols[2].button(
        "应用筛选",
        key="apply_filter_button",
        use_container_width=True,
        disabled=collection_locked or not st.session_state.raw_products,
    )
    clear_filters = action_cols[3].button(
        "清空筛选",
        key="clear_filters_button",
        use_container_width=True,
        on_click=reset_filter_widgets,
        disabled=collection_locked,
    )
    load_last_raw = action_cols[4].button(
        "载入最近采集池",
        key="load_last_raw_button",
        use_container_width=True,
        disabled=collection_locked,
        on_click=close_category_dialog_state,
    )
    collection_total_placeholder = st.empty()
    update_collection_total_status(collection_total_placeholder)
    failed_seed_details = st.session_state.get("collection_failed_seed_details", [])
    if failed_seed_details:
        failure_lines = [
            f"{index}. {item.get('label', '未知入口')}：{item.get('error', '未知错误')}"
            for index, item in enumerate(failed_seed_details, start=1)
        ]
        with st.expander(
            f"本轮失败入口及原因（{len(failed_seed_details)} 个）",
            expanded=True,
        ):
            st.warning("\n\n".join(failure_lines))
    warning_seed_details = st.session_state.get("collection_warning_seed_details", [])
    if warning_seed_details:
        warning_lines = [
            f"{index}. {item.get('label', '未知入口')}：{item.get('warning', '产品数量偏少')}"
            for index, item in enumerate(warning_seed_details, start=1)
        ]
        with st.expander(
            f"本轮数量偏少警告（{len(warning_seed_details)} 个，不计为失败）",
            expanded=False,
        ):
            st.info("\n\n".join(warning_lines))

    raw_count = len(st.session_state.raw_products)
    filtered_count = len(st.session_state.products)
    if raw_count:
        st.caption(
            f"原始采集池：{raw_count} 条｜当前筛选保留：{filtered_count} 条｜筛掉：{raw_count - filtered_count} 条。"
            "如果只是调整筛选条件，点击“应用筛选”即可，不需要重新采集。"
        )
    if st.session_state.last_raw_products_message:
        st.info(st.session_state.last_raw_products_message)

if clear_filters and st.session_state.raw_products:
    apply_filters_to_raw_pool(current_filters)
    log("Filter widgets reset and filters re-applied to raw product pool.")
    st.rerun()

if load_last_raw:
    st.session_state.show_category_dialog = False
    loaded_products, message = load_raw_products()
    st.session_state.raw_products = loaded_products
    st.session_state.last_raw_products_message = message
    if loaded_products:
        apply_filters_to_raw_pool(current_filters)
        log(f"Loaded raw product pool from disk: {len(loaded_products)} products.")
    else:
        st.session_state.products = []
        st.session_state.last_collection_summary = ""
    st.rerun()

if load_history:
    st.session_state.show_category_dialog = False
    selected_history_path = dict(history_options).get(selected_history_label)
    if selected_history_path:
        loaded_products, payload, error = load_raw_products_payload(selected_history_path)
        if error:
            st.session_state.raw_products = []
            st.session_state.products = []
            st.session_state.last_collection_summary = ""
            st.session_state.last_raw_products_message = error
        else:
            st.session_state.raw_products = loaded_products
            apply_filters_to_raw_pool(current_filters)
            saved_at = payload.get("saved_at", "-")
            label = payload.get("label") or "未命名采集"
            st.session_state.last_raw_products_message = (
                f"已载入历史原始采集池：{label}，{len(loaded_products)} 条，保存时间：{saved_at}。"
                "载入后已按当前筛选条件重新计算。"
            )
            log(f"Loaded historical raw product pool: {len(loaded_products)} products from {selected_history_path}.")
    st.rerun()

if apply_filter:
    apply_filters_to_raw_pool(current_filters)
    log(f"Re-applied filters to raw product pool: kept {len(st.session_state.products)}/{len(st.session_state.raw_products)}.")
    st.rerun()

if run:
    reset_collection_run_messages()
    update_collection_total_status(collection_total_placeholder)
    mark_collection_running()
    filters = current_filters
    st.session_state.collection_in_progress = True
    active_progress_bar = None
    try:
        collected_products = []
        if data_source == "卖家精灵插件":
            selected_seed_paths = compact_category_paths(selected_paths)
            should_batch_category_collect = (
                batch_category_collect
                or len(selected_seed_paths) > 1
                or selection_contains_parent_category(selected_seed_paths)
            )
            if should_batch_category_collect:
                seed_urls = resolve_category_seed_urls(selected_seed_paths, custom_url)
                target_url = seed_urls[0][1] if len(seed_urls) == 1 else ""
                collection_label = "；".join(selected_paths[:3]) if selected_paths else "自定义链接"
                if len(selected_paths) > 3:
                    collection_label += f" +{len(selected_paths) - 3}"
                log("Start batch category collection from category selection.")
                batch_bar = st.progress(0, text="正在准备大类批量采集...")
                active_progress_bar = batch_bar
                collected_products = collect_sellersprite_batch_from_seeds(
                    seed_urls,
                    list_type,
                    filters,
                    batch_bar,
                    total_status_placeholder=collection_total_placeholder,
                )
            else:
                target_label, target_url = resolve_primary_collection_url(selected_paths, custom_url)
                target_url = amazon_url_for_list_type(target_url, list_type)
                collection_label = "；".join(selected_paths[:3]) if selected_paths else target_label
                if len(selected_paths) > 3:
                    collection_label += f" +{len(selected_paths) - 3}"
                st.session_state.collection_total_seed_count = 1
                st.session_state.collection_completed_seed_count = 0
                st.session_state.collection_total_raw_count = 0
                st.session_state.collection_current_seed_index = 1
                st.session_state.collection_current_seed_label = target_label
                update_collection_total_status(collection_total_placeholder)
                if selected_paths and (len(selected_paths) > 1 or any(find_exact_category_url(path) for path in selected_paths)):
                    st.session_state.last_category_mapping_message = (
                        (st.session_state.last_category_mapping_message + " " if st.session_state.last_category_mapping_message else "")
                        + "当前未勾选“大类批量采集”，本次只采集一个榜单入口页。"
                    )
                log(f"Open Amazon page and wait for SellerSprite plugin data: {target_label}.")
                refresh_bar = st.progress(0, text="正在打开 Amazon，并等待卖家精灵插件加载...")
                active_progress_bar = refresh_bar

                def update_run_refresh_progress(percent: int, message: str):
                    st.session_state.collection_total_raw_count = len(
                        st.session_state.get("collection_staged_raw_products", [])
                    )
                    update_collection_total_status(collection_total_placeholder)
                    refresh_bar.progress(percent, text=message)

                collected_products, refresh_results, quality_ok, quality_message = collect_sellersprite_entry_with_quality_retry(
                    target_url,
                    list_type,
                    filters,
                    progress=update_run_refresh_progress,
                    page_count=2,
                    progress_label=target_label,
                    category_path=target_label,
                )
                if not quality_ok and is_sellersprite_load_failure(quality_message):
                    raise SellerSpriteLoadTimeout(
                        f"当前入口卖家精灵加载超时：{quality_message}",
                        collected_products,
                        len(refresh_results),
                    )
                st.session_state.collection_completed_seed_count = 1
                st.session_state.collection_current_seed_index = 0
                st.session_state.collection_current_seed_label = ""
                st.session_state.collection_total_raw_count = len(collected_products)
                update_collection_total_status(collection_total_placeholder)
                completed_pages = completed_collection_page_count(refresh_results)
                total_product_count = sum(result.product_count for result in refresh_results)
                total_hydrated_count = sum(result.hydrated_count for result in refresh_results)
                total_image_count = sum(result.image_count for result in refresh_results)
                log(
                    "SellerSprite plugin data refresh: "
                    f"{len(collected_products)} unique products, "
                    f"{total_product_count} page products, {total_hydrated_count} hydrated, "
                    f"{total_image_count} images."
                )
                st.session_state.last_cache_refresh_message = (
                    f"本入口采集完成：读取 {completed_pages} 页；页面识别 {total_product_count} 条；"
                    f"原始去重 {len(collected_products)} 条；卖家精灵字段完整 {total_hydrated_count} 条（销量/FBA/销售额等）。"
                    f"质量判断：{quality_message}"
                )
                if not quality_ok:
                    st.session_state.last_cache_refresh_message += " 这次疑似漏采，建议保持采集 Chrome 前台可见后重试。"
                    if quality_message != EMPTY_NEW_RELEASES_MESSAGE:
                        st.session_state.collection_failed_seed_count = 1
                        st.session_state.collection_failed_seed_details = [{
                            "label": target_label,
                            "url": target_url,
                            "error": quality_message,
                        }]
                if is_collection_quality_warning(quality_message):
                    st.session_state.collection_warning_seed_count = 1
                    st.session_state.collection_warning_seed_details = [{
                        "label": target_label,
                        "url": target_url,
                        "warning": quality_message,
                    }]
                save_collection_report(
                    [{
                        "label": target_label,
                        "url": target_url,
                        "added": len(collected_products),
                        "raw": len(collected_products),
                        "quality_ok": quality_ok,
                        "quality_message": quality_message,
                        "pages": completed_pages,
                        "failed": not quality_ok and quality_message != EMPTY_NEW_RELEASES_MESSAGE,
                        "warning": is_collection_quality_warning(quality_message),
                        "empty": quality_message == EMPTY_NEW_RELEASES_MESSAGE,
                        "error": "" if quality_ok else quality_message,
                    }],
                    len(collected_products),
                    "completed",
                    [(target_label, target_url)],
                )
                log(f"SellerSprite plugin collection finished. Parsed {len(collected_products)} unique products from {completed_pages} pages.")
        for product in collected_products:
            product.selected = False
        if collected_products:
            st.session_state.raw_products = collected_products
            apply_filters_to_raw_pool(filters)
            save_raw_products(collected_products, collection_label, target_url)
            st.session_state.last_raw_products_message = (
                f"已保存原始采集池：{len(collected_products)} 条。最近 5 次采集会保留在本地。"
            )
        else:
            st.session_state.last_collection_summary = "本次没有解析到产品。请检查 Amazon 页面是否正常打开、卖家精灵插件是否已加载。"
            if st.session_state.raw_products:
                st.session_state.last_raw_products_message = (
                    f"本次未解析到新产品，已保留当前原始采集池：{len(st.session_state.raw_products)} 条。"
                )
            else:
                st.session_state.last_raw_products_message = ""
        log(
            "Filters applied: "
            f"price {filters['min_price']}-{filters['max_price']}, "
            f"reviews {filters['min_reviews'] or 0}-{filters['max_reviews']}, "
            f"monthly sales {filters['min_bought']}-{filters['max_bought'] or 'any'}, "
            f"child sales {filters['min_child_sales'] or 0}-{filters['max_child_sales'] or 'any'}, "
            f"BSR {filters['min_bsr'] or 0}-{filters['max_bsr'] or 'any'}, "
            f"launch {filters['launch_window']}. "
            f"Kept {len(st.session_state.products)}/{len(collected_products)}, "
            f"removed {len(collected_products) - len(st.session_state.products)}."
        )
    except SellerSpriteLoadTimeout as exc:
        if not collected_products:
            collected_products = st.session_state.get("collection_staged_raw_products", [])
        timeout_label = target_label if "target_label" in locals() else (
            collection_label if "collection_label" in locals() else "当前入口"
        )
        timeout_url = target_url if "target_url" in locals() else ""
        is_batch_timeout = bool(
            "should_batch_category_collect" in locals()
            and should_batch_category_collect
        )
        if not is_batch_timeout:
            st.session_state.collection_failed_seed_count += 1
            st.session_state.collection_failed_seed_details.append({
                "label": timeout_label,
                "url": timeout_url,
                "error": str(exc),
            })
            save_collection_report(
                [{
                    "label": timeout_label,
                    "url": timeout_url,
                    "added": len(collected_products),
                    "raw": len(collected_products),
                    "quality_ok": False,
                    "quality_message": str(exc),
                    "pages": exc.page_count,
                    "failed": True,
                    "warning": False,
                    "empty": False,
                    "error": str(exc),
                }],
                len(collected_products),
                "paused",
                [(timeout_label, timeout_url)],
            )
        if collected_products:
            for product in collected_products:
                product.selected = False
            st.session_state.raw_products = collected_products
            apply_filters_to_raw_pool(filters)
            save_raw_products(
                collected_products,
                collection_label if "collection_label" in locals() else "卖家精灵加载超时",
                timeout_url,
            )
            st.session_state.last_raw_products_message = (
                f"卖家精灵加载超时，整轮采集已暂停；已保存当前原始采集池：{len(collected_products)} 条。"
                "请确认插件在当前 Amazon 页面加载正常后再重新开始。"
            )
        st.session_state.last_collection_summary = str(exc)
        log(f"Collection paused after SellerSprite load timeout. Kept {len(collected_products)} products.")
        st.warning(str(exc))
    except CollectionStopped as exc:
        if not collected_products:
            collected_products = st.session_state.get("collection_staged_raw_products", [])
        if collected_products:
            for product in collected_products:
                product.selected = False
            st.session_state.raw_products = collected_products
            apply_filters_to_raw_pool(filters)
            save_raw_products(collected_products, collection_label if "collection_label" in locals() else "手动停止采集", target_url if "target_url" in locals() else "")
            st.session_state.last_raw_products_message = (
                f"采集已停止，已保存当前原始采集池：{len(collected_products)} 条。"
                "页面列表已按当前筛选条件重新计算；可以继续查看/导出，也可以重新开始采集。"
            )
        st.session_state.last_collection_summary = str(exc)
        log(f"Collection stopped by user. Kept {len(collected_products)} products.")
        st.warning(str(exc))
    except Exception as exc:
        st.session_state.last_collection_summary = f"采集失败：{exc}"
        staged_products = st.session_state.get("collection_staged_raw_products", [])
        if staged_products:
            for product in staged_products:
                product.selected = False
            st.session_state.raw_products = staged_products
            apply_filters_to_raw_pool(filters)
            save_raw_products(
                staged_products,
                collection_label if "collection_label" in locals() else "部分采集失败",
                target_url if "target_url" in locals() else "",
            )
            st.session_state.last_raw_products_message = (
                f"采集过程中发生异常，但已保存当前完成的原始产品：{len(staged_products)} 条。"
                "可以继续应用筛选、查看或导出；需要完整数据时再重新采集。"
            )
        elif st.session_state.raw_products:
            st.session_state.last_raw_products_message = (
                f"采集失败，已保留当前原始采集池：{len(st.session_state.raw_products)} 条。"
            )
        log(f"{data_source} collection failed: {exc}.")
        st.warning(f"{data_source} 实时采集失败：{exc}。请检查采集 Chrome、Amazon 登录、卖家精灵插件或类目链接。")
    finally:
        clear_progress_bar(active_progress_bar)
        st.session_state.collection_current_seed_index = 0
        st.session_state.collection_current_seed_label = ""
        st.session_state.collection_in_progress = False
        clear_collection_running_flag()
        clear_stop_collection_flag()
    st.rerun()

products = st.session_state.products
sync_product_annotations(products)
sync_product_selection_from_widgets(products)
current_page_size = normalize_page_size(st.session_state.get("result_page_size"))
st.session_state.result_page_size = current_page_size
current_result_page = clamp_page(st.session_state.get("result_current_page", 1), len(products), current_page_size)
st.session_state.result_current_page = current_result_page
st.session_state.result_page_jump = current_result_page
current_page_products = page_slice(products, current_result_page, current_page_size)
current_page_display_start = page_start_index(len(products), current_result_page, current_page_size)
if st.session_state.get("result_view_mode") not in ["列表", "平铺"]:
    st.session_state.result_view_mode = "列表"

if st.session_state.last_collection_summary:
    if products:
        st.success(st.session_state.last_collection_summary)
    elif st.session_state.raw_products:
        st.info(st.session_state.last_collection_summary)
    else:
        st.warning(st.session_state.last_collection_summary)

summary_cols = st.columns(6)
summary_cols[0].metric("产品数" if UI_LANG == "中文" else "Products", len(products))
summary_cols[1].metric("已勾选" if UI_LANG == "中文" else "Selected", sum(1 for p in products if p.selected))
summary_cols[2].metric("A 级", sum(1 for p in products if p.potential_level == "A"))
summary_cols[3].metric("B 级", sum(1 for p in products if p.potential_level == "B"))
summary_cols[4].metric("风险" if UI_LANG == "中文" else "Risk", sum(1 for p in products if p.potential_level == "Risk"))
summary_cols[5].metric("平均分" if UI_LANG == "中文" else "Avg Score", round(sum(p.potential_score for p in products) / len(products), 1) if products else 0)

if products:
    render_result_pagination_controls(len(products), "top")

favorites_tab_label = "收藏" if UI_LANG == "中文" else "Favorites"
tab_cards, tab_table, tab_favorites, tab_log = st.tabs([T["cards"], T["table"], favorites_tab_label, T["log"]])

with tab_cards:
    if not products:
        st.info(empty_products_message())
    else:
        selected_products = [p for p in products if p.selected]
        st.markdown("<span class='cards-toolbar-anchor'></span>", unsafe_allow_html=True)
        st.markdown("<span class='cards-toolbar-actions-anchor'></span>", unsafe_allow_html=True)
        actions_toolbar = st.columns([0.24, 0.78, 0.86, 0.7, 0.82, 0.8, 0.8, 3.0], vertical_alignment="center")
        all_selected = bool(products) and len(selected_products) == len(products)
        current_page_all_selected = bool(current_page_products) and all(product.selected for product in current_page_products)
        select_summary = (
            f"已全选 <strong>{len(selected_products)}</strong> 条"
            if all_selected
            else f"已勾选 <strong>{len(selected_products)}</strong> / {len(products)} 条"
        )
        st.session_state["select_current_page_products"] = current_page_all_selected
        actions_toolbar[0].checkbox(
            "全选当前页",
            key="select_current_page_products",
            label_visibility="collapsed",
            on_change=handle_select_current_page_products_change,
            args=(current_page_products,),
        )
        selected_products = [p for p in products if p.selected]
        actions_toolbar[1].markdown(f"<div class='toolbar-meta'>{select_summary}</div>", unsafe_allow_html=True)
        actions_toolbar[2].button(
            "全选全部结果",
            use_container_width=True,
            disabled=all_selected,
            on_click=select_all_filtered_products,
            args=(products, current_page_products),
        )
        if actions_toolbar[3].button("复制ASIN", use_container_width=True, disabled=not selected_products):
            log(f"Copied {len(selected_products)} ASIN values.")
        export_scope = actions_toolbar[4].selectbox(
            "导出范围",
            ["导出已勾选", "导出当前页", "导出全部结果"],
            key="result_export_scope",
            label_visibility="collapsed",
        )
        export_products = export_products_for_scope(products, current_page_products, export_scope)
        render_lazy_export_button(
            actions_toolbar[5],
            "生成导出",
            export_products,
            "amazon_selection_export.xlsx",
            disabled=not export_products,
            use_container_width=True,
            key="cards_scope_xlsx",
        )
        render_lazy_export_button(
            actions_toolbar[6],
            "生成明细",
            products,
            "amazon_selection_all.xlsx",
            use_container_width=True,
            key="cards_all_xlsx",
        )
        st.markdown("<span class='cards-toolbar-controls-anchor'></span>", unsafe_allow_html=True)
        st.markdown("<span class='cards-toolbar-sort-anchor'></span>", unsafe_allow_html=True)
        controls_toolbar = st.columns([1.0, 0.64, 0.76, 0.68, 0.72, 3.2], vertical_alignment="center")
        controls_toolbar[0].markdown(f"<div class='toolbar-meta'>搜索结果数：<strong>{len(products):,}</strong></div>", unsafe_allow_html=True)
        result_view_mode = st.session_state.get("result_view_mode", "列表")
        with controls_toolbar[1]:
            st.markdown("<span class='result-view-toggle-anchor'></span>", unsafe_allow_html=True)
            view_cols = st.columns(2, gap="small")
            view_cols[0].button(
                "列表",
                key="result_view_list_button",
                type="primary" if result_view_mode == "列表" else "secondary",
                use_container_width=True,
                on_click=set_result_view_mode,
                args=("列表",),
            )
            view_cols[1].button(
                "平铺",
                key="result_view_tile_button",
                type="primary" if result_view_mode == "平铺" else "secondary",
                use_container_width=True,
                on_click=set_result_view_mode,
                args=("平铺",),
            )
            result_view_mode = st.session_state.get("result_view_mode", "列表")
        controls_toolbar[2].selectbox("排序字段", ["月销量", "评分", "价格", "上架时间"], label_visibility="collapsed")
        controls_toolbar[3].selectbox("排序", ["降序", "升序"], label_visibility="collapsed")
        controls_toolbar[4].button("应用排序", type="primary", use_container_width=True)
        st.markdown("<div class='toolbar-spacer'></div>", unsafe_allow_html=True)
        if result_view_mode == "平铺":
            render_tile_cards(current_page_products)
        else:
            render_cards(current_page_products, current_page_display_start)
            render_list_favorite_portal()
        render_result_pagination_controls(len(products), "cards_bottom")
        render_clipboard_bridge()

with tab_table:
    if products:
        st.dataframe(
            table_rows(current_page_products, current_page_display_start),
            use_container_width=True,
            hide_index=True,
        )
        export_scope = st.session_state.get("result_export_scope", "导出已勾选")
        export_products = export_products_for_scope(products, current_page_products, export_scope)
        render_lazy_export_button(
            st,
            f"生成{export_scope} Excel" if UI_LANG == "中文" else "Prepare Excel",
            export_products,
            "amazon_selection_export.xlsx",
            disabled=not export_products,
            key="table_scope_xlsx",
        )
        render_lazy_export_button(
            st,
            "生成全部筛选结果 Excel" if UI_LANG == "中文" else "Prepare all filtered Excel",
            products,
            "amazon_selection_all.xlsx",
            key="table_all_xlsx",
        )
        render_lazy_export_button(
            st,
            f"生成{export_scope} CSV" if UI_LANG == "中文" else "Prepare CSV",
            export_products,
            "amazon_selection_export.csv",
            file_format="csv",
            disabled=not export_products,
            key="table_scope_csv",
        )
        render_lazy_export_button(
            st,
            "生成全部筛选结果 CSV" if UI_LANG == "中文" else "Prepare all filtered CSV",
            products,
            "amazon_selection_all.csv",
            file_format="csv",
            key="table_all_csv",
        )
        render_result_pagination_controls(len(products), "table_bottom")
    else:
        st.info(empty_products_message())

with tab_favorites:
    render_favorites_panel()

with tab_log:
    st.caption("操作日志：用于回看载入、筛选、采集和停止等动作。")
    st.code("\n".join(readable_log_line(line) for line in st.session_state.run_log[-80:]))





